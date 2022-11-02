###########################################################################################################
# 稼働設定：解像度 1920*1080 表示スケール125%
###########################################################################################################
# モジュールインポート
import pyautogui as pg
import time
import MJSOpen

# pandasインポート
import pandas as pd

# 配列計算関数numpyインポート
import numpy as np

# osインポート
import os

# datetimeインポート

# 例外処理判定の為のtracebackインポート
import traceback

# pandas(pd)で関与先データCSVを取得
import pyautogui
import pyperclip  # クリップボードへのコピーで使用
import ExcelFileAction as EFA

import datetime
import openpyxl


import WarekiHenkan as WH
import RPA_Function as RPA

import Sub_GenkasyoukyakuUpdate as GenkasyoukyakuUpdate
import Sub_HoujinzeiUpdate as HoujinzeiUpdate
import Sub_HouteiUpdate as HouteiUpdate
import Sub_IkkatuUpDate as IkkatuUpDate
import Sub_KaikeiUpDate as KaikeiUpDate
import Sub_KessanUpDate as KessanUpDate

import Sub_NencyouUpdate as NencyouUpdate
import Sub_SyotokuzeiUpdate as SyotokuzeiUpdate
import Sub_ZaisanUpdate as ZaisanUpdate

import Control

# logger設定------------------------------------------------------------------------------------------------------------
import logging.config

logging.config.fileConfig(r"LogConf\loggingMJSSysUp.conf")
logger = logging.getLogger(__name__)
LURL = r"\\NAS-SV\B_監査etc\B2_電子ﾌｧｲﾙ\RPA_ミロクシステム次年更新\一括更新申請\MJSLog\初回起動_"
# ----------------------------------------------------------------------------------------------------------------------
# #######################################################################################################################
# 一括更新処理の該当年度はThisYearKey.pngなので、年度が変わったらスクリーンショットしなおす事
# #######################################################################################################################

# ------------------------------------------------------------------------------------------------
class Job:
    """
    処理全体(class)
    """

    def __init__(self, **kw):
        log_out("Jobクラス読込開始")
        # 自分のDir(str)
        self.dir = RPA.My_Dir("MJS_System_NextCreate")
        # 画像のDir(str)
        self.Img_dir = self.dir + r"\\img"
        # コントロール(class)
        self.control = Control.control()
        # 当年(int)
        self.Start_Year = WH.Wareki.from_ad(datetime.datetime.today().year).year
        # RPA用画像フォルダの作成
        self.FolURL = os.getcwd().replace("\\", "/")  # 先
        self.TFolURL = RPA.My_Dir("MJS_System_NextCreate")  # 先
        self.imgdir_url = self.TFolURL + r"\\img"  # 先
        self.XLSDir = r"\\NAS-SV\B_監査etc\B2_電子ﾌｧｲﾙ\RPA_ミロクシステム次年更新\一括更新申請"
        self.first_csv = self.XLSDir + r"\MJSLog\MJSSysUpLog.txt"  # 処理状況CSVのURL
        self.BatUrl = (
            os.getcwd() + r"\\bat\\AWADriverOpen.bat"
        )  # 4724ポート指定でappiumサーバー起動バッチを開く
        self.driver = MJSOpen.MainFlow(
            self.BatUrl, self.FolURL, self.Img_dir
        )  # MJSを起動しログイン後インスタンス化
        log_out("Jobクラス読込終了")

    def MainFlow(self, Exc):
        """
        概要: メイン処理
        @param FolURL : ミロク起動関数のフォルダ(str)
        @param TFolURL : このpyファイルのフォルダ(str)
        @param Exc : Excel指示シート(obj)
        @return : bool
        """
        try:
            log_out("xlsxをDataFrameに")
            open(LURL, "w").close()
            for Exc.Title in Exc.input_sheet_name:
                # DataFrameとしてsheetのデータ読込み
                if "更新申請" in Exc.Title:
                    Exc.Read_sheet(Exc.Title, self.first_csv)
                    MainStarter(
                        self,
                        Exc,
                    )  # データ送信画面までの関数
                    print("")
        except Exception as e:
            log_out(e)

    def KomonUpdate(self, ExRow):
        """
        MJSの顧問先情報更新
        """
        log_out("_顧問先情報更新開始")
        # 関与先コード入力ボックスをクリック------------------------------------
        RPA.ImgClick(self.imgdir_url, r"\Komonsaki_Icon.png", 0.9, 10)
        while (
            pg.locateOnScreen(self.imgdir_url + r"\Komonsaki_Open.png", confidence=0.9)
            is None
        ):
            time.sleep(1)

        p = pyautogui.locateOnScreen(
            self.imgdir_url + r"\Komonsaki_CodeTxt.png", confidence=0.9
        )
        x, y = pyautogui.center(p)
        pyautogui.click(x + 100, y)
        pg.press("delete")
        pyperclip.copy(str(ExRow["関与先番号"]))
        pg.hotkey("ctrl", "v")
        pg.press(["return", "return"])

        time.sleep(1)

        p = pyautogui.locateOnScreen(
            self.imgdir_url + r"\RensaouMeisyou.png", confidence=0.9
        )
        x, y = pyautogui.center(p)
        pyautogui.click(x + 100, y)
        pg.press("up")
        pg.press("down")
        pg.press("delete")
        pyperclip.copy(str(ExRow["関与先番号"]))
        pg.hotkey("ctrl", "v")
        pg.press(["return", "return"])

        pg.keyDown("alt")
        pg.press("u")
        pg.keyUp("alt")

        time.sleep(3)

        pg.keyDown("alt")
        pg.press("x")
        pg.keyUp("alt")

        while (
            pg.locateOnScreen(self.imgdir_url + r"\SyonaiKanri.png", confidence=0.9)
            is None
        ):
            time.sleep(1)

        pg.keyDown("alt")
        pg.press("f4")
        pg.keyUp("alt")
        time.sleep(1)
        log_out("_顧問先情報更新完了")


# ------------------------------------------------------------------------------------------------
class Sheet:
    """
    エクセルブック(class)
    """

    def __init__(self, XLSURL, **kw):
        log_out("_Excelブック読込開始")
        self.mybook_url = XLSURL
        self.sheet_title = ""
        Ex_file = EFA.XlsmRead(XLSURL)
        if Ex_file[0] is True:
            # エクセルブック
            self.book = Ex_file[1]
            # 全シート
            self.input_sheet_name = self.book.sheet_names
            # 全シート数
            self.num_sheet = len(self.input_sheet_name)
            log_out("_Excelブック読込終了")
        else:
            log_out("_Excelブック読込失敗")

    def Read_sheet(self, sheet_name, first_csv, **kw):
        log_out("_Excelシート読込開始")
        self.sheet_header = []
        ExSheet = ""
        NameSheet = ""
        ExSheet = self.book.parse(sheet_name, skiprows=0)
        NameSheet = self.book.parse("関与先一覧")
        print(ExSheet)
        # 初回読込時の保存--------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H-%M-%S")
        self.sheet_df = pd.DataFrame(ExSheet)
        self.name_df = pd.DataFrame(NameSheet)
        self.sheet_df.to_csv(
            first_csv + dt_s + ".csv",
            encoding="cp932",
            index=False,
        )
        # 列名整理--------------------------------
        self.sheet_column_count = self.sheet_df.shape[1]  # 列数
        for Ex in range(self.sheet_column_count):
            ExRow = ExSheet.iloc[0]  # 列名
            ExSecondRow = ExSheet.iloc[1]  # 列名2
            if ExRow[Ex] == ExRow[Ex]:  # nan判定
                # nanでない場合
                Txt = ExRow[Ex]
                if ExSecondRow[Ex] == ExSecondRow[Ex]:  # nan判定
                    # nanでない場合
                    self.sheet_header.append(ExRow[Ex] + "_" + ExSecondRow[Ex])
                else:
                    # nanの場合
                    self.sheet_header.append(ExRow[Ex])
            else:
                # nanの場合
                if ExSecondRow[Ex] == ExSecondRow[Ex]:  # nan判定
                    # nanでない場合
                    self.sheet_header.append(Txt + "_" + ExSecondRow[Ex])
                else:
                    # nanの場合
                    self.sheet_header.append(Txt)
        # データ整理--------------------------------
        # Df作成
        ExDf = pd.DataFrame(self.sheet_df.values[3:, :], columns=self.sheet_header)
        # Dfnan処理
        ExDf.dropna(how="all", inplace=True)
        print(ExDf)
        self.sheet_df = ExDf
        self.sheet_column_count = self.sheet_df.shape[1]  # 列数
        self.sheet_row_count = self.sheet_df.shape[0]  # 行数

        log_out("_Excelシート読込完了")

    def WriteExcel(self):
        """
        エクセルシート入力
        """
        dt_now = datetime.datetime.now()
        dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
        WriteEx = openpyxl.load_workbook(self.mybook_url, keep_vba=True)
        WriteExSheet = WriteEx[self.sheet_title]
        WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
        WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
        print("シート書き込み完了")
        print(WriteEx)
        WriteEx.save(XLSURL)
        WriteEx.close


# ------------------------------------------------------------------------------------------------
def log_out(txt):
    """
    logger出力
    """
    dt_s = datetime.datetime.now()
    dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
    logger.debug(dt_s + txt)


# ------------------------------------------------------------------------------------------------
def logcsv_out(txt):
    """
    loggercsv出力
    """
    dt_s = datetime.datetime.now()
    dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
    with open(LURL, "a") as f:
        print([dt_s, txt], file=f)


# ------------------------------------------------------------------------------------------------
def NameSearch(NameDF, Rno):
    try:
        NameDFColumn = np.array(NameDF.columns)
        NameDF = np.array(NameDF)
        NC = np.where(NameDFColumn == "コード")
        KC = np.where(NameDFColumn == "顧問先名称")
        NameDFIndex = NameDF[:, NC]
        NR = np.where(NameDFIndex == Rno)
        N_L = NameDF[NR, KC]
        N_L = (
            str(N_L[0])
            .replace("[", "")
            .replace("]", "")
            .replace("'", "")
            .replace('"', "")
            .replace("\u3000", "")
        )
        print(N_L)
        return N_L
    except:
        return "NameErr"


# ------------------------------------------------------------------------------------------------
def ChildFlow(Job, Exc):
    if "会計大将" == Exc.Title:
        # Log---------------------------------------------------------------------------------------
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + "_会計大将更新処理開始"
        )
        log_out(msg)
        logcsv_out(msg)
        # ------------------------------------------------------------------------------------------
        # 会計大将のアイコンを探す-------------------------------------------------
        ImgList = [r"\K_TaisyouIcon.png", r"\K_TaisyouIcon2.png"]
        ICFL = RPA.ImgCheckForList(Job.imgdir_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 会計大将のアイコンがあれば
            RPA.ImgClick(Job.imgdir_url, ICFL[1], 0.9, 10)  # 会計大将のアイコンをクリック
            SystemUp = KaikeiUpDate.KaikeiUpDate(Job, Exc)
            # Excel書き込み--------------------------------------------------
            if SystemUp[0] is True:
                Exc.WriteExcel()
                # Log---------------------------------------------------------------------------------------
                msg = (
                    "_関与先番号:"
                    + str(Exc.row_kanyo_no)
                    + ":"
                    + str(Exc.row_kanyo_name)
                    + "_会計大将更新処理終了"
                )
                log_out(msg)
                logcsv_out(msg)
                # ------------------------------------------------------------------------------------------
            else:
                if SystemUp[1] == "当年データ重複エラー":
                    # Log---------------------------------------------------------------------------------------
                    msg = (
                        "_関与先番号:"
                        + str(Exc.row_kanyo_no)
                        + ":"
                        + str(Exc.row_kanyo_name)
                        + "_会計大将当年データ重複エラー"
                    )
                    log_out(msg)
                    logcsv_out(msg)
                    # ------------------------------------------------------------------------------------------
                else:
                    # Log---------------------------------------------------------------------------------------
                    msg = (
                        "_関与先番号:"
                        + str(Exc.row_kanyo_no)
                        + ":"
                        + str(Exc.row_kanyo_name)
                        + "_会計大将更新処理エラー中断"
                    )
                    log_out(msg)
                    logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
    elif "決算内訳書" == Exc.Title:
        # Log---------------------------------------------------------------------------------------
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + "_決算内訳書更新処理開始"
        )
        log_out(msg)
        logcsv_out(msg)
        # ------------------------------------------------------------------------------------------
        SystemUp = KessanUpDate.KessanUpDate(Job, Exc)
        # Excel書き込み--------------------------------------------------
        if SystemUp[0] is True:
            Exc.WriteExcel()
            # Log---------------------------------------------------------------------------------------
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + "_決算内訳書更新処理終了"
            )
            log_out(msg)
            logcsv_out(msg)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "Noren":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "連動対象無エラー"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_決算内訳書更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "決算内訳書更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
    elif "減価償却" == Exc.Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_減価償却更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "減価償却更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = GenkasyoukyakuUpdate.GenkasyoukyakuUpdate(
            FolURL, TFolURL, ExRow, driver
        )
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_減価償却更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "減価償却更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "Noren":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "決算未確定更新"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
        elif SystemUp[1] == "当年データ重複エラー":
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_減価償却当年データ重複エラー"
            )
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "減価償却当年データ重複エラー"], file=f)
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "減価償却当年データ重複エラー"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
        else:
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_決算未確定減価償却更新処理終了"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "決算未確定_減価償却更新処理終了"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
        # ---------------------------------------------------------------
    elif "法人税申告書" == Exc.Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法人税申告書更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "法人税申告書更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = HoujinzeiUpdate.HoujinzeiUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法人税申告書更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "法人税申告書更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "要データ再計算":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "要データ再計算"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法人税申告書更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "法人税申告書更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "要申告指定":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "要申告指定"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法人税申告書申告指定無しの為中断"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "法人税申告書申告指定無しの為中断"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
    elif "所得税確定申告" == Exc.Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_所得税更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "所得税更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = SyotokuzeiUpdate.SyotokuzeiUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_所得税更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "所得税更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "当年データ重複エラー":
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_所得税当年データ重複エラー")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "所得税当年データ重複エラー"], file=f)
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "所得税当年データ重複エラー"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
        elif SystemUp[1] == "Nocalc":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "計算未処理更新"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_所得税更新処理計算未処理で終了"
            )
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "所得税更新処理計算未処理で終了"], file=f)
        elif SystemUp[0] is False:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "関与先無"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # -------------------------------------------
    elif "財産評価明細書" == Exc.Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_財産評価明細書更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "財産評価明細書更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = ZaisanUpdate.ZaisanUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_財産評価明細書更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "財産評価明細書更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "更新対象年度無し":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "更新対象年度無し"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_財産評価明細書更新処理終了_更新対象年度無し"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "財産評価明細書更新処理終了_更新対象年度無し"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
        # ---------------------------------------------------------------
    elif "年末調整" == Exc.Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_年末調整更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "年末調整更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = NencyouUpdate.NencyouUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_年末調整更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "年末調整更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "次年度あり":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_年末調整更新次年度あり処理終了"
            )
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "年末調整更新次年度あり処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        # ---------------------------------------------------------------
    elif "法定調書" == Exc.Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法定調書更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "法定調書更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = HouteiUpdate.HouteiUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[Title]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
        # ---------------------------------------------------------------
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法定調書更新処理終了")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "法定調書更新処理終了"], file=f)
        # ------------------------------------------------------------------------------------------
    else:
        print("NoSystem")


# ------------------------------------------------------------------------------------------------
def MainStarter(Job, Exc):
    try:
        for Ex in range(Exc.sheet_row_count):
            Exc.row_data = Exc.sheet_df.iloc[Ex]
            if Exc.row_data["関与先番号"] == Exc.row_data["関与先番号"]:  # nan判定
                # nanでない場合
                Exc.row_kanyo_no = Exc.row_data["関与先番号"]
                Exc.row_kanyo_name = NameSearch(Exc.name_df, Exc.row_kanyo_no)
                Job.KomonUpdate(Exc.row_data)  # 顧問先情報更新
                OpenSystem(Job, Exc)
                print("")
            else:
                # nanの場合
                print("nan")
        return True, ""
    except:
        return False, ""


# ------------------------------------------------------------------------------------------------
def OpenSystem(Job, Exc):
    try:
        Eh = 0
        for ExrcHeaderItem in Exc.sheet_header:
            if Eh < (len(Exc.sheet_header) - 1):
                if "_繰越対象" in ExrcHeaderItem:
                    SysN = ExrcHeaderItem.split("_")
                    Exc.Title = str(SysN[0])
                    if (
                        not Exc.row_data[Exc.Title + "_繰越対象"] == "-"
                        and str(Exc.row_data[Exc.Title + "_繰越対象"]) == "1"
                    ):
                        if "::" not in Exc.Title:
                            if (
                                Exc.row_data[Exc.Title + "_繰越対象"]
                                == Exc.row_data[Exc.Title + "_繰越対象"]
                            ):
                                # nanでない場合
                                if (
                                    Exc.row_data[Exc.Title + "_繰越処理日"]
                                    == Exc.row_data[Exc.Title + "_繰越処理日"]
                                ):
                                    # nanでない場合
                                    print(Exc.row_data[Exc.Title + "_繰越処理日"])
                                else:
                                    # nanの場合
                                    # Log--------------------------------------------
                                    msg = (
                                        "_関与先番号:"
                                        + str(Exc.row_kanyo_no)
                                        + ":"
                                        + str(Exc.row_kanyo_name)
                                        + "_メイン処理開始"
                                    )
                                    log_out(msg)
                                    logcsv_out(msg)
                                    # -----------------------------------------------
                                    ChildFlow(Job, Exc)
                            else:
                                # nanでない場合
                                if (
                                    Exc.row_data[Exc.Title + "_繰越処理日"]
                                    == Exc.row_data[Exc.Title + "_繰越処理日"]
                                ):
                                    # nanでない場合
                                    print("スタート")
                                else:
                                    # nanの場合
                                    # Log--------------------------------------------
                                    msg = (
                                        "_関与先番号:"
                                        + str(Exc.row_kanyo_no)
                                        + ":"
                                        + str(Exc.row_kanyo_name)
                                        + "_メイン処理開始"
                                    )
                                    log_out(msg)
                                    logcsv_out(msg)
                                    # -----------------------------------------------
                                    ChildFlow(Job, Exc)
        return True
    except:
        print("TEST")
        return False


if __name__ == "__main__":
    global Start_Year  # 当年
    j = Job()

    # Log--------------------------------------------
    dt_s = datetime.datetime.now()
    dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
    logger.debug(dt_s + "_MJSシステム更新開始")
    # -----------------------------------------------
    for fd_path, sb_folder, sb_file in os.walk(j.XLSDir):
        FDP = fd_path
        if not len(sb_folder) == 0:
            for sb_fileItem in sb_file:
                print(sb_fileItem)
                if (
                    "一括更新申請ミロク" in sb_fileItem
                    and not "一括更新申請ミロク(原本).xlsm" == sb_fileItem
                ):
                    XLSURL = FDP + r"\\" + sb_fileItem.replace("~", "").replace("$", "")
                    MoveXLSURL = (
                        FDP
                        + r"\\MJSLog\\"
                        + sb_fileItem.replace("~", "").replace("$", "")
                    )
                    os.rename(XLSURL, MoveXLSURL)
                    MoveXLSURL = (
                        FDP + r"\\" + sb_fileItem.replace("~", "").replace("$", "")
                    )
                    XLSURL = (
                        FDP
                        + r"\\MJSLog\\"
                        + sb_fileItem.replace("~", "").replace("$", "")
                    )
                    Ex_File = Sheet(XLSURL)
                    try:
                        j.MainFlow(Ex_File)
                    except:
                        traceback.print_exc()

                    os.rename(XLSURL, MoveXLSURL)
