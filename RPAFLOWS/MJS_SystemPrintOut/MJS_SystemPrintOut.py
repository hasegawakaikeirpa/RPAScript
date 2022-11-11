###########################################################################################################
# 稼働設定：解像度 1920*1080 表示スケール125%
###########################################################################################################
# モジュールインポート
import pyautogui as pg
import time
import pandas as pd
import numpy as np
import os
from threading import Thread, Lock
import traceback
import datetime
import openpyxl
import re

# pandas(pd)で関与先データCSVを取得
# import pyautogui
# import pyperclip  # クリップボードへのコピーで使用

# 自作モジュールインポート
import MJSOpen
import MJSSPOPDFMarge as PDFM
import ExcelFileAction as EFA
import WarekiHenkan as WH
import RPA_Function as RPA
import Sub_GenkasyoukyakuUpdate as GenkasyoukyakuUpdate
import Sub_HoujinzeiUpdate as HoujinzeiUpdate
import Sub_KaikeiUpDate as KaikeiUpDate
import Sub_KessanUpDate as KessanUpDate
import Sub_DensisinkokuUpDate as DensisinkokuUpDate

# logger設定------------------------------------------------------------------------------------------------------------
import logging.config

logging.config.fileConfig(r"LogConf\loggingMJSSysUp.conf")
logger = logging.getLogger(__name__)
LURL = (
    r"\\NAS-SV\B_監査etc\B2_電子ﾌｧｲﾙ\RPA_ミロクシステム次年更新\製本・電子ファイル印刷申請\MJSLog\MJSSysUpLog.txt"
)
open(LURL, "w").close()  # エクセル用実行ログをリセット
# ----------------------------------------------------------------------------------------------------------------------
# #######################################################################################################################
# 一括更新処理の該当年度はThisYearKey.pngなので、年度が変わったらスクリーンショットしなおす事
# #######################################################################################################################

# ------------------------------------------------------------------------------------------------
class Job:
    """
    処理全体(class)
    """

    def __init__(self, **kw):
        log_out("Jobクラス読込開始")
        # 自分のDir(str)
        self.dir = RPA.My_Dir("MJS_SystemPrintOut")
        # 画像のDir(str)
        self.Img_dir = self.dir + r"\\img"
        self.FolURL = os.getcwd().replace("\\", "/")  # 先
        # 当年(int)
        self.Start_Year = WH.Wareki.from_ad(datetime.datetime.today().year).year
        # RPA用画像フォルダの作成
        self.PrintOut_url = self.Img_dir + r"\\MJS_SystemPrintOut"  # 先
        self.All_url = self.PrintOut_url + r"\\All"  # 先
        self.NextCreate_url = self.Img_dir + r"\\MJS_SystemNextCreate"  # 先
        self.XLSDir = r"\\NAS-SV\B_監査etc\B2_電子ﾌｧｲﾙ\RPA_ミロクシステム次年更新\製本・電子ファイル印刷申請"
        self.SerchURL = r"\\NAS-SV\B_監査etc\B2_電子ﾌｧｲﾙ\03_法人決算"  # 先
        self.Log_dir = self.XLSDir + r"\MJSLog\BackUp"  # 処理状況CSVの保管場所
        self.filename = ""  # 処理状況CSVの元となるファイル名
        # self.BatUrl = (
        #     os.getcwd() + r"\\bat\\AWADriverOpen.bat"
        # )  # 4724ポート指定でappiumサーバー起動バッチを開く
        self.driver = MJSOpen.MainFlow(
            "self.BatUrl", self.FolURL, self.Img_dir
        )  # MJSを起動しログイン後インスタンス化
        log_out("Jobクラス読込終了")

    def MainFlow(self, Exc):

        """
        概要: プリントメイン処理
        @param FolURL : ミロク起動関数のフォルダ(str)
        @param TFolURL : このpyファイルのフォルダ(str)
        @param Exlsx : Excel指示シート(obj)
        @return : bool
        """
        try:
            log_out("xlsxをDataFrameに")
            for Exc.sheet_name in Exc.input_sheet_name:
                # DataFrameとしてsheetのデータ読込み
                if "印刷申請" in Exc.sheet_name:
                    Exc.Read_sheet(Exc.sheet_name, self.Log_dir + r"\\" + self.filename)
                    MainStarter(
                        self,
                        Exc,
                    )  # データ送信画面までの関数
                    print("")
        except Exception as e:
            log_out(e)


# ------------------------------------------------------------------------------------------------
class Sheet:
    """
    エクセルブック(class)
    """

    def __init__(self, XLSURL, **kw):
        log_out("_Excelブック読込開始")
        self.mybook_url = XLSURL
        Ex_file = EFA.XlsmRead(XLSURL)
        if Ex_file[0] is True:
            # エクセルブック
            self.book = Ex_file[1]
            # 全シート
            self.input_sheet_name = self.book.sheet_names
            # 全シート数
            self.num_sheet = len(self.input_sheet_name)
            log_out("_Excelブック読込終了")
        else:
            log_out("_Excelブック読込失敗")

    def Read_sheet(self, sheet_name, Log_csvurl, **kw):
        log_out("_Excelシート読込開始")
        self.sheet_header = []
        ExSheet = ""
        NameSheet = ""
        ExSheet = self.book.parse(sheet_name, skiprows=0)
        NameSheet = self.book.parse("関与先一覧")
        print(ExSheet)
        # 初回読込時の保存--------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H-%M-%S")
        self.sheet_df = pd.DataFrame(ExSheet)
        self.name_df = pd.DataFrame(NameSheet)
        self.sheet_df.to_csv(
            Log_csvurl + dt_s + ".csv",
            encoding="cp932",
            index=False,
        )
        # 列名整理--------------------------------
        self.sheet_column_count = self.sheet_df.shape[1]  # 列数
        for Ex in range(self.sheet_column_count):
            ExRow = ExSheet.iloc[0]  # 列名
            ExSecondRow = ExSheet.iloc[2]  # 列名2
            if ExRow[Ex] == ExRow[Ex]:  # nan判定
                # nanでない場合
                Txt = ExRow[Ex]
                if ExSecondRow[Ex] == ExSecondRow[Ex]:  # nan判定
                    # nanでない場合
                    self.sheet_header.append(ExRow[Ex] + "_" + ExSecondRow[Ex])
                else:
                    # nanの場合
                    self.sheet_header.append(ExRow[Ex])
            else:
                # nanの場合
                if ExSecondRow[Ex] == ExSecondRow[Ex]:  # nan判定
                    # nanでない場合
                    self.sheet_header.append(Txt + "_" + ExSecondRow[Ex])
                else:
                    # nanの場合
                    self.sheet_header.append(Txt)
        # データ整理--------------------------------
        self.sheet_PdfCell = self.sheet_header.index("保存先_URL")
        # Df作成
        ExDf = pd.DataFrame(self.sheet_df.values[3:, :], columns=self.sheet_header)
        # Dfnan処理
        ExDf.dropna(how="all", inplace=True)
        print(ExDf)
        self.sheet_df = ExDf
        self.sheet_column_count = self.sheet_df.shape[1]  # 列数
        self.sheet_row_count = self.sheet_df.shape[0]  # 行数

        log_out("_Excelシート読込完了")

    def WriteExcel(self, txt):
        """
        エクセルシート入力
        """
        dt_now = datetime.datetime.now()
        dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
        WriteEx = openpyxl.load_workbook(self.mybook_url, keep_vba=True)
        WriteExSheet = WriteEx[self.sheet_name]
        if txt == "○":
            WriteExSheet.cell(
                row=self.this_row_count + 5, column=self.this_col_count + 1
            ).value = dt_now
        else:
            WriteExSheet.cell(
                row=self.this_row_count + 5, column=self.this_col_count + 1
            ).value = txt
        print("シート書き込み完了")
        print(WriteEx)
        WriteEx.save(XLSURL)
        WriteEx.close

    def WritePDFURL(self, txt):
        """
        エクセルシート入力
        """
        dt_now = datetime.datetime.now()
        dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
        WriteEx = openpyxl.load_workbook(self.mybook_url, keep_vba=True)
        WriteExSheet = WriteEx[self.sheet_name]
        WriteExSheet.cell(
            row=self.this_row_count + 5, column=self.sheet_PdfCell + 1
        ).value = txt
        WriteExSheet.cell(
            row=self.this_row_count + 5, column=self.sheet_PdfCell + 1
        ).hyperlink = txt
        print("シート書き込み完了")
        print(WriteEx)
        WriteEx.save(XLSURL)
        WriteEx.close


# ------------------------------------------------------------------------------------------------
def log_out(txt):
    """
    logger出力
    """
    logger.debug(txt)


# ------------------------------------------------------------------------------------------------
def logcsv_out(txt):
    """
    loggercsv出力
    """
    dt_s = datetime.datetime.now()
    dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
    with open(LURL, "a") as f:
        print([dt_s, txt], file=f)


# ------------------------------------------------------------------------------------------------
def NameSearch(NameDF, Rno):
    """
    顧問先名称の検索
    """
    try:
        NameDFColumn = np.array(NameDF.columns)
        NameDF = np.array(NameDF)
        NC = np.where(NameDFColumn == "コード")
        KC = np.where(NameDFColumn == "顧問先名称")
        NameDFIndex = NameDF[:, NC]
        NR = np.where(NameDFIndex == Rno)
        N_L = NameDF[NR, KC]
        N_L = (
            str(N_L[0])
            .replace("[", "")
            .replace("]", "")
            .replace("'", "")
            .replace('"', "")
            .replace("\u3000", "")
        )
        print(N_L)
        return N_L
    except:
        return "NameErr"


# ------------------------------------------------------------------------------------------------------------------
def ChildFlow_sub(Job, Exc, txt):
    """
    シート列名に応じて処理分岐サブ
    """
    # Log---------------------------------------------------------------------------------------
    msg = "_関与先番号:" + str(Exc.row_kanyo_no) + ":" + str(Exc.row_kanyo_name) + txt + "開始"
    log_out(msg)
    logcsv_out(msg)

    if "会計大将" == Exc.Title:
        # 会計大将のアイコンを探す
        ImgList = [r"\K_TaisyouIcon.png", r"\K_TaisyouIcon2.png"]
        ICFL = RPA.ImgCheckForList(Job.All_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 会計大将のアイコンがあれば
            RPA.ImgClick(Job.All_url, ICFL[1], 0.9, 10)  # 会計大将のアイコンをクリック
            ret = KaikeiUpDate.KaikeiUpDate(Job, Exc)
            # Log
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + txt
                + "終了"
            )
            log_out(msg)
            logcsv_out(msg)
            return ret

    elif "決算内訳書" == Exc.Title:
        # 決算内訳書のアイコンを探す-------------------------------------------------
        ImgList = [r"\K_Uchiwake.png", r"\K_Uchiwake2.png"]
        ICFL = RPA.ImgCheckForList(Job.All_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 決算内訳書のアイコンがあれば
            RPA.ImgClick(Job.All_url, ICFL[1], 0.9, 10)  # 決算内訳書のアイコンをクリック
            ret = KessanUpDate.KessanUpDate(Job, Exc)
            # Log
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + txt
                + "終了"
            )
            log_out(msg)
            logcsv_out(msg)
            return ret

    elif "減価償却" == Exc.Title:

        # 減価償却のアイコンを探す-------------------------------------------------
        ImgList = [r"\G_Syoukyaku.png", r"\G_Syoukyaku2.png"]
        ICFL = RPA.ImgCheckForList(Job.All_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 減価償却のアイコンがあれば
            RPA.ImgClick(Job.All_url, ICFL[1], 0.9, 10)  # 減価償却のアイコンをクリック
            ret = GenkasyoukyakuUpdate.GenkasyoukyakuUpdate(Job, Exc)
            # Log
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + txt
                + "終了"
            )
            log_out(msg)
            logcsv_out(msg)
            return ret

    elif "法人税申告書" == Exc.Title:
        # 法人税のアイコンを探す-------------------------------------------------
        ImgList = [r"\Houjinzei.png", r"\Houjinzei2.png"]
        ICFL = RPA.ImgCheckForList(Job.All_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 法人税のアイコンがあれば
            RPA.ImgClick(Job.All_url, ICFL[1], 0.9, 10)  # 法人税のアイコンをクリック
            ret = HoujinzeiUpdate.HoujinzeiUpdate(Job, Exc)
            # Log
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + txt
                + "終了"
            )
            log_out(msg)
            logcsv_out(msg)
            return ret

    elif "電子申告" == Exc.Title:
        # 電子申告のアイコンを探す-------------------------------------------------
        ImgList = [r"\DensiIcon.png", r"\DensiIcon2.png"]
        ICFL = RPA.ImgCheckForList(Job.All_url, ImgList, 0.9, 10)
        if ICFL[0] is True:  # 電子申告のアイコンがあれば
            RPA.ImgClick(Job.All_url, ICFL[1], 0.9, 10)  # 電子申告のアイコンをクリック
            pg.keyDown("alt")
            pg.press("a")
            pg.keyUp("alt")
            ret = DensisinkokuUpDate.DensisinkokuUpDate(Job, Exc)
            # Log
            msg = (
                "_関与先番号:"
                + str(Exc.row_kanyo_no)
                + ":"
                + str(Exc.row_kanyo_name)
                + txt
                + "終了"
            )
            log_out(msg)
            logcsv_out(msg)
            return ret


# ------------------------------------------------------------------------------------------------------------------
def ChildFlow(Job, Exc):
    """
    シート列名に応じて処理分岐サブ
    """
    try:
        if "会計大将" == Exc.Title:
            Exc.Fname = Job.Img_dir + r"\PDF\\" + Exc.PN + ".pdf"
            # 処理
            SystemUp = ChildFlow_sub(Job, Exc, Exc.Title + Exc.PN + "_印刷処理")
            ChildFlow_err(Job, Exc, SystemUp)
            return True
        elif "決算内訳書" == Exc.Title:
            Exc.Fname = Job.Img_dir + r"\PDF\\" + Exc.PN + ".pdf"
            # 処理
            SystemUp = ChildFlow_sub(Job, Exc, Exc.Title + Exc.PN + "_印刷処理")
            ChildFlow_err(Job, Exc, SystemUp)
            return True
        elif "減価償却" == Exc.Title:
            Exc.Fname = Job.Img_dir + r"\PDF\\" + Exc.PN + ".pdf"
            # 処理
            SystemUp = ChildFlow_sub(Job, Exc, Exc.Title + Exc.PN + "_印刷処理")
            if SystemUp[0] is True:
                Exc.WriteExcel("○")  # シート書き込み
                # Log
                msg = (
                    "_関与先番号:"
                    + str(Exc.row_kanyo_no)
                    + ":"
                    + str(Exc.row_kanyo_name)
                    + Exc.Title
                    + Exc.PN
                    + "_印刷処理完了"
                )
                log_out(msg)
                logcsv_out(msg)
                return True
            elif SystemUp[0] == "年度なし":
                Exc.WriteExcel("年度なし")  # シート書き込み
                # Log
                msg = (
                    "_関与先番号:"
                    + str(Exc.row_kanyo_no)
                    + ":"
                    + str(Exc.row_kanyo_name)
                    + Exc.Title
                    + Exc.PN
                    + "_年度なし"
                )
                log_out(msg)
                logcsv_out(msg)
                return True
            elif SystemUp[1] == "物件無し":
                Exc.WriteExcel("物件無し")  # シート書き込み
                # Log
                msg = (
                    "_関与先番号:"
                    + str(Exc.row_kanyo_no)
                    + ":"
                    + str(Exc.row_kanyo_name)
                    + Exc.Title
                    + Exc.PN
                    + "_物件無し"
                )
                log_out(msg)
                logcsv_out(msg)
                return True
            elif SystemUp[0] is False:
                Exc.WriteExcel("資産無し")  # シート書き込み
                # Log
                msg = (
                    "_関与先番号:"
                    + str(Exc.row_kanyo_no)
                    + ":"
                    + str(Exc.row_kanyo_name)
                    + Exc.Title
                    + Exc.PN
                    + "_資産無し"
                )
                log_out(msg)
                logcsv_out(msg)
                return True
        elif "法人税申告書" == Exc.Title:
            Exc.Fname = Job.Img_dir + r"\PDF\\" + Exc.PN + ".pdf"
            # 処理
            SystemUp = ChildFlow_sub(Job, Exc, Exc.Title + Exc.PN + "_印刷処理")
            ChildFlow_err(Job, Exc, SystemUp)
            return True
        elif "電子申告" == Exc.Title:
            Exc.Fname = Job.Img_dir + r"\PDF\\" + Exc.PN + ".pdf"
            # 処理
            SystemUp = ChildFlow_sub(Job, Exc, Exc.Title + Exc.PN + "_印刷処理")
            ChildFlow_err(Job, Exc, SystemUp)
            return True
    except:
        return False


# ------------------------------------------------------------------------------------------------------------------
def ChildFlow_err(Job, Exc, SystemUp):
    """
    エラー処理
    """
    if SystemUp[0] is True:
        Exc.WriteExcel("○")  # シート書き込み
        # Log
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + Exc.Title
            + Exc.PN
            + "_印刷処理完了"
        )
        log_out(msg)
        logcsv_out(msg)
    elif SystemUp[1] == "要消費税基本情報登録":
        Exc.WriteExcel("要消費税基本情報登録")  # シート書き込み
        # Log
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + Exc.Title
            + Exc.PN
            + "_要消費税基本情報登録"
        )
        log_out(msg)
        logcsv_out(msg)
    elif SystemUp[1] == "出力履歴なし":
        Exc.WriteExcel("要消費税基本情報登録")  # シート書き込み
        # Log
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + Exc.Title
            + Exc.PN
            + "_出力履歴なし"
        )
        log_out(msg)
        logcsv_out(msg)
    elif SystemUp[1] == "決算順序未設定":
        Exc.WriteExcel("決算順序未設定")  # シート書き込み
        # Log
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + Exc.Title
            + Exc.PN
            + "_決算順序未設定"
        )
        log_out(msg)
        logcsv_out(msg)
    elif SystemUp[1] == "印刷様式未設定":
        Exc.WriteExcel("印刷様式未設定処理終了")  # シート書き込み
        # Log
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + Exc.Title
            + Exc.PN
            + "_印刷様式未設定処理終了"
        )
        log_out(msg)
        logcsv_out(msg)
    elif SystemUp[1] == "T_Err":
        Exc.WriteExcel("貸借バランスエラー")  # シート書き込み
        # Log
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + Exc.Title
            + Exc.PN
            + "_貸借バランスエラー"
        )
        log_out(msg)
        logcsv_out(msg)
    elif SystemUp[0] == "年度なし":
        Exc.WriteExcel("年度なし")  # シート書き込み
        # Log
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + Exc.Title
            + Exc.PN
            + "_年度なし"
        )
        log_out(msg)
        logcsv_out(msg)
    elif SystemUp[0] is False:
        Exc.WriteExcel("計算エラー")  # シート書き込み
        # Log
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + Exc.Title
            + Exc.PN
            + "_計算エラー"
        )
        log_out(msg)
        logcsv_out(msg)
    else:
        Exc.WriteExcel("失敗")  # シート書き込み
        # Log
        msg = (
            "_関与先番号:"
            + str(Exc.row_kanyo_no)
            + ":"
            + str(Exc.row_kanyo_name)
            + Exc.Title
            + Exc.PN
            + "_印刷処理失敗"
        )
        log_out(msg)
        logcsv_out(msg)


# ------------------------------------------------------------------------------------------------------------------
def OpenSystem(Job, Exc):
    """
    列ループ
    """
    try:
        Exc.this_col_count = 0
        # 列名からMJSシステムリストを作成------------------------------------------------
        Exc.SystemList = []
        Exc.CFFlag = False
        for Exc.this_col_name in Exc.sheet_header:
            # エクセルシートの値が1(str)なら
            if Exc.row_data[Exc.this_col_name] == "1":
                if Exc.this_col_count < (len(Exc.sheet_header) - 1):
                    if "_" in Exc.this_col_name:
                        SP = Exc.this_col_name.split("_")
                        Exc.Title = SP[0]  # MJS出力帳表名
                        Exc.PN = SP[1]  # MJS出力帳表名
                        Exc.SystemList.append(SP[1])
                        # Log---------------------------------------------------------------------------------------
                        msg = (
                            "_関与先番号:"
                            + str(Exc.row_kanyo_no)
                            + ":"
                            + str(Exc.row_kanyo_name)
                            + Exc.Title
                            + Exc.PN
                            + "_プリントメイン処理開始"
                        )
                        log_out(msg)
                        logcsv_out(msg)
                        # ------------------------------------------------------------------------------------------
                        if (
                            Exc.Title != "なし"
                            and Exc.Title != "決算フォルダ"
                            and Exc.Title != "繰越対象"
                            and "::" not in Exc.Title
                        ):
                            CF = ChildFlow(Job, Exc)
                            if CF is True:
                                Exc.CFFlag = True
            Exc.this_col_count += 1
        return True
    except:
        print("TEST")
        return False


# ------------------------------------------------------------------------------------------------------------------
def MainStarter(Job, Exc):
    """
    行ループ
    """
    try:
        for Exc.this_row_count in range(Exc.sheet_row_count):
            if Exc.this_row_count != 0:
                Exc.row_data = Exc.sheet_df.iloc[Exc.this_row_count]
                if Exc.row_data["関与先番号_関与先番号"] == Exc.row_data["関与先番号_関与先番号"]:  # nan判定
                    # nanでない場合
                    Exc.row_kanyo_no = Exc.row_data["関与先番号_関与先番号"]
                    Exc.row_kanyo_name = NameSearch(Exc.name_df, Exc.row_kanyo_no)
                    Exc.row_PDFTitle = (
                        str(Exc.row_kanyo_no)
                        + "_"
                        + str(Exc.row_kanyo_name)
                        + "_RPA決算書"
                    )
                    Exc.row_savedir = Exc.row_data["年度_(保管フォルダ名)"]
                    # マージPDFの保存先・対象システム年度をエクセルから抽出
                    if "_" in Exc.row_savedir:
                        sd = Exc.row_savedir.split("_")
                        Exc.row_savedir = sd[0]
                        Exc.year = re.sub(r"\D", "", sd[1])
                    else:
                        Exc.year = 0

                    OpenSystem(Job, Exc)
                    if Exc.CFFlag is True:
                        PMURL = PDFM.PDFMarge(
                            Job.All_url + r"\ListNumber.csv",
                            Job.Img_dir + r"\PDF",
                            Job.SerchURL,
                            Exc.row_PDFTitle,
                            Exc.row_kanyo_no,
                            Exc.row_savedir,
                        )
                        Exc.WritePDFURL(PMURL)
                else:
                    # nanの場合561
                    print("nan")
        return True, ""
    except:
        return False, ""


# ------------------------------------------------------------------------------------------------------------------
# 排他制御付print
def print_lock(ctx, str):
    with ctx["lock"]:
        print(str)


# ------------------------------------------------------------------------------------------------------------------
# スレッド処理クラス
class TestThread(Thread):
    def __init__(self, ctx, limit):
        super(TestThread, self).__init__()
        self.ctx = ctx
        self.run_flag = True  # 実行フラグ
        self.stop_flag = False  # 中断フラグ
        self.limit_time = limit
        self.start()

    def set(self, j, Ex_File):
        # self.j = j
        self.Ex_File = Ex_File
        j.stop_flag = False  # 中断フラグ

    def run(self):
        self.run_flag = True  # 実行フラグ
        th.stop_flag = False  # 中断フラグ
        print_lock(self.ctx, "sub loop start---")
        Count = 0
        while True:  # 無限ループ
            if self.ctx["stop"]:  # main側から終了を指示されたら終了
                break
            if Count == self.limit_time:  # 3600:  # カウント変数が指定に達したら終了
                j.driver.kill()
                # self.stop_flag = True  # 中断フラグ
                j.stop_flag = True  # 中断フラグ
                break
            time.sleep(0.5)
            Count += 0.5
            if Count % 10 == 0:
                print(f"{Count}秒経過")
            # print_lock(self.ctx, "sub  : " + str(datetime.datetime.today()))
        print_lock(self.ctx, "sub loop end---")
        return


# ------------------------------------------------------------------------------------------------------------------
# def call():
if __name__ == "__main__":
    global Start_Year  # 当年
    global XLSURL, MoveXLSURL
    try:
        # mainとスレッドで共有するデータ
        # ctx = {"lock": Lock(), "stop": False}
        # th = TestThread(ctx, 30)  # 別スレッドでタイマー起動
        j = Job()  # JobClass
        while_count = 0
        # Log--------------------------------------------595
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_MJS決算書印刷開始")
        # -----------------------------------------------
        # while th.run_flag is True and while_count < 2:
        #     while_count += 1
        # try:
        for curDir, dirs, files in os.walk(j.XLSDir):
            if curDir == j.XLSDir:        
                for sb_fileItem in files:
                    print(sb_fileItem)
                    if (
                        "製本・電子ファイル印刷申請ミロク" in sb_fileItem
                        and not "製本・電子ファイル印刷申請ミロク(原本).xlsm" == sb_fileItem
                    ):
                        XLSURL = (
                            curDir + r"\\" + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        MoveXLSURL = (
                            curDir
                            + r"\\MJSLog\\"
                            + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        os.rename(XLSURL, MoveXLSURL)
                        MoveXLSURL = (
                            curDir + r"\\" + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        XLSURL = (
                            curDir
                            + r"\\MJSLog\\"
                            + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        # フォルダ移動後のエクセル読込
                        Ex_File = Sheet(XLSURL)
                        # エクセルファイル名を取得
                        j.filename = os.path.splitext(os.path.basename(XLSURL))[0]

                        # print_lock(ctx, "main loop start---")
                        # JobClassがなければ
                        # if j.driver.poll() is not None:
                        #     j = Job()  # JobClass
                        #     # mainとスレッドで共有するデータ
                        #     ctx = {"lock": Lock(), "stop": False}
                        #     th = TestThread(ctx, 30)  # 別スレッドでタイマー起動
                        # メイン処理とエクセルシートクラスをタイマーに格納
                        # th.set(j, Ex_File)
                        try:
                            j.MainFlow(Ex_File)
                        except:
                            traceback.print_exc()
                        finally:
                            print("")
                            try:
                                del Ex_File.book  # エクセルブッククラスを解放
                                os.rename(XLSURL, MoveXLSURL)
                            except:
                                print("解放済")
                # 時間制限でミロクを閉じていた場合
                # if th.stop_flag is True:
                #     try:
                #         del Ex_File.book  # エクセルブッククラスを解放
                #         os.rename(XLSURL, MoveXLSURL)
                #     except:
                #         print("解放済")
        # if th.stop_flag is False:
        #     th.run_flag = False  # 実行フラグを中断状態に
        # except:
        #     continue
    finally:
        print("finally")
        # print_lock(ctx, "main loop end---")
        # ctx["stop"] = True  # スレッド側に終了を指示
        # th.join()  # スレッドの終了を待つ

# ------------------------------------------------------------------------------------------------------------------
# if __name__ == "__main__":
#     call()
#     print("")
# try:
#     con =Control.control()
# except: