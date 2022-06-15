# モジュールインポート
import pyautogui as pg
import time
import MJSOpen

# pandasインポート
import pandas as pd

# 配列計算関数numpyインポート
import numpy as np

# osインポート
import os

# datetimeインポート

# 例外処理判定の為のtracebackインポート
import traceback

# pandas(pd)で関与先データCSVを取得
import pyautogui
import pyperclip  # クリップボードへのコピーで使用
import Function.ExcelFileAction as EFA
import Function.CSVOut as FCO
import Function.MJSSPOPDFMarge as PDFM
import datetime
import openpyxl
from openpyxl.formatting.rule import Rule
from ctypes import windll

# logger設定------------------------------------------------------------------------------------------------------------
import logging.config

logging.config.fileConfig(r"LogConf\loggingMJSSysUp.conf")
logger = logging.getLogger(__name__)
# ----------------------------------------------------------------------------------------------------------------------


def DriverUIWaitXPATH(UIPATH, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        try:
            driver.find_element_by_xpath(UIPATH)
            Flag = 1
            return True
        except:
            Flag = 0
    if Flag == 0:
        return False


# ----------------------------------------------------------------------------------------------------------------------
def DriverUIWaitAutomationId(UIPATH, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        try:
            driver.find_element_by_accessibility_id(UIPATH)
            Flag = 1
            return True
        except:
            Flag = 0
    if Flag == 0:
        return False


# ----------------------------------------------------------------------------------------------------------------------
def DriverUIWaitName(UIPATH, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        try:
            driver.find_element_by_Name(UIPATH)
            Flag = 1
            return True
        except:
            Flag = 0
    if Flag == 0:
        return False


# ------------------------------------------------------------r----------------------------------------------------------
def DriverUIWaitclassname(UIPATH, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        try:
            driver.find_element_by_class_name(UIPATH)
            Flag = 1
            return True
        except:
            Flag = 0
    if Flag == 0:
        return False


# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def DriverFindClass(UIPATH, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        try:
            elList = driver.find_elements_by_class_name(UIPATH)
            Flag = 1
            return True, elList
        except:
            Flag = 0
    if Flag == 0:
        return False


# ----------------------------------------------------------------------------------------------------------------------
def DriverCheck(Hub, ObjName, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        if Hub == "AutomationID":
            if (
                DriverUIWaitAutomationId(ObjName, driver) is True
            ):  # OMSメニューの年調起動ボタンを判定して初期処理分け
                # 正常待機後処理
                driver.find_element_by_accessibility_id(ObjName)  # 一括電子申告送信ボタン
                return True
            else:
                # 異常待機後処理
                print("要素取得に失敗しました。")
        elif Hub == "XPATH":
            if DriverUIWaitXPATH(ObjName, driver) is True:  # OMSメニューの年調起動ボタンを判定して初期処理分け
                # 正常待機後処理
                driver.find_element_by_xpath(ObjName)  # 一括電子申告送信ボタン
                return True
            else:
                # 異常待機後処理
                print("要素取得に失敗しました。")
        elif Hub == "Name":
            if DriverUIWaitName(ObjName, driver) is True:  # OMSメニューの年調起動ボタンを判定して初期処理分け
                # 正常待機後処理
                driver.find_element_by_Name(ObjName)  # 一括電子申告送信ボタン
                return True
            else:
                # 異常待機後処理
                print("要素取得に失敗しました。")


# ----------------------------------------------------------------------------------------------------------------------
def DriverClick(Hub, ObjName, driver):
    if Hub == "AutomationID":
        if (
            DriverUIWaitAutomationId(ObjName, driver) is True
        ):  # OMSメニューの年調起動ボタンを判定して初期処理分け
            # 正常待機後処理
            OMSObj = driver.find_element_by_accessibility_id(ObjName)  # 一括電子申告送信ボタン
            OMSObj.click()
            return OMSObj
        else:
            # 異常待機後処理
            print("要素取得に失敗しました。")
    elif Hub == "XPATH":
        if DriverUIWaitXPATH(ObjName, driver) is True:  # OMSメニューの年調起動ボタンを判定して初期処理分け
            # 正常待機後処理
            OMSObj = driver.find_element_by_xpath(ObjName)  # 一括電子申告送信ボタン
            OMSObj.click()
            return OMSObj
        else:
            # 異常待機後処理
            print("要素取得に失敗しました。")
    elif Hub == "Name":
        if DriverUIWaitName(ObjName, driver) is True:  # OMSメニューの年調起動ボタンを判定して初期処理分け
            # 正常待機後処理
            OMSObj = driver.find_element_by_Name(ObjName)  # 一括電子申告送信ボタン
            OMSObj.click()
            return OMSObj
        else:
            # 異常待機後処理
            print("要素取得に失敗しました。")
    elif Hub == "class_name":
        if DriverUIWaitclassname(ObjName, driver) is True:  # OMSメニューの年調起動ボタンを判定して初期処理分け
            # 正常待機後処理
            OMSObj = driver.find_element_by_class_name(ObjName)  # 一括電子申告送信ボタン
            OMSObj.click()
            return OMSObj
        else:
            # 異常待機後処理
            print("要素取得に失敗しました。")


# ----------------------------------------------------------------------------------------------------------------------
def ImgCheck(FolURL2, FileName, conf, LoopVal):  # 画像があればTrueを返す関数
    ImgURL = FolURL2 + "/" + FileName
    for x in range(LoopVal):
        try:
            p = pyautogui.locateOnScreen(ImgURL, confidence=conf)
            x, y = pyautogui.center(p)
            return True, x, y
        except:
            Flag = 0
    if Flag == 0:
        return False, "", ""


# ----------------------------------------------------------------------------------------------------------------------
def ImgNothingCheck(FolURL2, FileName, conf, LoopVal):  # 画像がなければTrueを返す
    ImgURL = FolURL2 + "/" + FileName
    for x in range(LoopVal):
        try:
            p = pyautogui.locateOnScreen(ImgURL, confidence=conf)
            x, y = pyautogui.center(p)
            return False
        except:
            Flag = 0
    if Flag == 0:
        return True


# ----------------------------------------------------------------------------------------------------------------------
def ImgCheckForList(FolURL2, List, conf, LoopVal):  # リスト内の画像があればTrueと画像名を返す
    for x in range(LoopVal):
        for ListItem in List:
            ImgURL = FolURL2 + "/" + ListItem
            try:
                p = pyautogui.locateOnScreen(ImgURL, confidence=conf)
                x, y = pyautogui.center(p)
                return True, ListItem
                break
            except:
                Flag = 0
    if Flag == 0:
        return False, ""


# ----------------------------------------------------------------------------------------------------------------------
def ImgClick(FolURL2, FileName, conf, LoopVal):  # 画像があればクリックしてx,y軸を返す
    ImgURL = FolURL2 + "/" + FileName
    for x in range(10):
        if (
            ImgCheck(FolURL2, FileName, conf, LoopVal)[0] is True
        ):  # OMSメニューの年調起動ボタンを判定して初期処理分け
            # 正常待機後処理
            for y in range(10):
                try:
                    p = pyautogui.locateOnScreen(ImgURL, confidence=conf)
                    x, y = pyautogui.center(p)
                    pyautogui.click(x, y)
                    time.sleep(1)
                    return x, y
                except:
                    print("失敗")
        else:
            # 異常待機後処理
            print("要素取得に失敗しました。")


# ------------------------------------------------------------------------------------------------------------------
def NameSearch(NameDF, Rno):
    try:
        Nr = len(NameDF)
        for Nrx in range(Nr):
            NameDFRow = NameDF.iloc[Nrx]
            print(NameDFRow["コード"])
            if Rno == NameDFRow["コード"]:
                return NameDFRow["顧問先名称"]
    except:
        return "NameErr"


# ------------------------------------------------------------------------------------------------------------------
def ChildFlow(
    FolURL,
    TFolURL,
    CFolURL,
    ExRow,
    Ex,
    Eh,
    ExrcHeader,
    isnItem,
    Title,
    PN,
    driver,
    Rno,
    Rn,
):
    try:
        if "会計大将" == Title:
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_会計大将_" + PN + "印刷処理開始"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "_会計大将_" + PN + "印刷処理開始"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
            Fname = CFolURL + r"\PDF\\" + PN + ".pdf"
            SystemUp = KaikeiUpDate(FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname)
            # Excel書き込み--------------------------------------------------
            if SystemUp[0] is True:
                dt_now = datetime.datetime.now()
                dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
                WriteEx = openpyxl.load_workbook(XLSURL)
                WriteExSheet = WriteEx[isnItem]
                WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = dt_now
                print("シート書き込み完了")
                print(WriteEx)
                WriteEx.save(XLSURL)
                WriteEx.close
                # ---------------------------------------------------------------
                # Log---------------------------------------------------------------------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                logger.debug(
                    dt_s
                    + "_関与先番号:"
                    + str(Rno)
                    + ":"
                    + str(Rn)
                    + "_会計大将_"
                    + PN
                    + "印刷処理終了"
                )
                with open(LURL, "a") as f:
                    print(
                        [dt_s, "関与先番号:" + str(Rno), str(Rn), "_会計大将_" + PN + "印刷処理終了"],
                        file=f,
                    )
                # ------------------------------------------------------------------------------------------
            elif SystemUp[0] is False:
                dt_now = datetime.datetime.now()
                dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
                WriteEx = openpyxl.load_workbook(XLSURL)
                WriteExSheet = WriteEx[isnItem]
                WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "計算エラー" + dt_now
                print("シート書き込み完了")
                print(WriteEx)
                WriteEx.save(XLSURL)
                WriteEx.close
                # ---------------------------------------------------------------
                # Log---------------------------------------------------------------------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                logger.debug(
                    dt_s
                    + "_関与先番号:"
                    + str(Rno)
                    + ":"
                    + str(Rn)
                    + "_会計大将_"
                    + PN
                    + "計算エラーで処理終了"
                )
                with open(LURL, "a") as f:
                    print(
                        [
                            dt_s,
                            "関与先番号:" + str(Rno),
                            str(Rn),
                            "_会計大将_" + PN + "計算エラーで処理終了",
                        ],
                        file=f,
                    )
                # ------------------------------------------------------------------------------------------
        elif "決算内訳書" == Title:
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_決算内訳書_" + PN + "印刷処理開始"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "_決算内訳書_" + PN + "印刷処理開始"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
            Fname = CFolURL + r"\PDF\\" + PN + ".pdf"
            SystemUp = KessanUpDate(FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname)
            # Excel書き込み--------------------------------------------------
            if SystemUp[0] is True:
                dt_now = datetime.datetime.now()
                dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
                WriteEx = openpyxl.load_workbook(XLSURL)
                WriteExSheet = WriteEx[isnItem]
                WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = dt_now
                print("シート書き込み完了")
                WriteEx.save(XLSURL)
                WriteEx.close
                # Log---------------------------------------------------------------------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                logger.debug(
                    dt_s
                    + "_関与先番号:"
                    + str(Rno)
                    + ":"
                    + str(Rn)
                    + "_決算内訳書_"
                    + PN
                    + "印刷処理終了"
                )
                with open(LURL, "a") as f:
                    print(
                        [dt_s, "関与先番号:" + str(Rno), str(Rn), "_決算内訳書_" + PN + "印刷処理終了"],
                        file=f,
                    )
                # ------------------------------------------------------------------------------------------
        elif "減価償却" == Title:
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_減価償却_" + PN + "印刷処理開始"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "_減価償却_" + PN + "印刷処理開始"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
            Fname = CFolURL + r"\PDF\\" + PN + ".pdf"
            SystemUp = GenkasyoukyakuUpdate(
                FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname
            )
            # Excel書き込み---------------------------------------------------
            if SystemUp[0] is True:
                dt_now = datetime.datetime.now()
                dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
                WriteEx = openpyxl.load_workbook(XLSURL)
                WriteExSheet = WriteEx[isnItem]
                WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = dt_now
                print("シート書き込み完了")
                WriteEx.save(XLSURL)
                WriteEx.close
                # Log---------------------------------------------------------------------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                logger.debug(
                    dt_s
                    + "_関与先番号:"
                    + str(Rno)
                    + ":"
                    + str(Rn)
                    + "_減価償却_"
                    + PN
                    + "印刷処理終了"
                )
                with open(LURL, "a") as f:
                    print(
                        [dt_s, "関与先番号:" + str(Rno), str(Rn), "_減価償却_" + PN + "印刷処理終了"],
                        file=f,
                    )
            elif SystemUp[0] is False:
                dt_now = datetime.datetime.now()
                dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
                WriteEx = openpyxl.load_workbook(XLSURL)
                WriteExSheet = WriteEx[isnItem]
                WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "資産無し" + dt_now
                print("シート書き込み完了")
                print(WriteEx)
                WriteEx.save(XLSURL)
                WriteEx.close
                # ---------------------------------------------------------------
                # Log---------------------------------------------------------------------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                logger.debug(
                    dt_s
                    + "_関与先番号:"
                    + str(Rno)
                    + ":"
                    + str(Rn)
                    + "_会計大将_"
                    + PN
                    + "計算エラーで処理終了"
                )
                with open(LURL, "a") as f:
                    print(
                        [
                            dt_s,
                            "関与先番号:" + str(Rno),
                            str(Rn),
                            "_会計大将_" + PN + "資産無しで処理終了",
                        ],
                        file=f,
                    )
                # ------------------------------------------------------------------------------------------
        elif "法人税申告書" == Title:
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法人税申告書_" + PN + "印刷処理開始"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "_法人税申告書_" + PN + "印刷処理開始"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
            Fname = CFolURL + r"\PDF\\" + PN + ".pdf"
            SystemUp = HoujinzeiUpdate(
                FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname
            )
            # Excel書き込み---------------------------------------------------
            if SystemUp[0] is True:
                dt_now = datetime.datetime.now()
                dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
                WriteEx = openpyxl.load_workbook(XLSURL)
                WriteExSheet = WriteEx[isnItem]
                WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = dt_now
                print("シート書き込み完了")
                WriteEx.save(XLSURL)
                WriteEx.close
                # Log---------------------------------------------------------------------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                logger.debug(
                    dt_s
                    + "_関与先番号:"
                    + str(Rno)
                    + ":"
                    + str(Rn)
                    + "_法人税申告書_"
                    + PN
                    + "印刷処理終了"
                )
                with open(LURL, "a") as f:
                    print(
                        [
                            dt_s,
                            "関与先番号:" + str(Rno),
                            str(Rn),
                            "_法人税申告書_" + PN + "印刷処理終了",
                        ],
                        file=f,
                    )
                return True
            else:
                return False
                # ------------------------------------------------------------------------------------------
        elif "電子申告" == Title:
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_電子申告_" + PN + "印刷処理開始"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "_電子申告_" + PN + "印刷処理開始"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
            Fname = CFolURL + r"\PDF\\" + PN + ".pdf"
            SystemUp = DensisinkokuUpDate(
                FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname
            )
            # Excel書き込み--------------------------------------------------
            if SystemUp[0] is True:
                dt_now = datetime.datetime.now()
                dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
                WriteEx = openpyxl.load_workbook(XLSURL)
                WriteExSheet = WriteEx[isnItem]
                WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = dt_now
                print("シート書き込み完了")
                print(WriteEx)
                WriteEx.save(XLSURL)
                WriteEx.close
                # ---------------------------------------------------------------
                # Log---------------------------------------------------------------------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                logger.debug(
                    dt_s
                    + "_関与先番号:"
                    + str(Rno)
                    + ":"
                    + str(Rn)
                    + "_電子申告_"
                    + PN
                    + "印刷処理終了"
                )
                with open(LURL, "a") as f:
                    print(
                        [dt_s, "関与先番号:" + str(Rno), str(Rn), "_電子申告_" + PN + "印刷処理終了"],
                        file=f,
                    )
                return True
            else:
                return False
                # ------------------------------------------------------------------------------------------
        else:
            print("NoSystem")
            return True
    except:
        return False


# ------------------------------------------------------------------------------------------------------------------
def HoujinzeiUpdateSinkokuItiran(
    FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname, ThisNo, ThisYear, ThisMonth
):
    # 申告税一覧表印刷処理----------------------------------------------------
    OP = ImgCheckForList(
        CFolURL,
        [
            r"\Houjinzei\01SinkokuNyuuryoku.png",
            r"\Houjinzei\01SinkokuNyuuryoku2.png",
        ],
        0.9,
        10,
    )
    if OP[0] is True:
        ImgClick(CFolURL, OP[1], 0.9, 10)
        # 法人税メニューが表示されるまで待機------------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\HoujinOpen.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
        # --------------------------------------------------------------------
        if PN == "申告税一覧表":
            ImgClick(CFolURL, r"\Houjinzei\DownPrint.png", 0.9, 10)
            time.sleep(1)
            pg.press("n")
            # 一覧表メニューが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(
                    CFolURL + r"\Houjinzei\ItiranFlag.png", confidence=0.9
                )
                is None
            ):
                time.sleep(1)
                SRC = ImgCheck(CFolURL, r"\Houjinzei\S_RendouCheck.png", 0.9, 10)
                if SRC[0] is True:
                    pg.press("return")
                SRCC = ImgCheck(CFolURL, r"\Houjinzei\S_Rendou2.png", 0.9, 10)
                if SRCC[0] is True:
                    pg.press("n")
                SRCCC = ImgCheck(CFolURL, r"\Houjinzei\S_Rendou3.png", 0.9, 10)
                if SRCCC[0] is True:
                    pg.press("n")
                # 新規別表追加選択-------------------------------------------------
                SB = ImgCheck(CFolURL, r"\Houjinzei\SaiyouBeppyou.png", 0.9, 10)
                if SB[0] is True:
                    ImgClick(CFolURL, r"\Houjinzei\SBKousin.png", 0.9, 10)
            # --------------------------------------------------------------------
            ImgClick(CFolURL, r"\Houjinzei\Print.png", 0.9, 10)
            # 一覧表出力項目指定が表示されるまで待機---------------------------------
            while (
                pg.locateOnScreen(
                    CFolURL + r"\Houjinzei\ItiranSyutuQ.png", confidence=0.9
                )
                is None
            ):
                time.sleep(1)
            # --------------------------------------------------------------------
            pg.press("return")
            # 一覧表出力項目指定が表示されるまで待機---------------------------------
            while (
                pg.locateOnScreen(CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            # --------------------------------------------------------------------
            # 申告税一覧表印刷処理----------------------------------------------------
            FO = ImgCheckForList(
                CFolURL,
                [
                    r"\Houjinzei\FileOut.png",
                    r"\Houjinzei\FileOut2.png",
                ],
                0.9,
                10,
            )
            if FO[0] is True:
                ImgClick(CFolURL, FO[1], 0.9, 10)
            ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
            pg.press("return")
            pg.press("delete")
            pg.press("backspace")
            pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
            pg.hotkey("ctrl", "v")
            pg.press("return")
            time.sleep(1)
            ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
            time.sleep(2)
            # 印刷中が表示されるまで待機---------------------------------
            IC = 0
            while (
                pg.locateOnScreen(CFolURL + r"\Houjinzei\NowPrint.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
                IC += 1
                if IC == 5:
                    pg.press("tab")
                    break
                FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
                if FO[0] is True:
                    pg.press("y")
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\NowPrint.png",
                            confidence=0.9,
                        )
                        is None
                    ):
                        time.sleep(1)
            # --------------------------------------------------------------------
            # 印刷中が表示されなくなるまで待機---------------------------------
            while (
                pg.locateOnScreen(CFolURL + r"\Houjinzei\NowPrint.png", confidence=0.9)
                is not None
            ):
                time.sleep(1)
            # --------------------------------------------------------------------
            time.sleep(1)
            pg.keyDown("alt")
            pg.press("x")
            pg.keyUp("alt")
            time.sleep(3)
            # 確実に閉じる---------------------------------------------------------
            HME = ImgCheck(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
            if HME[0] is True:
                ImgClick(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
            else:
                pg.keyDown("alt")
                pg.press("x")
                pg.keyUp("alt")
            # --------------------------------------------------------------------
            # 終了確認が表示されるまで待機---------------------------------
            while (
                pg.locateOnScreen(
                    CFolURL + r"\Houjinzei\SinkokuEndQ.png", confidence=0.9
                )
                is None
            ):
                time.sleep(1)
            # --------------------------------------------------------------------
            pg.press("y")
            while (
                ImgCheckForList(
                    CFolURL,
                    [
                        r"\Houjinzei\01SinkokuNyuuryoku.png",
                        r"\Houjinzei\01SinkokuNyuuryoku2.png",
                    ],
                    0.9,
                    10,
                )[0]
                is False
            ):
                time.sleep(1)
                SEQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ2.png", 0.9, 10)
                if SEQ[0] is True:
                    ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ2Btn.png", 0.9, 10)
                SEQQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ3.png", 0.9, 10)
                if SEQQ[0] is True:
                    pg.press("return")
                    # 地方税一覧入力が表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\SinkokuEndQ4.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ4.png", 0.9, 10)
                    # --------------------------------------------------------------------
            # 閉じる処理--------------------------
            pg.keyDown("alt")
            pg.press("f4")
            pg.keyUp("alt")
            # -----------------------------------
            # 法人税フラグが表示されるまで待機-------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\HoujinFlag.png", confidence=0.9) is None
            ):
                time.sleep(1)
            # --------------------------------------------------------------------
            # 初期画面で開封された法人税項目を閉じる----------------------------------
            HoujinList = [
                r"\FastMenuHoujinzei.png",
                r"\FastMenuHoujinzei2.png",
            ]
            HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
            if HLI[0] is True:
                ImgClick(TFolURL, HLI[1], 0.9, 10)
            # --------------------------------------------------------------------
            return True, ThisNo, ThisYear, ThisMonth
        elif (
            PN == "第6号様式（県）"
            or PN == "第6号様式別表9（県）"
            or PN == "第20号様式（市）"
            or PN == "第22号の2様式"
            or PN == "別表１　緑色"
        ):
            ImgClick(CFolURL, r"\Houjinzei\HoujinzeiSelecter.png", 0.9, 10)
            # 申告書印刷メニューが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(
                    CFolURL + r"\Houjinzei\BeppyouIcon.png", confidence=0.9
                )
                is None
            ):
                time.sleep(1)
                SLQ = ImgCheck(CFolURL, r"\Houjinzei\SelectQ.png", 0.9, 10)
                if SLQ[0] is True:
                    ImgClick(CFolURL, r"\Houjinzei\Kousin.png", 0.9, 10)
            # --------------------------------------------------------------------
            if PN == "第6号様式（県）":
                ImgClick(CFolURL, r"\Houjinzei\HSCancel.png", 0.9, 10)
                time.sleep(1)
                pg.write("0600")
                pg.press("return")
            elif PN == "第6号様式別表9（県）":
                ImgClick(CFolURL, r"\Houjinzei\HSCancel.png", 0.9, 10)
                time.sleep(1)
                pg.write("0690")
                pg.press("return")
            elif PN == "第20号様式（市）":
                ImgClick(CFolURL, r"\Houjinzei\HSCancel.png", 0.9, 10)
                time.sleep(1)
                pg.write("2000")
                pg.press("return")
            elif PN == "第22号の2様式":
                ImgClick(CFolURL, r"\Houjinzei\HSCancel.png", 0.9, 10)
                time.sleep(1)
                pg.write("2202")
                pg.press("return")
            elif PN == "別表１　緑色":
                pg.press("home")
                pg.press("return")

            if PN == "第20号様式（市）":
                # 市町村の数だけループ-------------------------------------------------------------
                # 画像に名前を付ける
                SSN = "20ScreenShot.png"
                SSN2 = "20ScreenShot2.png"
                Win = "Window.png"
                while ImgCheck(CFolURL, r"\\All\\" + Win, 0.9, 10)[0] is True:
                    WC = ImgCheck(CFolURL, r"\\All\\" + Win, 0.9, 10)
                    pyautogui.click(WC[1], WC[2], button="right")
                    pg.press("x")
                time.sleep(1)
                ImgClick(CFolURL, r"\Houjinzei\PreviewIcon.png", 0.9, 10)
                # プレビュー画面が表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\HoujinOpen.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                time.sleep(3)
                # ロード完了まで待機----------------------------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\PreviewLoad.png", confidence=0.9
                    )
                    is not None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                TPI = ImgCheckForList(
                    CFolURL,
                    [r"\Houjinzei\ThisPIcon.png", r"\Houjinzei\ThisPIcon2.png"],
                    0.9,
                    10,
                )  # 現在項印刷アイコンを探す
                if TPI[0] is True:  # 現在項印刷アイコンがあれば
                    # pg.keyDown("alt")
                    # pg.press("c")
                    # pg.keyup("alt")
                    # ImgClick(TFolURL, r"\HoujinFlag.png", 0.9, 10)
                    ImgClick(CFolURL, TPI[1], 0.9, 10)  # 現在項印刷アイコンをクリック
                    # 印刷ダイアログ待機----------------------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\ThisPMenu.png",
                            confidence=0.9,
                        )
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                    pg.hotkey("ctrl", "v")
                    pg.press("return")
                    # 印刷完了まで待機----------------------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\ThisPMenu2.png",
                            confidence=0.9,
                        )
                        is not None
                    ):
                        time.sleep(1)
                        TPOQ = ImgCheck(
                            CFolURL, r"\Houjinzei\ThisPOverQ.png", 0.9, 10
                        )  # 上書き確認
                        if TPOQ[0] is True:
                            pg.press("y")  # yで上書き
                    # -------------------------------------------------------------------
                    # 印刷完了まで待機----------------------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\NowPrint.png", confidence=0.9
                        )
                        is not None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                # 市町村の数だけループ-------------------------------------------------------------
                for CNo in range(1, 100):
                    ICFL = ImgCheckForList(
                        CFolURL,
                        [r"\\All\\TihouKey.png", r"\\All\\TihouKey.png"],
                        0.9,
                        10,
                    )
                    if ICFL[0] is True:
                        # 一回目のスクショ----------------------------------------------
                        s = pyautogui.screenshot(region=(820, 250, 450, 40))  # スクショ
                        s.save(CFolURL + r"\\All\\" + SSN)  # 保存
                        # -------------------------------------------------------------
                        ImgClick(CFolURL, ICFL[1], 0.9, 10)
                        CNo += 1
                        pg.write(str(CNo))
                        pg.press("return")
                        # 二回目以降のスクショ------------------------------------------
                        s = pyautogui.screenshot(region=(820, 250, 450, 40))  # スクショ
                        s.save(CFolURL + r"\\All\\" + SSN2)  # 保存
                        # -------------------------------------------------------------
                    WC = ImgCheck(CFolURL, r"\\All\\" + SSN, 0.9, 10)
                    if WC[0] is True:
                        break
                    else:
                        time.sleep(1)
                        ImgClick(CFolURL, r"\Houjinzei\PreviewIcon.png", 0.9, 10)
                        # プレビュー画面が表示されるまで待機------------------------------------
                        while (
                            pg.locateOnScreen(
                                CFolURL + r"\Houjinzei\HoujinOpen.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        time.sleep(3)
                        # ロード完了まで待機----------------------------------------------------
                        while (
                            pg.locateOnScreen(
                                CFolURL + r"\Houjinzei\PreviewLoad.png", confidence=0.9
                            )
                            is not None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        TPI = ImgCheckForList(
                            CFolURL,
                            [r"\Houjinzei\ThisPIcon.png", r"\Houjinzei\ThisPIcon2.png"],
                            0.9,
                            10,
                        )  # 現在項印刷アイコンを探す
                        if TPI[0] is True:  # 現在項印刷アイコンがあれば
                            # pg.keyDown("alt")
                            # pg.press("c")
                            # pg.keyup("alt")
                            # ImgClick(TFolURL, r"\HoujinFlag.png", 0.9, 10)
                            ImgClick(CFolURL, TPI[1], 0.9, 10)  # 現在項印刷アイコンをクリック
                            # 印刷ダイアログ待機----------------------------------------------------
                            while (
                                pg.locateOnScreen(
                                    CFolURL + r"\Houjinzei\ThisPMenu.png",
                                    confidence=0.9,
                                )
                                is None
                            ):
                                time.sleep(1)
                            # --------------------------------------------------------------------
                            Fname = Fname.replace(".pdf", "") + str(CNo) + ".pdf"
                            pyperclip.copy(
                                Fname.replace("\\\\", "\\").replace("/", "\\")
                            )
                            pg.hotkey("ctrl", "v")
                            pg.press("return")
                            # 印刷完了まで待機----------------------------------------------------
                            while (
                                pg.locateOnScreen(
                                    CFolURL + r"\Houjinzei\ThisPMenu2.png",
                                    confidence=0.9,
                                )
                                is not None
                            ):
                                time.sleep(1)
                                TPOQ = ImgCheck(
                                    CFolURL, r"\Houjinzei\ThisPOverQ.png", 0.9, 10
                                )  # 上書き確認
                                if TPOQ[0] is True:
                                    pg.press("y")  # yで上書き
                            # -------------------------------------------------------------------
                            # 印刷完了まで待機----------------------------------------------------
                            while (
                                pg.locateOnScreen(
                                    CFolURL + r"\Houjinzei\NowPrint.png", confidence=0.9
                                )
                                is not None
                            ):
                                time.sleep(1)
                            # --------------------------------------------------------------------
                # 確実に閉じる---------------------------------------------------------
                HME = ImgCheck(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
                if HME[0] is True:
                    ImgClick(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
                else:
                    pg.keyDown("alt")
                    pg.press("x")
                    pg.keyUp("alt")
                # --------------------------------------------------------------------
                # 終了確認が表示されるまで待機---------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\SinkokuEndQ.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                pg.press("y")
                while (
                    ImgCheckForList(
                        CFolURL,
                        [
                            r"\Houjinzei\01SinkokuNyuuryoku.png",
                            r"\Houjinzei\01SinkokuNyuuryoku2.png",
                        ],
                        0.9,
                        10,
                    )[0]
                    is False
                ):
                    time.sleep(1)
                    SEQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ2.png", 0.9, 10)
                    if SEQ[0] is True:
                        ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ2Btn.png", 0.9, 10)
                    SEQQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ3.png", 0.9, 10)
                    if SEQQ[0] is True:
                        pg.press("return")
                        # 地方税一覧入力が表示されるまで待機---------------------------------
                        while (
                            pg.locateOnScreen(
                                CFolURL + r"\Houjinzei\SinkokuEndQ4.png",
                                confidence=0.9,
                            )
                            is None
                        ):
                            time.sleep(1)
                        ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ4.png", 0.9, 10)
                        # --------------------------------------------------------------------
                # 閉じる処理--------------------------
                pg.keyDown("alt")
                pg.press("f4")
                pg.keyUp("alt")
                # -----------------------------------
                # 法人税フラグが表示されるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\HoujinFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                # 初期画面で開封された法人税項目を閉じる----------------------------------
                HoujinList = [
                    r"\FastMenuHoujinzei.png",
                    r"\FastMenuHoujinzei2.png",
                ]
                HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
                if HLI[0] is True:
                    ImgClick(TFolURL, HLI[1], 0.9, 10)
                # --------------------------------------------------------------------
                return True, ThisNo, ThisYear, ThisMonth
                # --------------------------------------------------------------------------------
            else:
                time.sleep(1)
                ImgClick(CFolURL, r"\Houjinzei\PreviewIcon.png", 0.9, 10)
                # プレビュー画面が表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\HoujinOpen.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                time.sleep(3)
                # ロード完了まで待機----------------------------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\PreviewLoad.png", confidence=0.9
                    )
                    is not None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                TPI = ImgCheckForList(
                    CFolURL,
                    [r"\Houjinzei\ThisPIcon.png", r"\Houjinzei\ThisPIcon2.png"],
                    0.9,
                    10,
                )  # 現在項印刷アイコンを探す
                if TPI[0] is True:  # 現在項印刷アイコンがあれば
                    # pg.keyDown("alt")
                    # pg.press("c")
                    # pg.keyup("alt")
                    # ImgClick(TFolURL, r"\HoujinFlag.png", 0.9, 10)
                    ImgClick(CFolURL, TPI[1], 0.9, 10)  # 現在項印刷アイコンをクリック
                    # 印刷ダイアログ待機----------------------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\ThisPMenu.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                    pg.hotkey("ctrl", "v")
                    pg.press("return")
                    # 印刷完了まで待機----------------------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\ThisPMenu2.png", confidence=0.9
                        )
                        is not None
                    ):
                        time.sleep(1)
                        TPOQ = ImgCheck(
                            CFolURL, r"\Houjinzei\ThisPOverQ.png", 0.9, 10
                        )  # 上書き確認
                        if TPOQ[0] is True:
                            pg.press("y")  # yで上書き
                    # -------------------------------------------------------------------
                    # 印刷完了まで待機----------------------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\NowPrint.png", confidence=0.9
                        )
                        is not None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    # 確実に閉じる---------------------------------------------------------
                    HME = ImgCheck(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
                    if HME[0] is True:
                        ImgClick(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("x")
                        pg.keyUp("alt")
                    # --------------------------------------------------------------------
                    # 終了確認が表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\SinkokuEndQ.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    pg.press("y")
                    while (
                        ImgCheckForList(
                            CFolURL,
                            [
                                r"\Houjinzei\01SinkokuNyuuryoku.png",
                                r"\Houjinzei\01SinkokuNyuuryoku2.png",
                            ],
                            0.9,
                            10,
                        )[0]
                        is False
                    ):
                        time.sleep(1)
                        SEQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ2.png", 0.9, 10)
                        if SEQ[0] is True:
                            ImgClick(
                                CFolURL, r"\Houjinzei\SinkokuEndQ2Btn.png", 0.9, 10
                            )
                        SEQQ = ImgCheck(
                            CFolURL, r"\Houjinzei\SinkokuEndQ3.png", 0.9, 10
                        )
                        if SEQQ[0] is True:
                            pg.press("return")
                            # 地方税一覧入力が表示されるまで待機---------------------------------
                            while (
                                pg.locateOnScreen(
                                    CFolURL + r"\Houjinzei\SinkokuEndQ4.png",
                                    confidence=0.9,
                                )
                                is None
                            ):
                                time.sleep(1)
                            ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ4.png", 0.9, 10)
                            # --------------------------------------------------------------------
                    # 閉じる処理--------------------------
                    pg.keyDown("alt")
                    pg.press("f4")
                    pg.keyUp("alt")
                    # -----------------------------------
                    # 法人税フラグが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(TFolURL + r"\HoujinFlag.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    # 初期画面で開封された法人税項目を閉じる----------------------------------
                    HoujinList = [
                        r"\FastMenuHoujinzei.png",
                        r"\FastMenuHoujinzei2.png",
                    ]
                    HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
                    if HLI[0] is True:
                        ImgClick(TFolURL, HLI[1], 0.9, 10)
                    # --------------------------------------------------------------------
                    return True, ThisNo, ThisYear, ThisMonth


# ------------------------------------------------------------------------------------------------------------------
def HoujinzeiUpdateZeimuDairi(
    FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname, ThisNo, ThisYear, ThisMonth
):
    # 税務代理権限証書印刷処理----------------------------------------------------
    OP = ImgCheckForList(
        CFolURL,
        [
            r"\Houjinzei\15Zeimudairi.png",
            r"\Houjinzei\15Zeimudairi2.png",
        ],
        0.9,
        10,
    )
    if OP[0] is True:
        ImgClick(CFolURL, OP[1], 0.9, 10)
        # 法人税メニューが表示されるまで待機------------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\HoujinOpen.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
            # 新規別表追加選択-------------------------------------------------
            SB = ImgCheck(CFolURL, r"\Houjinzei\SaiyouBeppyou.png", 0.9, 10)
            if SB[0] is True:
                ImgClick(CFolURL, r"\Houjinzei\SBKousin.png", 0.9, 10)
        # --------------------------------------------------------------------
        ImgClick(CFolURL, r"\Houjinzei\ZeimuPrint.png", 0.9, 10)
        time.sleep(1)
        # 一覧表出力項目指定が表示されるまで待機---------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
        # --------------------------------------------------------------------
        # 申告税一覧表印刷処理----------------------------------------------------
        FO = ImgCheckForList(
            CFolURL,
            [
                r"\Houjinzei\FileOut.png",
                r"\Houjinzei\FileOut2.png",
            ],
            0.9,
            10,
        )
        if FO[0] is True:
            ImgClick(CFolURL, FO[1], 0.9, 10)
        ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
        pg.press("return")
        pg.press("delete")
        pg.press("backspace")
        pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
        pg.hotkey("ctrl", "v")
        pg.press("return")
        time.sleep(1)
        ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
        time.sleep(2)
        # 印刷中が表示されなくなるまで待機---------------------------------
        while (
            pg.locateOnScreen(
                CFolURL + r"\Houjinzei\ZeimuPrintFlag.png", confidence=0.9
            )
            is not None
        ):
            time.sleep(1)
            FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
            if FO[0] is True:
                pg.press("y")
                time.sleep(2)
        # --------------------------------------------------------------------
        time.sleep(1)
        # 確実に閉じる---------------------------------------------------------
        HME = ImgCheck(CFolURL, r"\Houjinzei\ZeimuMEnd.png", 0.9, 10)
        if HME[0] is True:
            ImgClick(CFolURL, r"\Houjinzei\ZeimuMEnd.png", 0.9, 10)
        else:
            pg.keyDown("alt")
            pg.press("x")
            pg.keyUp("alt")
        # --------------------------------------------------------------------
        # 終了確認が表示されるまで待機---------------------------------
        while (
            ImgCheckForList(
                CFolURL,
                [
                    r"\Houjinzei\15Zeimudairi.png",
                    r"\Houjinzei\15Zeimudairi2.png",
                ],
                0.9,
                10,
            )[0]
            is False
        ):
            time.sleep(1)
            SEQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ2.png", 0.9, 10)
            if SEQ[0] is True:
                ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ2Btn.png", 0.9, 10)
            SEQQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ3.png", 0.9, 10)
            if SEQQ[0] is True:
                pg.press("return")
                # 地方税一覧入力が表示されるまで待機---------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\SinkokuEndQ4.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ4.png", 0.9, 10)
                # --------------------------------------------------------------------
        # 閉じる処理--------------------------
        pg.keyDown("alt")
        pg.press("f4")
        pg.keyUp("alt")
        # -----------------------------------
        # 法人税フラグが表示されるまで待機-------------------------------------
        while pg.locateOnScreen(TFolURL + r"\HoujinFlag.png", confidence=0.9) is None:
            time.sleep(1)
        # --------------------------------------------------------------------
        # 初期画面で開封された法人税項目を閉じる----------------------------------
        HoujinList = [
            r"\FastMenuHoujinzei.png",
            r"\FastMenuHoujinzei2.png",
        ]
        HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
        if HLI[0] is True:
            ImgClick(TFolURL, HLI[1], 0.9, 10)
        # --------------------------------------------------------------------
        return True, ThisNo, ThisYear, ThisMonth


# ------------------------------------------------------------------------------------------------------------------
def HoujinzeiUpdateSyomen(
    FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname, ThisNo, ThisYear, ThisMonth
):
    # 税務代理権限証書印刷処理----------------------------------------------------
    OP = ImgCheckForList(
        CFolURL,
        [
            r"\Houjinzei\16TenpuSyomen.png",
            r"\Houjinzei\16TenpuSyomen2.png",
        ],
        0.9,
        10,
    )
    if OP[0] is True:
        ImgClick(CFolURL, OP[1], 0.9, 10)
        # 法人税メニューが表示されるまで待機------------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\HoujinOpen.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
            # 新規別表追加選択-------------------------------------------------
            SB = ImgCheck(CFolURL, r"\Houjinzei\SaiyouBeppyou.png", 0.9, 10)
            if SB[0] is True:
                ImgClick(CFolURL, r"\Houjinzei\SBKousin.png", 0.9, 10)
        # --------------------------------------------------------------------
        time.sleep(3)
        AYR = ImgCheck(CFolURL, r"\Houjinzei\AYear.png", 0.9, 10)
        if AYR[0] is True:
            pg.press("n")
        ImgClick(CFolURL, r"\Houjinzei\ZeimuPrint.png", 0.9, 10)
        time.sleep(1)
        # 用紙選択が表示されるまで待機---------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\YousiSentaku.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
        # --------------------------------------------------------------------
        ImgClick(CFolURL, r"\Houjinzei\YousiOK.png", 0.9, 10)
        # 申告税一覧表印刷処理----------------------------------------------------
        FO = ImgCheckForList(
            CFolURL,
            [
                r"\Houjinzei\FileOut.png",
                r"\Houjinzei\FileOut2.png",
            ],
            0.9,
            10,
        )
        if FO[0] is True:
            ImgClick(CFolURL, FO[1], 0.9, 10)
        ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
        pg.press("return")
        pg.press("delete")
        pg.press("backspace")
        pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
        pg.hotkey("ctrl", "v")
        pg.press("return")
        time.sleep(1)
        ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
        time.sleep(2)
        # 操作ガイドが表示されなくなるまで待機---------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\HPCFlag.png", confidence=0.9)
            is not None
        ):
            time.sleep(1)
            FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
            if FO[0] is True:
                pg.press("y")
        time.sleep(1)
        # 確実に閉じる---------------------------------------------------
        ZME = ImgCheck(CFolURL, r"\Houjinzei\ZeimuMEnd.png", 0.9, 10)
        if ZME[0] is True:
            ImgClick(CFolURL, r"\Houjinzei\ZeimuMEnd.png", 0.9, 10)
        else:
            pg.keyDown("alt")
            pg.press("x")
            pg.keyUp("alt")
        # ---------------------------------------------------------------
        # 終了確認が表示されるまで待機---------------------------------
        while (
            ImgCheckForList(
                CFolURL,
                [
                    r"\Houjinzei\16TenpuSyomen.png",
                    r"\Houjinzei\16TenpuSyomen2.png",
                ],
                0.9,
                10,
            )[0]
            is False
        ):
            time.sleep(1)
            SEQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ2.png", 0.9, 10)
            if SEQ[0] is True:
                ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ2Btn.png", 0.9, 10)
            SEQQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ3.png", 0.9, 10)
            if SEQQ[0] is True:
                pg.press("return")
                # 地方税一覧入力が表示されるまで待機---------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\SinkokuEndQ4.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ4.png", 0.9, 10)
                # --------------------------------------------------------------------
        # 閉じる処理--------------------------
        pg.keyDown("alt")
        pg.press("f4")
        pg.keyUp("alt")
        # -----------------------------------
        # 法人税フラグが表示されるまで待機-------------------------------------
        while pg.locateOnScreen(TFolURL + r"\HoujinFlag.png", confidence=0.9) is None:
            time.sleep(1)
        # --------------------------------------------------------------------
        # 初期画面で開封された法人税項目を閉じる----------------------------------
        HoujinList = [
            r"\FastMenuHoujinzei.png",
            r"\FastMenuHoujinzei2.png",
        ]
        HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
        if HLI[0] is True:
            ImgClick(TFolURL, HLI[1], 0.9, 10)
        # --------------------------------------------------------------------
        return True, ThisNo, ThisYear, ThisMonth


# ------------------------------------------------------------------------------------------------------------------
def HoujinzeiUpdateBeppyou(
    FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname, ThisNo, ThisYear, ThisMonth
):
    # 別表2-16印刷処理--------------------------------------------------------
    OP = ImgCheckForList(
        CFolURL,
        [
            r"\Houjinzei\01SinkokuNyuuryoku.png",
            r"\Houjinzei\01SinkokuNyuuryoku2.png",
        ],
        0.9,
        10,
    )
    if OP[0] is True:
        ImgClick(CFolURL, OP[1], 0.9, 10)
        # 法人税メニューが表示されるまで待機------------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\HoujinOpen.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
            # 新規別表追加選択-------------------------------------------------
            SB = ImgCheck(CFolURL, r"\Houjinzei\SaiyouBeppyou.png", 0.9, 10)
            if SB[0] is True:
                ImgClick(CFolURL, r"\Houjinzei\SBKousin.png", 0.9, 10)
        # --------------------------------------------------------------------
        ImgClick(CFolURL, r"\Houjinzei\DownPrint.png", 0.9, 10)
        time.sleep(1)
        pg.press("s")
        # 一覧表メニューが表示されるまで待機------------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\SItiranFlag.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
            HEQ = ImgCheck(CFolURL, r"\Houjinzei\HNoEntryQ.png", 0.9, 10)
            if HEQ[0] is True:  # 法人番号未登録ダイアログが表示されていたら
                pg.press("y")  # yで確定
        # --------------------------------------------------------------------
        if ImgCheck(CFolURL, r"\Houjinzei\Teisyutu.png", 0.99999, 10)[0] is True:
            ImgClick(CFolURL, r"\Houjinzei\Teisyutu.png", 0.99999, 10)
        if ImgCheck(CFolURL, r"\Houjinzei\Nyuuryoku.png", 0.99999, 10)[0] is True:
            ImgClick(CFolURL, r"\Houjinzei\Nyuuryoku.png", 0.99999, 10)
        if ImgCheck(CFolURL, r"\Houjinzei\Hikae.png", 0.99999, 10)[0] is True:
            ImgClick(CFolURL, r"\Houjinzei\Hikae.png", 0.99999, 10)
        if ImgCheck(CFolURL, r"\Houjinzei\Beppyou.png", 0.99999, 10)[0] is True:
            ImgClick(CFolURL, r"\Houjinzei\Beppyou.png", 0.99999, 10)
            pg.press("space")
        # --------------------------------------------------------------------
        # while (
        #     ImgCheckForList(
        #         CFolURL,
        #         [
        #             r"\Houjinzei\6Icon.png",
        #             r"\Houjinzei\69Icon.png",
        #             r"\Houjinzei\20Icon.png",
        #         ],
        #         0.99999,
        #         10,
        #     )[0]
        #     is False
        # ):
        #     pg.press(["pagedown", "pagedown"])
        # while (
        #     ImgCheckForList(
        #         CFolURL,
        #         [
        #             r"\Houjinzei\6Icon.png",
        #             r"\Houjinzei\69Icon.png",
        #             r"\Houjinzei\20Icon.png",
        #         ],
        #         0.99999,
        #         10,
        #     )[0]
        #     is True
        # ):
        #     ImgClick(
        #         CFolURL,
        #         ImgCheckForList(
        #             CFolURL,
        #             [
        #                 r"\Houjinzei\6Icon.png",
        #                 r"\Houjinzei\69Icon.png",
        #                 r"\Houjinzei\20Icon.png",
        #             ],
        #             0.99999,
        #             10,
        #         )[1],
        #         0.99999,
        #         10,
        #     )
        #     pg.press("space")
        ImgClick(CFolURL, r"\Houjinzei\SPrint.png", 0.9, 10)
        # 一覧表出力項目指定が表示されるまで待機---------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
        # --------------------------------------------------------------------
        # 申告税一覧表印刷処理----------------------------------------------------
        FO = ImgCheckForList(
            CFolURL,
            [
                r"\Houjinzei\FileOut.png",
                r"\Houjinzei\FileOut2.png",
            ],
            0.9,
            10,
        )
        if FO[0] is True:
            ImgClick(CFolURL, FO[1], 0.9, 10)
        ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
        pg.press("return")
        pg.press("delete")
        pg.press("backspace")
        pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
        pg.hotkey("ctrl", "v")
        pg.press("return")
        time.sleep(1)
        ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
        time.sleep(2)
        # 印刷中が表示されるまで待機---------------------------------
        IC = 0
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\NowPrint.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
            IC += 1
            if IC == 5:
                pg.press("tab")
                break
            FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
            if FO[0] is True:
                pg.press("y")
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\NowPrint.png",
                        confidence=0.9,
                    )
                    is None
                ):
                    time.sleep(1)
        # --------------------------------------------------------------------
        # 印刷中が表示されなくなるまで待機---------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\NowPrint.png", confidence=0.9)
            is not None
        ):
            time.sleep(1)
        # --------------------------------------------------------------------
        time.sleep(3)
        # 確実に閉じる---------------------------------------------------
        pg.keyDown("alt")
        pg.press("x")
        pg.keyUp("alt")
        # ---------------------------------------------------------------
        time.sleep(3)
        # 確実に閉じる---------------------------------------------------
        ZME = ImgCheck(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
        if ZME[0] is True:
            ImgClick(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
        else:
            pg.keyDown("alt")
            pg.press("x")
            pg.keyUp("alt")
        # ---------------------------------------------------------------
        # 終了確認が表示されるまで待機---------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\SinkokuEndQ.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
        # --------------------------------------------------------------------
        pg.press("y")
        while (
            ImgCheckForList(
                CFolURL,
                [
                    r"\Houjinzei\01SinkokuNyuuryoku.png",
                    r"\Houjinzei\01SinkokuNyuuryoku2.png",
                ],
                0.9,
                10,
            )[0]
            is False
        ):
            time.sleep(1)
            SEQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ2.png", 0.9, 10)
            if SEQ[0] is True:
                ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ2Btn.png", 0.9, 10)
            SEQQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ3.png", 0.9, 10)
            if SEQQ[0] is True:
                pg.press("return")
                # 地方税一覧入力が表示されるまで待機---------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\SinkokuEndQ4.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ4.png", 0.9, 10)
                # --------------------------------------------------------------------
        # 閉じる処理--------------------------
        pg.keyDown("alt")
        pg.press("f4")
        pg.keyUp("alt")
        # -----------------------------------
        # 法人税フラグが表示されるまで待機-------------------------------------
        while pg.locateOnScreen(TFolURL + r"\HoujinFlag.png", confidence=0.9) is None:
            time.sleep(1)
        # --------------------------------------------------------------------
        # 初期画面で開封された法人税項目を閉じる----------------------------------
        HoujinList = [
            r"\FastMenuHoujinzei.png",
            r"\FastMenuHoujinzei2.png",
        ]
        HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
        if HLI[0] is True:
            ImgClick(TFolURL, HLI[1], 0.9, 10)
        # --------------------------------------------------------------------
        PDFM.BeppyouPDFSplit(
            Fname.replace("\\\\", "\\").replace("/", "\\"), CFolURL + r"\PDF"
        )
        return True, ThisNo, ThisYear, ThisMonth


# ------------------------------------------------------------------------------------------------------------------
def HoujinzeiUpdateGaikyou(
    FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname, ThisNo, ThisYear, ThisMonth
):
    # 事業概況説明書印刷処理--------------------------------------------------------
    OP = ImgCheckForList(
        CFolURL,
        [
            r"\Houjinzei\02JigyouGaikyou.png",
            r"\Houjinzei\02JigyouGaikyou2.png",
        ],
        0.9,
        10,
    )
    if OP[0] is True:
        ImgClick(CFolURL, OP[1], 0.9, 10)
        # 法人税メニューが表示されるまで待機------------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\HoujinOpen.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
            # 新規別表追加選択-------------------------------------------------
            SB = ImgCheck(CFolURL, r"\Houjinzei\SaiyouBeppyou.png", 0.9, 10)
            if SB[0] is True:
                ImgClick(CFolURL, r"\Houjinzei\SBKousin.png", 0.9, 10)
        # --------------------------------------------------------------------
        ImgClick(CFolURL, r"\Houjinzei\GaikyouPrint.png", 0.9, 10)
        # 一覧表出力項目指定が表示されるまで待機---------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
            # 印刷条件設定----------------------------------------------------------
            GP = ImgCheck(CFolURL, r"\Houjinzei\GPMenu.png", 0.9, 10)
            if GP[0] is True:
                # 白紙チェックボックス確認
                HP = ImgCheck(CFolURL, r"\Houjinzei\HPCheck.png", 0.9, 10)
                if HP[0] is True:
                    # 白紙チェックボックスが未指定ならクリック
                    ImgClick(CFolURL, r"\Houjinzei\HPCheck.png", 0.9, 10)
                    # 両面チェックボックス確認
                    RM = ImgCheck(CFolURL, r"\Houjinzei\Ryoumen.png", 0.9, 10)
                    if RM[0] is True:
                        # 両面チェックボックスが未指定ならクリック
                        ImgClick(CFolURL, r"\Houjinzei\Ryoumen.png", 0.9, 10)
                        # 印刷ボタンクリック
                        ImgClick(CFolURL, r"\Houjinzei\GPPrint.png", 0.9, 10)
                    else:
                        # 両面チェックボックスが指定済の場合
                        # 印刷ボタンクリック
                        ImgClick(CFolURL, r"\Houjinzei\GPPrint.png", 0.9, 10)
                else:
                    # 白紙チェックボックスが指定済の場合
                    # 両面チェックボックス確認
                    RM = ImgCheck(CFolURL, r"\Houjinzei\Ryoumen.png", 0.9, 10)
                    if RM[0] is True:
                        # 両面チェックボックスが未指定ならクリック
                        ImgClick(CFolURL, r"\Houjinzei\Ryoumen.png", 0.9, 10)
                        # 印刷ボタンクリック
                        ImgClick(CFolURL, r"\Houjinzei\GPPrint.png", 0.9, 10)
                    else:
                        # 両面チェックボックスが指定済の場合
                        # 印刷ボタンクリック
                        ImgClick(CFolURL, r"\Houjinzei\GPPrint.png", 0.9, 10)
        # --------------------------------------------------------------------------
        # 申告税一覧表印刷処理----------------------------------------------------
        FO = ImgCheckForList(
            CFolURL,
            [
                r"\Houjinzei\FileOut.png",
                r"\Houjinzei\FileOut2.png",
            ],
            0.9,
            10,
        )
        if FO[0] is True:
            ImgClick(CFolURL, FO[1], 0.9, 10)
        ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
        pg.press("return")
        pg.press("delete")
        pg.press("backspace")
        pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
        pg.hotkey("ctrl", "v")
        pg.press("return")
        time.sleep(1)
        ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
        time.sleep(2)
        # 印刷中が表示されるまで待機---------------------------------
        IC = 0
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\HPCFlag.png", confidence=0.9)
            is not None
        ):
            time.sleep(1)
            IC += 1
            if IC == 5:
                pg.press("tab")
                break
            FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
            if FO[0] is True:
                pg.press("y")
        # --------------------------------------------------------------------
        # 印刷中が表示されなくなるまで待機---------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\NowPrint.png", confidence=0.9)
            is not None
        ):
            time.sleep(1)
        # --------------------------------------------------------------------
        time.sleep(1)
        # 確実に閉じる---------------------------------------------------
        ZME = ImgCheck(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
        if ZME[0] is True:
            ImgClick(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
        else:
            pg.keyDown("alt")
            pg.press("x")
            pg.keyUp("alt")
        # ---------------------------------------------------------------
        time.sleep(3)
        # 確実に閉じる---------------------------------------------------
        ZME = ImgCheck(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
        if ZME[0] is True:
            ImgClick(CFolURL, r"\Houjinzei\MenuEnd.png", 0.9, 10)
        else:
            pg.keyDown("alt")
            pg.press("x")
            pg.keyUp("alt")
        # ---------------------------------------------------------------
        # 終了確認が表示されるまで待機---------------------------------
        while (
            pg.locateOnScreen(CFolURL + r"\Houjinzei\GaikyouEndQ.png", confidence=0.9)
            is None
        ):
            time.sleep(1)
        # --------------------------------------------------------------------
        pg.press("y")
        while (
            ImgCheckForList(
                CFolURL,
                [
                    r"\Houjinzei\02JigyouGaikyou.png",
                    r"\Houjinzei\02JigyouGaikyou2.png",
                ],
                0.9,
                10,
            )[0]
            is False
        ):
            time.sleep(1)
            SEQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ2.png", 0.9, 10)
            if SEQ[0] is True:
                ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ2Btn.png", 0.9, 10)
            SEQQ = ImgCheck(CFolURL, r"\Houjinzei\SinkokuEndQ3.png", 0.9, 10)
            if SEQQ[0] is True:
                pg.press("return")
                # 地方税一覧入力が表示されるまで待機---------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\SinkokuEndQ4.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                ImgClick(CFolURL, r"\Houjinzei\SinkokuEndQ4.png", 0.9, 10)
                # --------------------------------------------------------------------
        # 閉じる処理--------------------------
        pg.keyDown("alt")
        pg.press("f4")
        pg.keyUp("alt")
        # -----------------------------------
        # 法人税フラグが表示されるまで待機-------------------------------------
        while pg.locateOnScreen(TFolURL + r"\HoujinFlag.png", confidence=0.9) is None:
            time.sleep(1)
        # --------------------------------------------------------------------
        # 初期画面で開封された法人税項目を閉じる----------------------------------
        HoujinList = [
            r"\FastMenuHoujinzei.png",
            r"\FastMenuHoujinzei2.png",
        ]
        HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
        if HLI[0] is True:
            ImgClick(TFolURL, HLI[1], 0.9, 10)
        # --------------------------------------------------------------------
        return True, ThisNo, ThisYear, ThisMonth


# ------------------------------------------------------------------------------------------------------------------
def HoujinzeiUpdate(FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname):
    """
    概要: 法人税更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        # 法人税のアイコンを探す-------------------------------------------------
        ImgList = [r"\Houjinzei.png", r"\Houjinzei2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 法人税のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 法人税のアイコンをクリック
            # 法人税フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\HoujinFlag.png", confidence=0.9) is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 他システムとメニューが違う-------------------------------------------------------
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return"])
            # 申告種類が確定申告になっているか確認-----------------------------------
            KF = ImgCheck(CFolURL, r"\Houjinzei\KakuteiFlag.png", 0.9, 10)
            if KF[0] is False:
                while (
                    ImgCheck(CFolURL, r"\Houjinzei\KakuteiFlag.png", 0.9, 10)[0]
                    is False
                ):
                    ImgClick(CFolURL, r"\Houjinzei\SinkokuArrow.png", 0.9, 10)
                    pg.press("down")
                    pg.press("return")
            # -----------------------------------
            if ImgCheck(TFolURL, r"\NotData.png", 0.9, 10)[0] is True:
                # 入力した関与先コードを取得------------
                pg.press(["return", "return"])
                pg.keyDown("shift")
                pg.press(["tab", "tab", "tab", "tab"])
                pg.keyUp("shift")
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisNo = pyperclip.paste()
                pg.press("return")
                # クリップボードをクリア----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                # ------------------------------------
                # 表示された年度を取得-----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisYear = pyperclip.paste()
                # -----------------------------------
                pg.press("return")
                # 表示された申告種類を取得---------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisMonth = pyperclip.paste()
                pg.press("return")
                # -----------------------------------
            else:
                # 入力した関与先コードを取得------------
                pg.press(["return", "return"])
                pg.keyDown("shift")
                pg.press(["tab", "tab", "tab", "tab"])
                pg.keyUp("shift")
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisNo = pyperclip.paste()
                pg.press("return")
                # 表示された年度を取得-----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisYear = pyperclip.paste()
                # -----------------------------------
                pg.press("return")
                # 表示された申告種類を取得---------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisMonth = pyperclip.paste()
                pg.press("return")
                # -----------------------------------
            # 他システムとメニューが違う-------------------------------------------------------
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 法人税メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\HoujinzeiMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    HQ = ImgCheck(TFolURL, r"\HoujinOpenQ.png", 0.9, 10)
                    if HQ[0] is True:
                        ImgClick(TFolURL, r"\HoujinOpenQCansel.png", 0.9, 10)
                    # 顧問先情報更新ダイアログ確認------------------------------------------
                    KK = ImgCheck(TFolURL, r"\KomonKoushin.png", 0.9, 10)
                    if KK[0] is True:
                        pg.press("a")
                    # 新規別表追加選択-------------------------------------------------
                    SB = ImgCheck(CFolURL, r"\Houjinzei\SaiyouBeppyou.png", 0.9, 10)
                    if SB[0] is True:
                        ImgClick(CFolURL, r"\Houjinzei\SBKousin.png", 0.9, 10)
                # 指示内容で処理分け----------------------------------------------------------
                if (
                    PN == "申告税一覧表"
                    or PN == "第6号様式（県）"
                    or PN == "第6号様式別表9（県）"
                    or PN == "第20号様式（市）"
                    or PN == "第22号の2様式"
                    or PN == "別表１　緑色"
                ):
                    PNFlow = HoujinzeiUpdateSinkokuItiran(
                        FolURL,
                        TFolURL,
                        CFolURL,
                        ExRow,
                        driver,
                        PN,
                        Fname,
                        ThisNo,
                        ThisYear,
                        ThisMonth,
                    )
                    return PNFlow[0], PNFlow[1], PNFlow[2], PNFlow[3]
                elif PN == "税務代理権限書":
                    PNFlow = HoujinzeiUpdateZeimuDairi(
                        FolURL,
                        TFolURL,
                        CFolURL,
                        ExRow,
                        driver,
                        PN,
                        Fname,
                        ThisNo,
                        ThisYear,
                        ThisMonth,
                    )
                    return PNFlow[0], PNFlow[1], PNFlow[2], PNFlow[3]
                elif PN == "書面添付　法人税":
                    PNFlow = HoujinzeiUpdateSyomen(
                        FolURL,
                        TFolURL,
                        CFolURL,
                        ExRow,
                        driver,
                        PN,
                        Fname,
                        ThisNo,
                        ThisYear,
                        ThisMonth,
                    )
                    return PNFlow[0], PNFlow[1], PNFlow[2], PNFlow[3]
                elif PN == "別表2-16":
                    PNFlow = HoujinzeiUpdateBeppyou(
                        FolURL,
                        TFolURL,
                        CFolURL,
                        ExRow,
                        driver,
                        PN,
                        Fname,
                        ThisNo,
                        ThisYear,
                        ThisMonth,
                    )
                    return PNFlow[0], PNFlow[1], PNFlow[2], PNFlow[3]
                elif PN == "法人事業概況説明書":
                    PNFlow = HoujinzeiUpdateGaikyou(
                        FolURL,
                        TFolURL,
                        CFolURL,
                        ExRow,
                        driver,
                        PN,
                        Fname,
                        ThisNo,
                        ThisYear,
                        ThisMonth,
                    )
                    return PNFlow[0], PNFlow[1], PNFlow[2], PNFlow[3]
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "法人税起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def GenkasyoukyakuUpdate(FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname):
    """
    概要: 減価償却更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        # 減価償却のアイコンを探す-------------------------------------------------
        ImgList = [r"\G_Syoukyaku.png", r"\G_Syoukyaku2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 減価償却のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 減価償却のアイコンをクリック
            # 減価償却フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\G_SyoukyakuFlag.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return", "return"])
            # 入力した関与先コードを取得------------
            pg.keyDown("shift")
            pg.press(["tab", "tab", "tab"])
            pg.keyUp("shift")
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisNo = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された年度を取得-----------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisYear = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された月を取得-------------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisMonth = pyperclip.paste()
            # -----------------------------------
            time.sleep(1)
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 減価償却メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\G_SyoukyakuMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    # アップデート情報画面が出たら閉じる-------------------------------
                    GSUM = ImgCheck(TFolURL, r"\G_SyoukyakuUpMsg.png", 0.9, 10)
                    if GSUM[0] is True:
                        ImgClick(TFolURL, r"\G_SyoukyakuUpMsgCansel.png", 0.9, 10)
                    # 顧問先情報更新ダイアログ確認------------------------------------------
                    KK = ImgCheck(TFolURL, r"\KomonKoushin.png", 0.9, 10)
                    if KK[0] is True:
                        pg.press("a")
                # --------------------------------------------------------------------
                ImgClick(
                    CFolURL, r"\GenkaSyoukyaku\G_Insatu.png", 0.9, 10
                )  # 2.印刷処理アイコンをクリック
                # 印刷処理メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\GenkaSyoukyaku\G_InsatuFlag.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                if PN == "固定資産台帳":
                    GIF = ImgCheckForList(
                        CFolURL,
                        [
                            r"\GenkaSyoukyaku\01G_Uchiwake.png",
                            r"\GenkaSyoukyaku\01G_Uchiwake2.png",
                        ],
                        0.9,
                        10,
                    )
                    if GIF[0] is True:
                        ImgClick(CFolURL, GIF[1], 0.9, 10)
                    # 出力条件ウィンドウが表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\GenkaSyoukyaku\01G_PrintWait.png",
                            confidence=0.9,
                        )
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    time.sleep(1)
                    ImgClick(
                        CFolURL, r"\GenkaSyoukyaku\01G_PrintOK.png", 0.9, 10
                    )  # 出力条件設定OKをクリック
                    # 出力条件ウィンドウが表示されるなくなるまで待機--------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\GenkaSyoukyaku\01G_PrintWait.png",
                            confidence=0.9,
                        )
                        is not None
                    ):
                        time.sleep(1)
                    time.sleep(3)
                    # --------------------------------------------------------------------
                    ImgClick(
                        CFolURL, r"\GenkaSyoukyaku\01G_PrintBtn.png", 0.9, 10
                    )  # 印刷ボタンをクリック
                    # 印刷設定が表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    # 申告税一覧表印刷処理----------------------------------------------------
                    FO = ImgCheckForList(
                        CFolURL,
                        [
                            r"\Houjinzei\FileOut.png",
                            r"\Houjinzei\FileOut2.png",
                        ],
                        0.9,
                        10,
                    )
                    if FO[0] is True:
                        ImgClick(CFolURL, FO[1], 0.9, 10)
                    ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
                    pg.press("return")
                    pg.press("delete")
                    pg.press("backspace")
                    pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                    pg.hotkey("ctrl", "v")
                    pg.press("return")
                    time.sleep(1)
                    ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
                    time.sleep(2)
                    #  印刷設定が表示されなくなるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                        )
                        is not None
                    ):
                        time.sleep(1)
                        FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
                        if FO[0] is True:
                            pg.press("y")
                    # --------------------------------------------------------------------
                    time.sleep(1)
                    # 確実に閉じる---------------------------------------------------
                    ZME = ImgCheck(CFolURL, r"\GenkaSyoukyaku\01G_End.png", 0.9, 10)
                    if ZME[0] is True:
                        ImgClick(CFolURL, r"\GenkaSyoukyaku\01G_End.png", 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("x")
                        pg.keyUp("alt")
                    # ---------------------------------------------------------------
                    time.sleep(1)
                    while (
                        ImgCheckForList(
                            CFolURL,
                            [
                                r"\GenkaSyoukyaku\01G_Uchiwake.png",
                                r"\GenkaSyoukyaku\01G_Uchiwake2.png",
                            ],
                            0.9,
                            10,
                        )[0]
                        is False
                    ):
                        time.sleep(1)
                    # 閉じる処理--------------------------
                    pg.keyDown("alt")
                    pg.press("f4")
                    pg.keyUp("alt")
                    # -----------------------------------
                    # 減価償却フラグが表示されるまで待機------------------------------------
                    while (
                        pg.locateOnScreen(
                            TFolURL + r"\G_SyoukyakuFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ------------------------------------------------------------------
                    return True, ThisNo, ThisYear, ThisMonth
                elif PN == "一括償却資産":
                    Nod = ""
                    GIF = ImgCheckForList(
                        CFolURL,
                        [
                            r"\GenkaSyoukyaku\03G_Meisai.png",
                            r"\GenkaSyoukyaku\03G_Meisai2.png",
                        ],
                        0.9,
                        10,
                    )
                    if GIF[0] is True:
                        ImgClick(CFolURL, GIF[1], 0.9, 10)
                    # 出力条件ウィンドウが表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\GenkaSyoukyaku\03G_PrintWait.png",
                            confidence=0.9,
                        )
                        is None
                    ):
                        time.sleep(1)
                        GN = ImgCheck(
                            CFolURL, r"\GenkaSyoukyaku\03G_Nodata.png", 0.9, 10
                        )
                        if GN[0] is True:
                            pg.press("return")
                            Nod = "Nodata"
                        if Nod == "Nodata":
                            break
                    # --------------------------------------------------------------------
                    time.sleep(1)
                    if Nod == "":
                        ImgClick(
                            CFolURL, r"\GenkaSyoukyaku\03G_PrintOK.png", 0.9, 10
                        )  # 出力条件設定OKをクリック
                        # 出力条件ウィンドウが表示されるなくなるまで待機--------------------------
                        while (
                            pg.locateOnScreen(
                                CFolURL + r"\GenkaSyoukyaku\03G_PrintWait.png",
                                confidence=0.9,
                            )
                            is not None
                        ):
                            time.sleep(1)
                        time.sleep(3)
                        # --------------------------------------------------------------------
                        ImgClick(
                            CFolURL, r"\GenkaSyoukyaku\01G_PrintBtn.png", 0.9, 10
                        )  # 印刷ボタンをクリック
                        # 印刷設定が表示されるまで待機---------------------------------
                        while (
                            pg.locateOnScreen(
                                CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                            SJS = ImgCheck(CFolURL, r"\Houjinzei\Nodata.png", 0.9, 10)
                            if SJS[0] is True:
                                pg.press("return")
                                ImgClick(CFolURL, r"\Houjinzei\NodataCan.png", 0.9, 10)
                                Nod = "Nodata"
                            if Nod == "Nodata":
                                break
                        # --------------------------------------------------------------------
                    if Nod == "":
                        # 申告税一覧表印刷処理----------------------------------------------------
                        FO = ImgCheckForList(
                            CFolURL,
                            [
                                r"\Houjinzei\FileOut.png",
                                r"\Houjinzei\FileOut2.png",
                            ],
                            0.9,
                            10,
                        )
                        if FO[0] is True:
                            ImgClick(CFolURL, FO[1], 0.9, 10)
                        ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
                        pg.press("return")
                        pg.press("delete")
                        pg.press("backspace")
                        pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                        pg.hotkey("ctrl", "v")
                        pg.press("return")
                        time.sleep(1)
                        ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
                        time.sleep(2)
                        #  印刷設定が表示されなくなるまで待機---------------------------------
                        while (
                            pg.locateOnScreen(
                                CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                            )
                            is not None
                        ):
                            time.sleep(1)
                            FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
                            if FO[0] is True:
                                pg.press("y")
                        # --------------------------------------------------------------------
                        time.sleep(1)
                        # 確実に閉じる---------------------------------------------------
                        ZME = ImgCheck(CFolURL, r"\GenkaSyoukyaku\01G_End.png", 0.9, 10)
                        if ZME[0] is True:
                            ImgClick(CFolURL, r"\GenkaSyoukyaku\01G_End.png", 0.9, 10)
                        else:
                            pg.keyDown("alt")
                            pg.press("x")
                            pg.keyUp("alt")
                        # ---------------------------------------------------------------
                        time.sleep(1)
                        while (
                            ImgCheckForList(
                                CFolURL,
                                [
                                    r"\GenkaSyoukyaku\01G_Uchiwake.png",
                                    r"\GenkaSyoukyaku\01G_Uchiwake2.png",
                                ],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        # 閉じる処理--------------------------
                        pg.keyDown("alt")
                        pg.press("f4")
                        pg.keyUp("alt")
                        # -----------------------------------
                        # 減価償却フラグが表示されるまで待機------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\G_SyoukyakuFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # ------------------------------------------------------------------
                        return True, ThisNo, ThisYear, ThisMonth
                    else:
                        while (
                            ImgCheckForList(
                                CFolURL,
                                [
                                    r"\GenkaSyoukyaku\01G_Uchiwake.png",
                                    r"\GenkaSyoukyaku\01G_Uchiwake2.png",
                                ],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        # 閉じる処理--------------------------
                        pg.keyDown("alt")
                        pg.press("f4")
                        pg.keyUp("alt")
                        # -----------------------------------
                        # 減価償却フラグが表示されるまで待機------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\G_SyoukyakuFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # ------------------------------------------------------------------
                        return False, ThisNo, ThisYear, Nod
                elif PN == "少額償却資産":
                    Nod = ""
                    GIF = ImgCheckForList(
                        CFolURL,
                        [
                            r"\GenkaSyoukyaku\04G_Syougaku.png",
                            r"\GenkaSyoukyaku\04G_Syougaku2.png",
                        ],
                        0.9,
                        10,
                    )
                    if GIF[0] is True:
                        ImgClick(CFolURL, GIF[1], 0.9, 10)
                    # 出力条件ウィンドウが表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\GenkaSyoukyaku\04G_PrintWait.png",
                            confidence=0.9,
                        )
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    time.sleep(1)
                    ImgClick(
                        CFolURL, r"\GenkaSyoukyaku\04G_PrintOK.png", 0.9, 10
                    )  # 出力条件設定OKをクリック
                    # 出力条件ウィンドウが表示されるなくなるまで待機--------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\GenkaSyoukyaku\04G_PrintWait.png",
                            confidence=0.9,
                        )
                        is not None
                    ):
                        time.sleep(1)
                    time.sleep(3)
                    GN = ImgCheck(CFolURL, r"\GenkaSyoukyaku\01G_Nodata.png", 0.9, 10)
                    if GN[0] is True:
                        pg.press("return")
                        Nod = "Nodata"
                    if Nod == "":
                        # --------------------------------------------------------------------
                        ImgClick(
                            CFolURL, r"\GenkaSyoukyaku\01G_PrintBtn.png", 0.9, 10
                        )  # 印刷ボタンをクリック
                        # 印刷設定が表示されるまで待機---------------------------------
                        while (
                            pg.locateOnScreen(
                                CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                            SJS = ImgCheck(CFolURL, r"\Houjinzei\Nodata.png", 0.9, 10)
                            if SJS[0] is True:
                                pg.press("return")
                                ImgClick(CFolURL, r"\Houjinzei\NodataCan.png", 0.9, 10)
                                Nod = "Nodata"
                            if Nod == "Nodata":
                                break
                        # --------------------------------------------------------------------
                    else:
                        ImgClick(CFolURL, r"\GenkaSyoukyaku\03G_Cancel.png", 0.9, 10)
                    if Nod == "":
                        # 申告税一覧表印刷処理----------------------------------------------------
                        FO = ImgCheckForList(
                            CFolURL,
                            [
                                r"\Houjinzei\FileOut.png",
                                r"\Houjinzei\FileOut2.png",
                            ],
                            0.9,
                            10,
                        )
                        if FO[0] is True:
                            ImgClick(CFolURL, FO[1], 0.9, 10)
                        ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
                        pg.press("return")
                        pg.press("delete")
                        pg.press("backspace")
                        pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                        pg.hotkey("ctrl", "v")
                        pg.press("return")
                        time.sleep(1)
                        ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
                        time.sleep(2)
                        #  印刷設定が表示されなくなるまで待機---------------------------------
                        while (
                            pg.locateOnScreen(
                                CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                            )
                            is not None
                        ):
                            time.sleep(1)
                            FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
                            if FO[0] is True:
                                pg.press("y")
                        # --------------------------------------------------------------------
                        time.sleep(1)
                        # 確実に閉じる---------------------------------------------------
                        ZME = ImgCheck(CFolURL, r"\GenkaSyoukyaku\01G_End.png", 0.9, 10)
                        if ZME[0] is True:
                            ImgClick(CFolURL, r"\GenkaSyoukyaku\01G_End.png", 0.9, 10)
                        else:
                            pg.keyDown("alt")
                            pg.press("x")
                            pg.keyUp("alt")
                        # ---------------------------------------------------------------
                        time.sleep(1)
                        while (
                            ImgCheckForList(
                                CFolURL,
                                [
                                    r"\GenkaSyoukyaku\01G_Uchiwake.png",
                                    r"\GenkaSyoukyaku\01G_Uchiwake2.png",
                                ],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        # 閉じる処理--------------------------
                        pg.keyDown("alt")
                        pg.press("f4")
                        pg.keyUp("alt")
                        # -----------------------------------
                        # 減価償却フラグが表示されるまで待機------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\G_SyoukyakuFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # ------------------------------------------------------------------
                        return True, ThisNo, ThisYear, ThisMonth
                    else:
                        while (
                            ImgCheckForList(
                                CFolURL,
                                [
                                    r"\GenkaSyoukyaku\01G_Uchiwake.png",
                                    r"\GenkaSyoukyaku\01G_Uchiwake2.png",
                                ],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        # 閉じる処理--------------------------
                        pg.keyDown("alt")
                        pg.press("f4")
                        pg.keyUp("alt")
                        # -----------------------------------
                        # 減価償却フラグが表示されるまで待機------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\G_SyoukyakuFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # ------------------------------------------------------------------
                        return False, ThisNo, ThisYear, Nod
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "減価償却起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def KessanUpDate(FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname):
    """
    概要: 決算内訳書更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        # 決算内訳書のアイコンを探す-------------------------------------------------
        ImgList = [r"\K_Uchiwake.png", r"\K_Uchiwake2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 決算内訳書のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 決算内訳書のアイコンをクリック
            # 決算内訳書フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\Kessan_CFlag.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return", "return"])
            # 入力した関与先コードを取得------------
            pg.keyDown("shift")
            pg.press(["tab", "tab", "tab"])
            pg.keyUp("shift")
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisNo = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された年度を取得-----------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisYear = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された月を取得-------------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisMonth = pyperclip.paste()
            # -----------------------------------
            time.sleep(1)
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 決算内訳書メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\KessanMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    # 顧問先情報更新ダイアログ確認------------------------------------------
                    KK = ImgCheck(TFolURL, r"\KomonKoushin.png", 0.9, 10)
                    if KK[0] is True:
                        pg.press("a")
                # --------------------------------------------------------------------
                # 内訳書印刷メニューが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Uchiwake\11U_Flag.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                # 内訳書印刷のアイコンを探す-------------------------------------------------
                ImgList = [
                    r"\Uchiwake\11U_Uchiwake.png",
                    r"\Uchiwake\11U_Uchiwake2.png",
                ]
                ICFL = ImgCheckForList(CFolURL, ImgList, 0.9, 10)
                # -----------------------------------------------------------------------
                if ICFL[0] is True:
                    ImgClick(CFolURL, ICFL[1], 0.9, 10)  #  内訳書印刷アイコンをクリック
                # 印刷ボタンが表示されるまで待機---------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Uchiwake\11U_PrintBtn.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                ImgClick(CFolURL, r"\Uchiwake\11U_PrintBtn.png", 0.9, 10)  # 印刷ボタンをクリック
                # 印刷設定が表示されるまで待機---------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                # 申告税一覧表印刷処理----------------------------------------------------
                FO = ImgCheckForList(
                    CFolURL,
                    [
                        r"\Houjinzei\FileOut.png",
                        r"\Houjinzei\FileOut2.png",
                    ],
                    0.9,
                    10,
                )
                if FO[0] is True:
                    ImgClick(CFolURL, FO[1], 0.9, 10)
                ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
                pg.press("return")
                pg.press("delete")
                pg.press("backspace")
                pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                pg.hotkey("ctrl", "v")
                pg.press("return")
                time.sleep(1)
                ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
                time.sleep(2)
                #  印刷設定が表示されなくなるまで待機---------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                    )
                    is not None
                ):
                    time.sleep(1)
                    FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
                    if FO[0] is True:
                        pg.press("y")
                # --------------------------------------------------------------------
                time.sleep(1)
                #  印刷中が表示されるまで待機-------------------------------------------
                IC = 0
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\NowPrint.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                    IC += 1
                    if IC == 5:
                        pg.press("tab")
                        break
                #  印刷中が表示されなくなるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(
                        CFolURL + r"\Houjinzei\NowPrint.png", confidence=0.9
                    )
                    is not None
                ):
                    time.sleep(1)
                #  確実に閉じる--------------------------------------------------------
                UED = ImgCheck(CFolURL, r"\Uchiwake\11U_End.png", 0.9, 10)
                if UED[0] is True:
                    ImgClick(CFolURL, r"\Uchiwake\11U_End.png", 0.9, 10)
                else:
                    pg.keyDown("alt")
                    pg.press("x")
                    pg.keyUp("alt")
                # --------------------------------------------------------------------
                time.sleep(1)
                while (
                    ImgCheckForList(
                        CFolURL,
                        [
                            r"\Uchiwake\11U_Uchiwake.png",
                            r"\Uchiwake\11U_Uchiwake2.png",
                        ],
                        0.9,
                        10,
                    )[0]
                    is False
                ):
                    time.sleep(1)
                # 閉じる処理--------------------------
                pg.keyDown("alt")
                pg.press("f4")
                pg.keyUp("alt")
                # -----------------------------------
                # 決算内訳書フラグが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\Kessan_CFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # ------------------------------------------------------------------
                return True, ThisNo, ThisYear, ThisMonth
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "決算内訳書起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def KaikeiUpDate(FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname):
    """
    概要: 会計大将更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        # 会計大将のアイコンを探す-------------------------------------------------
        ImgList = [r"\K_TaisyouIcon.png", r"\K_TaisyouIcon2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 会計大将のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 会計大将のアイコンをクリック
            # 会計大将フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\Kaikei_CFlag.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return", "return"])
            # 入力した関与先コードを取得------------
            pg.keyDown("shift")
            pg.press(["tab", "tab", "tab"])
            pg.keyUp("shift")
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisNo = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された年度を取得-----------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisYear = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された月を取得-------------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisMonth = pyperclip.paste()
            # -----------------------------------
            time.sleep(1)
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 会計大将メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\K_TaisyouMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    # 顧問先情報更新ダイアログ確認------------------------------------------
                    KK = ImgCheck(TFolURL, r"\KomonKoushin.png", 0.9, 10)
                    if KK[0] is True:
                        pg.press("a")
                # --------------------------------------------------------------------
                # 指示内容で処理分け----------------------------------------------------------
                if PN == "消費税確定申告書":
                    ImgClick(
                        CFolURL, r"\KTaisyou\KessanSinkoku.png", 0.9, 10
                    )  # 決算申告書アイコンをクリック
                    # 決算申告書が表示されるまで待機----------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KessanFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # サイドメニューより消費税申告書を確認------------------------------------------
                    SSM = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\Syouhizei.png", r"\KTaisyou\Syouhizei2.png"],
                        0.9,
                        10,
                    )
                    if SSM[0] is True:  # 消費税申告書があれば
                        ImgClick(CFolURL, SSM[1], 0.9, 10)  # 消費税申告書アイコンをクリック
                        time.sleep(2)
                    # 申告書・付表入力を確認------------------------------------------------
                    SFI = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\SFIcon.png", r"\KTaisyou\SFIcon2.png"],
                        0.9,
                        10,
                    )
                    if SFI[0] is True:  # 申告書・付表入力があれば
                        ImgClick(CFolURL, SFI[1], 0.9, 10)  # 申告書・付表入力アイコンをクリック
                    # 申告選択ウィンドウが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\SinkokuSentaku.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                        # 申告回数設定ウィンドウがひょうじされたら
                        SCQ = ImgCheck(CFolURL, r"\KTaisyou\ScountQ.png", 0.9, 10)
                        if SCQ[0] is True:
                            pg.press("n")  # nでいいえ
                    # 申告種で処理分け----------------------------------------------------------------------------------
                    SL = ImgCheck(
                        CFolURL, r"\KTaisyou\SinkokuList.png", 0.9, 10
                    )  # 申告種類選択ボックスが開かれてるか確認
                    if SL[0] is True:
                        pg.press("return")
                        # if ExRow["会計大将_消費税確定申告書"] == "1":  # Excel数値が確定申告なら
                        #     while (
                        #         pg.locateOnScreen(
                        #             CFolURL + r"\KTaisyou\Kakutei.png", confidence=0.9
                        #         )
                        #         is None
                        #     ):
                        #         pg.press("down")
                        #     pg.press("return")
                        # elif ExRow["会計大将_消費税確定申告書"] == "2":  # Excel数値が修正確定なら
                        #     while (
                        #         pg.locateOnScreen(
                        #             CFolURL + r"\KTaisyou\Syuusei.png", confidence=0.9
                        #         )
                        #         is None
                        #     ):
                        #         pg.press("down")
                        #     pg.press("return")
                        # elif ExRow["会計大将_消費税確定申告書"] == "3":  # Excel数値が中間申告なら
                        #     while (
                        #         pg.locateOnScreen(
                        #             CFolURL + r"\KTaisyou\Cyuukan.png", confidence=0.9
                        #         )
                        #         is None
                        #     ):
                        #         pg.press("down")
                        #     pg.press("return")
                        # elif ExRow["会計大将_消費税確定申告書"] == "4":  # Excel数値が修正中間なら
                        #     while (
                        #         pg.locateOnScreen(
                        #             CFolURL + r"\KTaisyou\SCyuukan.png", confidence=0.9
                        #         )
                        #         is None
                        #     ):
                        #         pg.press("down")
                        #     pg.press("return")
                        # elif ExRow["会計大将_消費税確定申告書"] == "5":  # Excel数値が予定申告なら
                        #     while (
                        #         pg.locateOnScreen(
                        #             CFolURL + r"\KTaisyou\Yotei.png", confidence=0.9
                        #         )
                        #         is None
                        #     ):
                        #         pg.press("down")
                        #     pg.press("return")
                    else:
                        pg.press("return")
                        # SA = ImgCheck(
                        #     CFolURL, r"\KTaisyou\SinkokuArrow.png", 0.9, 10
                        # )  # 申告区分ドロップボックスが開けるか確認
                        # if SA[0] is True:
                        #     ImgClick(
                        #         CFolURL, r"\KTaisyou\SinkokuArrow.png", 0.9, 10
                        #     )  # 一括更新開始のアイコンをクリック
                        #     if ExRow["会計大将_消費税確定申告書"] == "1":  # Excel数値が確定申告なら
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\Kakutei.png",
                        #                 confidence=0.9,
                        #             )
                        #             is None
                        #         ):
                        #             pg.press("down")
                        #         pg.press("return")
                        #     elif ExRow["会計大将_消費税確定申告書"] == "2":  # Excel数値が修正確定なら
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\Syuusei.png",
                        #                 confidence=0.9,
                        #             )
                        #             is None
                        #         ):
                        #             pg.press("down")
                        #         pg.press("return")
                        #     elif ExRow["会計大将_消費税確定申告書"] == "3":  # Excel数値が中間申告なら
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\Cyuukan.png",
                        #                 confidence=0.9,
                        #             )
                        #             is None
                        #         ):
                        #             pg.press("down")
                        #         pg.press("return")
                        #     elif ExRow["会計大将_消費税確定申告書"] == "4":  # Excel数値が修正中間なら
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\SCyuukan.png",
                        #                 confidence=0.9,
                        #             )
                        #             is None
                        #         ):
                        #             pg.press("down")
                        #         pg.press("return")
                        #     elif ExRow["会計大将_消費税確定申告書"] == "5":  # Excel数値が予定申告なら
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\Yotei.png", confidence=0.9
                        #             )
                        #             is None
                        #         ):
                        #             pg.press("down")
                        #         pg.press("return")
                        # else:
                        #     # 現在の申告回数を取得
                        #     pg.hotkey("ctrl", "c")
                        #     SinNo = pyperclip.paste()
                        #     NextSinNo = int(SinNo) + 1
                        #     pg.write(str(NextSinNo))
                        #     pg.press("return")
                        #     if ExRow["会計大将_消費税確定申告書"] == "1":  # Excel数値が確定申告なら
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\Kakutei.png",
                        #                 confidence=0.9,
                        #             )
                        #             is None
                        #         ):
                        #             pg.press("down")
                        #             # 申告種選択リスト最後尾か調べる--------------------------------
                        #             LA = ImgCheck(
                        #                 CFolURL, r"\KTaisyou\SinkokuList.png", 0.9, 10
                        #             )
                        #             if LA[0] is True:
                        #                 pg.press("home")
                        #             # -------------------------------------------------------------
                        #         pg.press("return")
                        #     elif ExRow["会計大将_消費税確定申告書"] == "2":  # Excel数値が修正確定なら
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\Syuusei.png",
                        #                 confidence=0.9,
                        #             )
                        #             is None
                        #         ):
                        #             pg.press("down")
                        #             # 申告種選択リスト最後尾か調べる--------------------------------
                        #             LA = ImgCheck(
                        #                 CFolURL, r"\KTaisyou\SinkokuList.png", 0.9, 10
                        #             )
                        #             if LA[0] is True:
                        #                 pg.press("home")
                        #             # -------------------------------------------------------------
                        #         pg.press("return")
                        #     elif ExRow["会計大将_消費税確定申告書"] == "3":  # Excel数値が中間申告なら
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\Cyuukan.png",
                        #                 confidence=0.9,
                        #             )
                        #             is None
                        #         ):
                        #             pg.press("down")
                        #             # 申告種選択リスト最後尾か調べる--------------------------------
                        #             LA = ImgCheck(
                        #                 CFolURL, r"\KTaisyou\SinkokuList.png", 0.9, 10
                        #             )
                        #             if LA[0] is True:
                        #                 pg.press("home")
                        #             # -------------------------------------------------------------
                        #         pg.press("return")
                        #     elif ExRow["会計大将_消費税確定申告書"] == "4":  # Excel数値が修正中間なら
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\SCyuukan.png",
                        #                 confidence=0.9,
                        #             )
                        #             is None
                        #         ):
                        #             pg.press("down")
                        #             # 申告種選択リスト最後尾か調べる--------------------------------
                        #             LA = ImgCheck(
                        #                 CFolURL, r"\KTaisyou\SinkokuList.png", 0.9, 10
                        #             )
                        #             if LA[0] is True:
                        #                 pg.press("home")
                        #             # -------------------------------------------------------------
                        #         pg.press("return")
                        #     elif ExRow["会計大将_消費税確定申告書"] == "5":  # Excel数値が予定申告なら
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\Yotei.png", confidence=0.9
                        #             )
                        #             is None
                        #         ):
                        #             pg.press("down")
                        #             # 申告種選択リスト最後尾か調べる--------------------------------
                        #             LA = ImgCheck(
                        #                 CFolURL, r"\KTaisyou\SinkokuList.png", 0.9, 10
                        #             )
                        #             if LA[0] is True:
                        #                 pg.press("home")
                        #             # -------------------------------------------------------------
                        #         pg.press("return")
                        #     time.sleep(1)
                        #     ImgClick(
                        #         CFolURL, r"\KTaisyou\SinkokuOK.png", 0.9, 10
                        #     )  # 申告書選択OKボタンクリック
                        #     time.sleep(1)
                        #     # 申告回数変更を求められた場合--------------------------------------------
                        #     SCQ = ImgCheck(CFolURL, r"\KTaisyou\S_CountQ.png", 0.9, 10)
                        #     if SCQ[0] is True:
                        #         pg.press("N")  # 申告回数変更しない
                        #         time.sleep(1)
                        #         pg.write(str(SinNo))  # 取得していた現在申告区分Noを入力
                        #         pg.press("return")
                        #         ImgClick(CFolURL, r"\KTaisyou\S_Delete.png", 0.9, 10)
                        #         # 申告データ削除ダイアログ表示まで待機----------------------------------
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\S_DeleteQ.png",
                        #                 confidence=0.9,
                        #             )
                        #             is None
                        #         ):
                        #             time.sleep(1)
                        #         pg.press("y")
                        #         # ---------------------------------------------------------------------
                        #         # 申告データ削除完了まで待機----------------------------------
                        #         while (
                        #             pg.locateOnScreen(
                        #                 CFolURL + r"\KTaisyou\S_DeleteQ2.png",
                        #                 confidence=0.9,
                        #             )
                        #             is None
                        #         ):
                        #             time.sleep(1)
                        #         pg.press("return")
                        #         # ---------------------------------------------------------------------
                        #         time.sleep(1)
                        #         pg.press("return")
                        #         time.sleep(1)
                        #         pg.press("home")
                        #         time.sleep(1)
                        #         SL = ImgCheck(
                        #             CFolURL, r"\KTaisyou\SinkokuList.png", 0.9, 10
                        #         )  # 申告種類選択ボックスが開かれてるか確認
                        #         if SL[0] is True:
                        #             if ExRow["会計大将_消費税確定申告書"] == "1":  # Excel数値が確定申告なら
                        #                 while (
                        #                     pg.locateOnScreen(
                        #                         CFolURL + r"\KTaisyou\Kakutei.png",
                        #                         confidence=0.9,
                        #                     )
                        #                     is None
                        #                 ):
                        #                     pg.press("down")
                        #                 pg.press("return")
                        #             elif (
                        #                 ExRow["会計大将_消費税確定申告書"] == "2"
                        #             ):  # Excel数値が修正確定なら
                        #                 while (
                        #                     pg.locateOnScreen(
                        #                         CFolURL + r"\KTaisyou\Syuusei.png",
                        #                         confidence=0.9,
                        #                     )
                        #                     is None
                        #                 ):
                        #                     pg.press("down")
                        #                 pg.press("return")
                        #             elif (
                        #                 ExRow["会計大将_消費税確定申告書"] == "3"
                        #             ):  # Excel数値が中間申告なら
                        #                 while (
                        #                     pg.locateOnScreen(
                        #                         CFolURL + r"\KTaisyou\Cyuukan.png",
                        #                         confidence=0.9,
                        #                     )
                        #                     is None
                        #                 ):
                        #                     pg.press("down")
                        #                 pg.press("return")
                        #             elif (
                        #                 ExRow["会計大将_消費税確定申告書"] == "4"
                        #             ):  # Excel数値が修正中間なら
                        #                 while (
                        #                     pg.locateOnScreen(
                        #                         CFolURL + r"\KTaisyou\SCyuukan.png",
                        #                         confidence=0.9,
                        #                     )
                        #                     is None
                        #                 ):
                        #                     pg.press("down")
                        #                 pg.press("return")
                        #             elif (
                        #                 ExRow["会計大将_消費税確定申告書"] == "5"
                        #             ):  # Excel数値が予定申告なら
                        #                 while (
                        #                     pg.locateOnScreen(
                        #                         CFolURL + r"\KTaisyou\Yotei.png",
                        #                         confidence=0.9,
                        #                     )
                        #                     is None
                        #                 ):
                        #                     pg.press("down")
                        #                 pg.press("return")
                    # ----------------------------------------------------------------------------------------------------
                    time.sleep(1)
                    # OKボタンにフォーカスするまでエンター押下-------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\SinkokuAllOK.png", confidence=0.9
                        )
                        is None
                    ):
                        pg.press("return")
                    # ---------------------------------------------------------------------
                    pg.press("return")  # 確定
                    # 確認ウィンドウが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\SinkokuWin.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                        SIQ = ImgCheck(CFolURL, r"\KTaisyou\SiwakeInputQ.png", 0.9, 10)
                        if SIQ[0] is True:
                            pg.press("y")
                        # 再計算確認ウィンドウがあるか確認
                        RCQ = ImgCheck(CFolURL, r"\KTaisyou\ReCalcQ.png", 0.9, 10)
                        if RCQ[0] is True:
                            pg.press("n")
                        # 消費税不明取引確認ウィンドウがあるか確認
                        KN = ImgCheck(CFolURL, r"\KTaisyou\S_Humei.png", 0.9, 10)
                        if KN[0] is True:
                            pg.press("y")
                    # --------------------------------------------------------------------
                    time.sleep(1)
                    SKQ = ImgCheck(CFolURL, r"\KTaisyou\KakuninQ.png", 0.9, 10)
                    if SKQ[0] is True:
                        pg.press("return")
                    time.sleep(1)
                    ImgClick(CFolURL, r"\KTaisyou\SyouhiPrint.png", 0.9, 10)
                    # 確認ウィンドウが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\S_Huhyou.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                        # 再計算確認ウィンドウがあるか確認
                        RCQ = ImgCheck(CFolURL, r"\KTaisyou\ReCalcQ.png", 0.9, 10)
                        if RCQ[0] is True:
                            pg.press("n")
                        # 消費税不明取引確認ウィンドウがあるか確認
                        KN = ImgCheck(CFolURL, r"\KTaisyou\S_Humei.png", 0.9, 10)
                        if KN[0] is True:
                            pg.press("y")
                    # --------------------------------------------------------------------
                    pg.press("p")  # 決定
                    time.sleep(1)
                    # 法人番号未登録確認が表示されているか確認
                    SAL = ImgCheck(CFolURL, r"\KTaisyou\S_Alert.png", 0.9, 10)
                    if SAL[0] is True:  # 法人番号未登録確認が表示された場合
                        pg.press("y")  # yで確定
                    # 一覧表出力項目指定が表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    # 申告税一覧表印刷処理----------------------------------------------------
                    FO = ImgCheckForList(
                        CFolURL,
                        [
                            r"\Houjinzei\FileOut.png",
                            r"\Houjinzei\FileOut2.png",
                        ],
                        0.9,
                        10,
                    )
                    if FO[0] is True:
                        ImgClick(CFolURL, FO[1], 0.9, 10)
                    ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
                    pg.press("return")
                    pg.press("delete")
                    pg.press("backspace")
                    pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                    pg.hotkey("ctrl", "v")
                    pg.press("return")
                    time.sleep(1)
                    ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
                    time.sleep(2)
                    # 印刷中が表示されるまで待機---------------------------------
                    IC = 0
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\NowSyouhiPrint.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                        IC += 1
                        if IC == 5:
                            pg.press("tab")
                            break
                        FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
                        if FO[0] is True:
                            pg.press("y")
                            while (
                                pg.locateOnScreen(
                                    CFolURL + r"\KTaisyou\NowSyouhiPrint.png",
                                    confidence=0.9,
                                )
                                is None
                            ):
                                time.sleep(1)
                    # --------------------------------------------------------------------
                    # 印刷中が表示されなくなるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\NowSyouhiPrint.png", confidence=0.9
                        )
                        is not None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    time.sleep(1)
                    # 確実に閉じる---------------------------------------------------------------------------------
                    KME = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\MenuEnd.png", r"\KTaisyou\MenuEnd2.png"],
                        0.9,
                        10,
                    )
                    if KME[0] is True:
                        ImgClick(CFolURL, KME[1], 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("x")
                        pg.keyUp("alt")
                    # --------------------------------------------------------------------------------------------
                    pg.keyDown("alt")
                    pg.press("x")
                    pg.keyUp("alt")
                    time.sleep(3)
                    # 確実に閉じる---------------------------------------------------------------------------------
                    KME = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\MenuEnd.png", r"\KTaisyou\MenuEnd2.png"],
                        0.9,
                        10,
                    )
                    if KME[0] is True:
                        ImgClick(CFolURL, KME[1], 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("x")
                        pg.keyUp("alt")
                    # --------------------------------------------------------------------------------------------
                    # 終了確認が表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KTaisyouEndQ.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    pg.press("y")
                    while (
                        ImgCheckForList(
                            CFolURL,
                            [r"\KTaisyou\Syouhizei.png", r"\KTaisyou\Syouhizei2.png"],
                            0.9,
                            10,
                        )[0]
                        is False
                    ):
                        time.sleep(1)
                    # 閉じる処理--------------------------
                    pg.keyDown("alt")
                    pg.press("f4")
                    pg.keyUp("alt")
                    # -----------------------------------
                    # 会計大将フラグが表示されるまで待機------------------------------------
                    while (
                        pg.locateOnScreen(
                            TFolURL + r"\Kaikei_CFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ------------------------------------------------------------------
                    return True, ThisNo, ThisYear, ThisMonth
                elif PN == "書面添付　消費税":
                    ImgClick(
                        CFolURL, r"\KTaisyou\KessanSinkoku.png", 0.9, 10
                    )  # 決算申告書アイコンをクリック
                    # 決算申告書が表示されるまで待機----------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KessanFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # サイドメニューより消費税申告書を確認------------------------------------------
                    SSM = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\Syouhizei.png", r"\KTaisyou\Syouhizei2.png"],
                        0.9,
                        10,
                    )
                    if SSM[0] is True:  # 消費税申告書があれば
                        ImgClick(CFolURL, SSM[1], 0.9, 10)  # 消費税申告書アイコンをクリック
                        time.sleep(2)
                    # 添付書面を確認------------------------------------------------
                    SFI = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\14Tenpu.png", r"\KTaisyou\14Tenpu2.png"],
                        0.9,
                        10,
                    )
                    if SFI[0] is True:  # 添付書面があれば
                        ImgClick(CFolURL, SFI[1], 0.9, 10)  # 添付書面アイコンをクリック
                    # 添付書面申告種類選択ウィンドウが表示されるまで待機--------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\TenpuBar.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    pg.press(["return", "return"])
                    # 添付書面印刷ウィンドウが表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\TenpuFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    time.sleep(3)
                    BEF = ImgCheck(CFolURL, r"\KTaisyou\BeforeQ.png", 0.9, 10)
                    if BEF[0] is True:
                        pg.press("n")
                    # ----------------------------------------------------------------------
                    WPC = ImgCheckForList(
                        CFolURL,
                        [
                            r"\KTaisyou\TenpuFlagPrint.png",
                            r"\KTaisyou\TenpuFlagPrint2.png",
                        ],
                        0.9,
                        10,
                    )
                    if WPC[0] is True:
                        ImgClick(CFolURL, WPC[1], 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("return")
                        pg.keyUp("alt")
                    # ----------------------------------------------------------------------
                    # 添付書面印刷サイズ選択が表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\TenpuPrintType.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    A4T = ImgCheck(CFolURL, r"\KTaisyou\A4Box.png", 0.99999, 10)
                    if A4T[0] is True:
                        ImgClick(CFolURL, r"\KTaisyou\A4Box.png", 0.99999, 10)
                    ImgClick(CFolURL, r"\KTaisyou\A4BoxOK.png", 0.9, 10)
                    # 一覧表出力項目指定が表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    # 申告税一覧表印刷処理----------------------------------------------------
                    FO = ImgCheckForList(
                        CFolURL,
                        [
                            r"\Houjinzei\FileOut.png",
                            r"\Houjinzei\FileOut2.png",
                        ],
                        0.9,
                        10,
                    )
                    if FO[0] is True:
                        ImgClick(CFolURL, FO[1], 0.9, 10)
                    ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
                    pg.press("return")
                    pg.press("delete")
                    pg.press("backspace")
                    pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                    pg.hotkey("ctrl", "v")
                    pg.press("return")
                    time.sleep(1)
                    ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
                    time.sleep(2)
                    # 印刷中が表示されるまで待機---------------------------------
                    IC = 0
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\TenpuPBar.png", confidence=0.9
                        )
                        is not None
                    ):
                        time.sleep(1)
                        IC += 1
                        if IC == 5:
                            pg.press("tab")
                            break
                        FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
                        if FO[0] is True:
                            pg.press("y")
                    # --------------------------------------------------------------------
                    time.sleep(2)
                    # 確実に閉じる---------------------------------------------------------------------------------
                    KME = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\MenuEnd.png", r"\KTaisyou\MenuEnd2.png"],
                        0.9,
                        10,
                    )
                    if KME[0] is True:
                        ImgClick(CFolURL, KME[1], 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("x")
                        pg.keyUp("alt")
                    # --------------------------------------------------------------------------------------------
                    # 終了確認が表示されるまで待機---------------------------------
                    while (
                        ImgCheckForList(
                            CFolURL,
                            [r"\KTaisyou\14Tenpu.png", r"\KTaisyou\14Tenpu2.png"],
                            0.9,
                            10,
                        )[0]
                        is False
                    ):
                        time.sleep(1)
                    # 閉じる処理--------------------------
                    pg.keyDown("alt")
                    pg.press("f4")
                    pg.keyUp("alt")
                    # -----------------------------------
                    # 会計大将フラグが表示されるまで待機------------------------------------
                    while (
                        pg.locateOnScreen(
                            TFolURL + r"\Kaikei_CFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ------------------------------------------------------------------
                    return True, ThisNo, ThisYear, ThisMonth
                elif PN == "決算報告書":
                    CalcErr = "Err"
                    ImgClick(
                        CFolURL, r"\KTaisyou\KessanSinkoku.png", 0.9, 10
                    )  # 決算申告書アイコンをクリック
                    # 決算申告書が表示されるまで待機----------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KessanFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # 01決算書アイコンを確認------------------------------------------
                    SSM = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\01Kessansyo.png", r"\KTaisyou\01Kessansyo2.png"],
                        0.9,
                        10,
                    )
                    if SSM[0] is True:  # 01決算書アイコンがあれば
                        ImgClick(CFolURL, SSM[1], 0.9, 10)  # 01決算書アイコンをクリック
                        time.sleep(2)
                    # 01決算書ウィンドウが表示されるまで待機--------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KessansyoFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    ImgClick(
                        CFolURL, r"\KTaisyou\KessansyoPrint.png", 0.9, 10
                    )  # 印刷ボタンクリック
                    # 一覧表出力項目指定が表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                        # 金額エラーが表示されていないか確認
                        KCE = ImgCheck(CFolURL, r"\Houjinzei\K_CalcErr.png", 0.9, 10)
                        if KCE[0] is True:
                            CalcErr = "Err"  # 金額エラーフラグを立てる
                            pg.press("y")  # 金額エラー無視で印刷
                    # --------------------------------------------------------------------
                    # 申告税一覧表印刷処理----------------------------------------------------
                    FO = ImgCheckForList(
                        CFolURL,
                        [
                            r"\Houjinzei\FileOut.png",
                            r"\Houjinzei\FileOut2.png",
                        ],
                        0.9,
                        10,
                    )
                    if FO[0] is True:
                        ImgClick(CFolURL, FO[1], 0.9, 10)
                    ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
                    pg.press("return")
                    pg.press("delete")
                    pg.press("backspace")
                    pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                    pg.hotkey("ctrl", "v")
                    pg.press("return")
                    time.sleep(1)
                    ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
                    time.sleep(2)
                    # 印刷ボタンが表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\NotEnd.png", confidence=0.99999
                        )
                        is not None
                    ):
                        time.sleep(1)
                        KOW = ImgCheck(CFolURL, r"\KTaisyou\K_OverWrite2.png", 0.9, 10)
                        if KOW[0] is True:
                            pg.press("y")
                    # --------------------------------------------------------------------
                    time.sleep(2)
                    # 確実に閉じる---------------------------------------------------------------------------------
                    KME = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\MenuEnd.png", r"\KTaisyou\MenuEnd2.png"],
                        0.9,
                        10,
                    )
                    if KME[0] is True:
                        ImgClick(CFolURL, KME[1], 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("x")
                        pg.keyUp("alt")
                    # --------------------------------------------------------------------------------------------
                    time.sleep(2)
                    # 閉じる処理--------------------------
                    pg.keyDown("alt")
                    pg.press("f4")
                    pg.keyUp("alt")
                    # -----------------------------------
                    # 会計大将フラグが表示されるまで待機------------------------------------
                    while (
                        pg.locateOnScreen(
                            TFolURL + r"\Kaikei_CFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ------------------------------------------------------------------
                    PDFM.BeppyouPDFSplit(
                        Fname.replace("\\\\", "\\").replace("/", "\\"),
                        CFolURL + r"\PDF",
                    )
                    if CalcErr == "":
                        return True, ThisNo, ThisYear, ThisMonth
                    else:
                        return False, ThisNo, ThisYear, CalcErr
                elif PN == "株主（社員）資本変動計算書":
                    CalcErr = ""
                    ImgClick(
                        CFolURL, r"\KTaisyou\KessanSinkoku.png", 0.9, 10
                    )  # 決算申告書アイコンをクリック
                    # 決算申告書が表示されるまで待機----------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KessanFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # 01決算書アイコンを確認------------------------------------------
                    SSM = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\01Kessansyo.png", r"\KTaisyou\01Kessansyo2.png"],
                        0.9,
                        10,
                    )
                    if SSM[0] is True:  # 01決算書アイコンがあれば
                        ImgClick(CFolURL, SSM[1], 0.9, 10)  # 01決算書アイコンをクリック
                        time.sleep(2)
                    # 01決算書ウィンドウが表示されるまで待機--------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KessansyoFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    ImgClick(CFolURL, r"\KTaisyou\K_Preview.png", 0.9, 10)  # 印刷ボタンクリック
                    # プレビュー画面が表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\K_AllRight.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                        # 金額エラーが表示されていないか確認
                        KCE = ImgCheck(CFolURL, r"\KTaisyou\K_CalcErr.png", 0.9, 10)
                        if KCE[0] is True:
                            CalcErr = "Err"  # 金額エラーフラグを立てる
                            pg.press("y")  # 金額エラー無視で印刷
                    # ---------------------------------------------------------------------
                    ImgClick(CFolURL, r"\KTaisyou\K_AllRight.png", 0.9, 10)  # ページ最後尾へ
                    # 左矢印が表示されるまで待機---------------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\K_Left.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ---------------------------------------------------------------------
                    ImgClick(CFolURL, r"\KTaisyou\K_Left.png", 0.9, 10)  # 1ページ戻る
                    time.sleep(1)
                    ImgClick(
                        CFolURL, r"\KTaisyou\K_ThisPrint.png", 0.9, 10
                    )  # 現在項印刷をクリック
                    time.sleep(1)
                    # 現在項印刷画面が表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\K_PrintBar.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ---------------------------------------------------------------------
                    pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                    pg.hotkey("ctrl", "v")
                    pg.press("return")
                    time.sleep(1)
                    # 現在項印刷画面が表示されなくなるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\K_PrintBar2.png", confidence=0.9
                        )
                        is not None
                    ):
                        time.sleep(1)
                        OVC = ImgCheck(CFolURL, r"\KTaisyou\K_OverWrite.png", 0.9, 10)
                        if OVC[0] is True:
                            pg.press("y")
                    # ---------------------------------------------------------------------
                    time.sleep(1)
                    # 確実に閉じる---------------------------------------------------------------------------------
                    KME = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\MenuEnd.png", r"\KTaisyou\MenuEnd2.png"],
                        0.9,
                        10,
                    )
                    if KME[0] is True:
                        ImgClick(CFolURL, KME[1], 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("x")
                        pg.keyUp("alt")
                    # --------------------------------------------------------------------------------------------
                    # プレビューボタンが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\K_Preview.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ---------------------------------------------------------------------
                    time.sleep(1)
                    # 確実に閉じる---------------------------------------------------------------------------------
                    KME = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\MenuEnd.png", r"\KTaisyou\MenuEnd2.png"],
                        0.9,
                        10,
                    )
                    if KME[0] is True:
                        ImgClick(CFolURL, KME[1], 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("x")
                        pg.keyUp("alt")
                    # --------------------------------------------------------------------------------------------
                    # 決算申告書が表示されるまで待機----------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KessanFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # 閉じる処理--------------------------
                    pg.keyDown("alt")
                    pg.press("f4")
                    pg.keyUp("alt")
                    # -----------------------------------
                    # 会計大将フラグが表示されるまで待機------------------------------------
                    while (
                        pg.locateOnScreen(
                            TFolURL + r"\Kaikei_CFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ------------------------------------------------------------------
                    if CalcErr == "":
                        return True, ThisNo, ThisYear, ThisMonth
                    else:
                        return False, ThisNo, ThisYear, CalcErr
                elif PN == "個別注記表":
                    CalcErr = ""
                    ImgClick(
                        CFolURL, r"\KTaisyou\KessanSinkoku.png", 0.9, 10
                    )  # 決算申告書アイコンをクリック
                    # 決算申告書が表示されるまで待機----------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KessanFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # 01決算書アイコンを確認------------------------------------------
                    SSM = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\01Kessansyo.png", r"\KTaisyou\01Kessansyo2.png"],
                        0.9,
                        10,
                    )
                    if SSM[0] is True:  # 01決算書アイコンがあれば
                        ImgClick(CFolURL, SSM[1], 0.9, 10)  # 01決算書アイコンをクリック
                        time.sleep(2)
                    # 01決算書ウィンドウが表示されるまで待機--------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KessansyoFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    ImgClick(CFolURL, r"\KTaisyou\K_Preview.png", 0.9, 10)  # 印刷ボタンクリック
                    # プレビュー画面が表示されるまで待機---------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\K_AllRight.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                        # 金額エラーが表示されていないか確認
                        KCE = ImgCheck(CFolURL, r"\KTaisyou\K_CalcErr.png", 0.9, 10)
                        if KCE[0] is True:
                            CalcErr = "Err"  # 金額エラーフラグを立てる
                            pg.press("y")  # 金額エラー無視で印刷
                    # ---------------------------------------------------------------------
                    ImgClick(CFolURL, r"\KTaisyou\K_AllRight.png", 0.9, 10)  # ページ最後尾へ
                    # 左矢印が表示されるまで待機---------------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\K_Left.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ---------------------------------------------------------------------
                    time.sleep(1)
                    ImgClick(
                        CFolURL, r"\KTaisyou\K_ThisPrint.png", 0.9, 10
                    )  # 現在項印刷をクリック
                    time.sleep(1)
                    # 現在項印刷画面が表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\K_PrintBar.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ---------------------------------------------------------------------
                    pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
                    pg.hotkey("ctrl", "v")
                    pg.press("return")
                    time.sleep(1)
                    # 現在項印刷画面が表示されなくなるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\K_PrintBar2.png", confidence=0.9
                        )
                        is not None
                    ):
                        time.sleep(1)
                        OVC = ImgCheck(CFolURL, r"\KTaisyou\K_OverWrite.png", 0.9, 10)
                        if OVC[0] is True:
                            pg.press("y")
                    # ---------------------------------------------------------------------
                    time.sleep(1)
                    # 確実に閉じる---------------------------------------------------------------------------------
                    KME = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\MenuEnd.png", r"\KTaisyou\MenuEnd2.png"],
                        0.9,
                        10,
                    )
                    if KME[0] is True:
                        ImgClick(CFolURL, KME[1], 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("x")
                        pg.keyUp("alt")
                    # --------------------------------------------------------------------------------------------
                    # プレビューボタンが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\K_Preview.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ---------------------------------------------------------------------
                    time.sleep(1)
                    # 確実に閉じる---------------------------------------------------------------------------------
                    KME = ImgCheckForList(
                        CFolURL,
                        [r"\KTaisyou\MenuEnd.png", r"\KTaisyou\MenuEnd2.png"],
                        0.9,
                        10,
                    )
                    if KME[0] is True:
                        ImgClick(CFolURL, KME[1], 0.9, 10)
                    else:
                        pg.keyDown("alt")
                        pg.press("x")
                        pg.keyUp("alt")
                    # --------------------------------------------------------------------------------------------
                    # 決算申告書が表示されるまで待機----------------------------------
                    while (
                        pg.locateOnScreen(
                            CFolURL + r"\KTaisyou\KessanFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # 閉じる処理--------------------------
                    pg.keyDown("alt")
                    pg.press("f4")
                    pg.keyUp("alt")
                    # -----------------------------------
                    # 会計大将フラグが表示されるまで待機------------------------------------
                    while (
                        pg.locateOnScreen(
                            TFolURL + r"\Kaikei_CFlag.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # ------------------------------------------------------------------
                    if CalcErr == "":
                        return True, ThisNo, ThisYear, ThisMonth
                    else:
                        return False, ThisNo, ThisYear, CalcErr
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "会計大将起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def DensisinkokuUpDate(FolURL, TFolURL, CFolURL, ExRow, driver, PN, Fname):
    """
    概要: 電子申告同意書印刷処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        # 電子申告のアイコンを探す-------------------------------------------------
        ICFL = ImgCheckForList(
            CFolURL, [r"\Densi\DensiIcon.png", r"\Densi\DensiIcon2.png"], 0.9, 10
        )
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 電子申告のアイコンがあれば
            ImgClick(CFolURL, ICFL[1], 0.9, 10)  # 電子申告のアイコンをクリック
            time.sleep
            ImgClick(CFolURL, r"\Densi\14D_Open.png", 0.9, 10)  # 電子申告のアイコンをクリック
            # 電子申告フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(CFolURL + r"\Densi\DensiFlag.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
                UPT = ImgCheck(
                    CFolURL, r"\Densi\14D_UpDateText.png", 0.9, 10
                )  # アップデート情報ウィンドウがあるかチェック
                if UPT[0] is True:
                    ImgClick(CFolURL, r"\Densi\14D_UDTClose.png", 0.9, 10)  # 閉じるをクリック
            # ------------------------------------------------------------------
            time.sleep(1)
            # 利用同意書アイコンを探す-------------------------------------------------
            DDR = ImgCheckForList(
                CFolURL, [r"\Densi\14Doui.png", r"\Densi\14Doui2.png"], 0.9, 10
            )
            # -----------------------------------------------------------------------
            if DDR[0] is True:  # 利用同意書アイコンがあれば
                ImgClick(CFolURL, DDR[1], 0.9, 10)  # 利用同意書アイコンをクリック
            # 利用同意書メニューが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(CFolURL + r"\Densi\14D_Print.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
                # ------------------------------------------------------------------
                ImgClick(CFolURL, r"\Densi\14D_Search.png", 0.9, 10)  # 利用同意書アイコンをクリック
            # 検索メニューが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(
                    CFolURL + r"\Densi\14D_SearchFlag.png", confidence=0.9
                )
                is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            # キャンセルボタンが選択されるまで待機------------------------------------
            while (
                pg.locateOnScreen(CFolURL + r"\Densi\CanselBtn.png", confidence=0.99999)
                is None
            ):
                time.sleep(1)
                pg.keyDown("shift")
                pg.press("tab")
                pg.keyUp("shift")
            # ------------------------------------------------------------------
            # 関与先コードで指定--------------------------------------------------
            kno = str(ExRow["関与先番号"])
            pyperclip.copy(kno)
            pg.press("tab")
            pg.hotkey("ctrl", "v")
            pg.press("return")
            pg.hotkey("ctrl", "v")
            pg.press("return")
            ImgClick(CFolURL, r"\Densi\14D_SearchOK.png", 0.9, 10)  # OKボタンをクリック
            # ------------------------------------------------------------------
            # チェックボックスが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(CFolURL + r"\Densi\14D_CheckBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
                # ------------------------------------------------------------------
            # チェックボックスが表示されなくなるまで待機------------------------------------
            while (
                pg.locateOnScreen(CFolURL + r"\Densi\14D_CheckBox.png", confidence=0.9)
                is not None
            ):
                time.sleep(1)
                # ------------------------------------------------------------------
                ImgClick(CFolURL, r"\Densi\14D_CheckBox.png", 0.9, 10)  # チェックボックスをクリック
            time.sleep(1)
            ImgClick(CFolURL, r"\Densi\14D_Print.png", 0.9, 10)  # 印刷開始ボタンをクリック
            # 印刷設定が表示されるまで待機---------------------------------
            while (
                pg.locateOnScreen(CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            # --------------------------------------------------------------------
            # 申告税一覧表印刷処理----------------------------------------------------
            FO = ImgCheckForList(
                CFolURL,
                [
                    r"\Houjinzei\FileOut.png",
                    r"\Houjinzei\FileOut2.png",
                ],
                0.9,
                10,
            )
            if FO[0] is True:
                ImgClick(CFolURL, FO[1], 0.9, 10)
            ImgClick(CFolURL, r"\Houjinzei\PDFBar.png", 0.9, 10)
            pg.press("return")
            pg.press("delete")
            pg.press("backspace")
            pyperclip.copy(Fname.replace("\\\\", "\\").replace("/", "\\"))
            pg.hotkey("ctrl", "v")
            pg.press("return")
            time.sleep(1)
            ImgClick(CFolURL, r"\Houjinzei\PrintOut.png", 0.9, 10)
            time.sleep(2)
            #  印刷設定が表示されなくなるまで待機---------------------------------
            while (
                pg.locateOnScreen(CFolURL + r"\Houjinzei\PrintBar.png", confidence=0.9)
                is not None
            ):
                time.sleep(1)
                FO = ImgCheck(CFolURL, r"\Houjinzei\FileOver.png", 0.9, 10)
                if FO[0] is True:
                    pg.press("y")
            # --------------------------------------------------------------------
            time.sleep(1)
            #  確実に閉じる--------------------------------------------
            DED = ImgCheck(CFolURL, r"\Densi\14D_End.png", 0.9, 10)
            if DED[0] is True:
                ImgClick(CFolURL, r"\Densi\14D_End.png", 0.9, 10)
            else:
                pg.keyDown("alt")
                pg.press("x")
                pg.keyUp("alt")
            # ---------------------------------------------------------
            time.sleep(1)
            while (
                ImgCheckForList(
                    CFolURL,
                    [
                        r"\Densi\14Doui.png",
                        r"\Densi\14Doui2.png",
                    ],
                    0.9,
                    10,
                )[0]
                is False
            ):
                time.sleep(1)
            # 閉じる処理--------------------------
            pg.keyDown("alt")
            pg.press("f4")
            pg.keyUp("alt")
            # -----------------------------------
            # 電子申告フラグが表示されるまで待機------------------------------------
            while (
                ImgCheckForList(
                    CFolURL,
                    [
                        r"\Densi\DensiIcon.png",
                        r"\Densi\DensiIcon2.png",
                    ],
                    0.9,
                    10,
                )[0]
                is False
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            return True, str(ExRow["関与先番号"]), "ThisYear", "ThisMonth"
        else:
            return False, "電子申告起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def OpenSystem(
    FolURL, TFolURL, CFolURL, NameDF, ExRow, Ex, ExrcHeader, isnItem, driver
):
    try:
        Eh = 0
        # 列名からMJSシステムリストを作成------------------------------------------------
        SystemList = []
        CountList = []
        for ExrcHeaderItem in ExrcHeader:
            if "_" in ExrcHeaderItem:
                SP = ExrcHeaderItem.split("_")
                SystemList.append(SP[0])
        for SystemListItem in SystemList:
            CountList.append(int(SystemList.count(SystemListItem)))
        SL = np.array(SystemList)
        CL = np.array(CountList)
        FMaxSystemList = SL[CL.argsort(axis=0)[::-1]]
        FMaxSystemList = list(dict.fromkeys(FMaxSystemList))
        print(FMaxSystemList)
        ChildFFlag = False
        # ----------------------------------------------------------------------------
        for ExrcHeaderItem in ExrcHeader:
            for FMaxSystemListItem in FMaxSystemList:
                if FMaxSystemListItem in ExrcHeaderItem:
                    Title = FMaxSystemListItem  # MJSSytem名
                    PN = ExrcHeaderItem.split("_")[1]  # MJS出力帳表名
                    # nanの場合
                    # Log--------------------------------------------
                    dt_s = datetime.datetime.now()
                    dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                    Rno = ExRow["関与先番号"]
                    Rn = NameSearch(NameDF, Rno)
                    Rn = Rn.replace("\u3000", " ")
                    NumList = [1, 2, 3, 4, 5]
                    strNumList = ["1", "2", "3", "4", "5"]
                    if (
                        ExRow[ExrcHeaderItem] in NumList
                        or ExRow[ExrcHeaderItem] in strNumList
                    ):
                        logger.debug(
                            dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_プリントメイン処理開始"
                        )
                        with open(LURL, "a") as f:
                            print(
                                [dt_s, "関与先番号:" + str(Rno), str(Rn), "プリントメイン処理開始"],
                                file=f,
                            )
                        # -----------------------------------------------
                        if not Title == "なし" or not Title == "決算フォルダ":
                            ChildFFlag = True
                            ChildFlow(
                                FolURL,
                                TFolURL,
                                CFolURL,
                                ExRow,
                                Ex,
                                Eh,
                                ExrcHeader,
                                isnItem,
                                Title,
                                PN,
                                driver,
                                Rno,
                                Rn,
                            )
            Eh += 1
        if ChildFFlag is True:
            return True
        else:
            return False
    except:
        print("TEST")
        return False


# ------------------------------------------------------------------------------------------------------------------
def MainStarter(
    FolURL, TFolURL, CFolURL, SerchURL, NameDF, ExSheet, ExrcHeader, isnItem, driver
):
    try:
        print(ExSheet[3 : len(ExSheet)])
        li = np.array(ExSheet[3 : len(ExSheet)])
        ExDf = pd.DataFrame(li, columns=ExrcHeader)
        print(ExDf)
        Exrc = np.array(ExDf).shape[1]  # 行数
        for Ex in range(1, Exrc):
            try:
                ExRow = ExDf.iloc[Ex]
                EXNo = int(ExRow["関与先番号"])
                EXName = NameSearch(NameDF, EXNo)
                Title = str(EXNo) + "_" + str(EXName) + "_RPA決算書"
                EXdir = str(ExRow["年度_(保管フォルダ名)"])
                if ExRow["関与先番号"] == ExRow["関与先番号"]:  # nan判定
                    # nanでない場合
                    OSM = OpenSystem(
                        FolURL,
                        TFolURL,
                        CFolURL,
                        NameDF,
                        ExRow,
                        Ex,
                        ExrcHeader,
                        isnItem,
                        driver,
                    )
                    if OSM is True:
                        PMURL = PDFM.PDFMarge(
                            CFolURL + r"\All\ListNumber.csv",
                            CFolURL + r"\PDF",
                            SerchURL,
                            Title,
                            EXNo,
                            EXdir,
                        )
                        Eh = len(ExrcHeader)
                        WriteEx = openpyxl.load_workbook(XLSURL)
                        WriteExSheet = WriteEx[isnItem]
                        WriteExSheet.cell(row=Ex + 5, column=Eh).hyperlink = PMURL
                        print("シート書き込み完了")
                        WriteEx.save(XLSURL)
                        WriteEx.close
                else:
                    # nanの場合
                    print("nan")
            except:
                print("データ無")
    except:
        return False, ""


# ------------------------------------------------------------------------------------------------------------------
def MainFlow(FolURL, TFolURL, CFolURL, SerchURL, Exlsx, driver):
    """
    概要: プリントメイン処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param Exlsx : Excel指示シート(obj)
    @return : bool
    """
    try:
        # ----------------------------------------------------------------------------------------------------------------------
        logger.debug("xlsxをDataFrameに")
        # sheet_namesメソッドでExcelブック内の各シートの名前をリストで取得できる
        input_sheet_name = Exlsx.sheet_names
        # lenでシートの総数を確認
        num_sheet = len(input_sheet_name)
        # シートの数とシートの名前のリストの表示
        print("Sheet の数:", num_sheet)
        print(input_sheet_name)
        for isnItem in input_sheet_name:
            # DataFrameとしてsheetのデータ読込み
            if isnItem == "印刷申請":
                ExSheet = Exlsx.parse(isnItem, skiprows=0)
                NameSheet = Exlsx.parse("顧問先別システム採用一覧表")
                print(ExSheet)
                # 初回読込時の保存--------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H-%M-%S")
                DF = pd.DataFrame(ExSheet)
                NameDF = pd.DataFrame(NameSheet)
                DF.to_csv(
                    r"\\Sv05121a\e\C 作業台\RPA\RPA_ミロクシステム次年更新\MJSLog\初回起動_プリント"
                    + dt_s
                    + ".csv",
                    encoding="cp932",
                    index=False,
                )
                # ----------------------------------------
                Exrc = np.array(ExSheet).shape[1]  # 列数
                ExrcHeader = []
                for Ex in range(Exrc):
                    ExRow = ExSheet.iloc[1]
                    print(ExRow)
                    ExSecondRow = ExSheet.iloc[2]
                    print(ExSecondRow)
                    if ExRow[Ex] == ExRow[Ex]:  # nan判定
                        # nanでない場合
                        if ExSecondRow[Ex] == ExSecondRow[Ex]:  # nan判定
                            # nanでない場合
                            ExrcHeader.append(ExRow[Ex] + "_" + ExSecondRow[Ex])
                        else:
                            # nanの場合
                            ExrcHeader.append(ExRow[Ex])
                    else:
                        # nanの場合
                        ExrcHeader.append(ExSecondRow[Ex])
                # Log--------------------------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                logger.debug(dt_s + "_Excelシート読込完了")
                with open(LURL, "a") as f:
                    print([dt_s, "Excelシート読込完了"], file=f)
                # -----------------------------------------------
                MainStarter(
                    FolURL,
                    TFolURL,
                    CFolURL,
                    SerchURL,
                    NameDF,
                    ExSheet,
                    ExrcHeader,
                    isnItem,
                    driver,
                )  # データ送信画面までの関数
    except Exception as e:
        logger.debug(e)


# RPA用画像フォルダの作成---------------------------------------------------------
FolURL = os.getcwd().replace("\\", "/")  # 先
TFolURL = FolURL + r"\RPAPhoto\MJS_SystemNextCreate"  # 先
CFolURL = FolURL + r"\RPAPhoto\MJS_SystemPrintOut"  # 先
SerchURL = r"\\Sv05121a\e\電子ファイル\(3)法人決算"  # 先
XLSDir = r"\\Sv05121a\e\C 作業台\RPA\RPA_ミロクシステム次年更新"
LURL = r"\\Sv05121a\e\C 作業台\RPA\RPA_ミロクシステム次年更新\MJSLog\MJSSysUpLog.txt"  # 処理状況CSVのURL
# --------------------------------------------------------------------------------
# Log--------------------------------------------
dt_s = datetime.datetime.now()
dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
logger.debug(dt_s + "_MJSシステム更新開始")
# -----------------------------------------------
BatUrl = FolURL + "/bat/AWADriverOpen.bat"  # 4724ポート指定でappiumサーバー起動バッチを開く
driver = MJSOpen.MainFlow(
    BatUrl, FolURL, "RPAPhoto/MJS_DensiSinkoku"
)  # MJSを起動しログイン後インスタンス化
# driver = []
FolURL = FolURL + "/RPAPhoto/MJS_DensiSinkoku"
for fd_path, sb_folder, sb_file in os.walk(XLSDir):
    FDP = fd_path
    for sb_fileItem in sb_file:
        print(sb_fileItem)
        if "ミロク更新項目" in sb_fileItem and not "ミロク更新項目(原本).xlsx" == sb_fileItem:
            XLSURL = FDP + r"\\" + sb_fileItem.replace("~", "").replace("$", "")
            MoveXLSURL = (
                FDP + r"\\MJSLog\\" + sb_fileItem.replace("~", "").replace("$", "")
            )
            os.rename(XLSURL, MoveXLSURL)
            MoveXLSURL = FDP + r"\\" + sb_fileItem.replace("~", "").replace("$", "")
            XLSURL = FDP + r"\\MJSLog\\" + sb_fileItem.replace("~", "").replace("$", "")
            open(LURL, "w").close()
            Exlsx = EFA.XlsmRead(XLSURL)
            if Exlsx[0] is True:
                try:
                    MainFlow(FolURL, TFolURL, CFolURL, SerchURL, Exlsx[1], driver)
                except:
                    traceback.print_exc()
                Exlsx = "閉じろや"
                os.rename(XLSURL, MoveXLSURL)
            else:
                print("Excel読み込みエラー")
                logger.debug("Excel読み込みエラー")
