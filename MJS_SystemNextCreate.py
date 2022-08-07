# モジュールインポート
import pyautogui as pg
import time
import MJSOpen

# pandasインポート
import pandas as pd

# 配列計算関数numpyインポート
import numpy as np

# osインポート
import os

# datetimeインポート

# 例外処理判定の為のtracebackインポート
import traceback

# pandas(pd)で関与先データCSVを取得
import pyautogui
import pyperclip  # クリップボードへのコピーで使用
import Function.ExcelFileAction as EFA

import datetime
import openpyxl

from ctypes import windll

# logger設定------------------------------------------------------------------------------------------------------------
import logging.config

logging.config.fileConfig(r"LogConf\loggingMJSSysUp.conf")
logger = logging.getLogger(__name__)
# ----------------------------------------------------------------------------------------------------------------------


def DriverUIWaitXPATH(UIPATH, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        try:
            driver.find_element_by_xpath(UIPATH)
            Flag = 1
            return True
        except:
            Flag = 0
    if Flag == 0:
        return False


# ----------------------------------------------------------------------------------------------------------------------
def DriverUIWaitAutomationId(UIPATH, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        try:
            driver.find_element_by_accessibility_id(UIPATH)
            Flag = 1
            return True
        except:
            Flag = 0
    if Flag == 0:
        return False


# ----------------------------------------------------------------------------------------------------------------------
def DriverUIWaitName(UIPATH, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        try:
            driver.find_element_by_Name(UIPATH)
            Flag = 1
            return True
        except:
            Flag = 0
    if Flag == 0:
        return False


# ------------------------------------------------------------r----------------------------------------------------------
def DriverUIWaitclassname(UIPATH, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        try:
            driver.find_element_by_class_name(UIPATH)
            Flag = 1
            return True
        except:
            Flag = 0
    if Flag == 0:
        return False


# ----------------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
def DriverFindClass(UIPATH, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        try:
            elList = driver.find_elements_by_class_name(UIPATH)
            Flag = 1
            return True, elList
        except:
            Flag = 0
    if Flag == 0:
        return False


# ----------------------------------------------------------------------------------------------------------------------
def DriverCheck(Hub, ObjName, driver):  # XPATH要素を取得するまで待機
    for x in range(10):
        if Hub == "AutomationID":
            if (
                DriverUIWaitAutomationId(ObjName, driver) is True
            ):  # OMSメニューの年調起動ボタンを判定して初期処理分け
                # 正常待機後処理
                driver.find_element_by_accessibility_id(ObjName)  # 一括電子申告送信ボタン
                return True
            else:
                # 異常待機後処理
                print("要素取得に失敗しました。")
        elif Hub == "XPATH":
            if DriverUIWaitXPATH(ObjName, driver) is True:  # OMSメニューの年調起動ボタンを判定して初期処理分け
                # 正常待機後処理
                driver.find_element_by_xpath(ObjName)  # 一括電子申告送信ボタン
                return True
            else:
                # 異常待機後処理
                print("要素取得に失敗しました。")
        elif Hub == "Name":
            if DriverUIWaitName(ObjName, driver) is True:  # OMSメニューの年調起動ボタンを判定して初期処理分け
                # 正常待機後処理
                driver.find_element_by_Name(ObjName)  # 一括電子申告送信ボタン
                return True
            else:
                # 異常待機後処理
                print("要素取得に失敗しました。")


# ----------------------------------------------------------------------------------------------------------------------
def DriverClick(Hub, ObjName, driver):
    if Hub == "AutomationID":
        if (
            DriverUIWaitAutomationId(ObjName, driver) is True
        ):  # OMSメニューの年調起動ボタンを判定して初期処理分け
            # 正常待機後処理
            OMSObj = driver.find_element_by_accessibility_id(ObjName)  # 一括電子申告送信ボタン
            OMSObj.click()
            return OMSObj
        else:
            # 異常待機後処理
            print("要素取得に失敗しました。")
    elif Hub == "XPATH":
        if DriverUIWaitXPATH(ObjName, driver) is True:  # OMSメニューの年調起動ボタンを判定して初期処理分け
            # 正常待機後処理
            OMSObj = driver.find_element_by_xpath(ObjName)  # 一括電子申告送信ボタン
            OMSObj.click()
            return OMSObj
        else:
            # 異常待機後処理
            print("要素取得に失敗しました。")
    elif Hub == "Name":
        if DriverUIWaitName(ObjName, driver) is True:  # OMSメニューの年調起動ボタンを判定して初期処理分け
            # 正常待機後処理
            OMSObj = driver.find_element_by_Name(ObjName)  # 一括電子申告送信ボタン
            OMSObj.click()
            return OMSObj
        else:
            # 異常待機後処理
            print("要素取得に失敗しました。")
    elif Hub == "class_name":
        if DriverUIWaitclassname(ObjName, driver) is True:  # OMSメニューの年調起動ボタンを判定して初期処理分け
            # 正常待機後処理
            OMSObj = driver.find_element_by_class_name(ObjName)  # 一括電子申告送信ボタン
            OMSObj.click()
            return OMSObj
        else:
            # 異常待機後処理
            print("要素取得に失敗しました。")


# ----------------------------------------------------------------------------------------------------------------------
def ImgCheck(FolURL2, FileName, conf, LoopVal):  # 画像があればTrueを返す関数
    ImgURL = FolURL2 + "/" + FileName
    for x in range(LoopVal):
        try:
            p = pyautogui.locateOnScreen(ImgURL, confidence=conf)
            x, y = pyautogui.center(p)
            return True, x, y
        except:
            Flag = 0
    if Flag == 0:
        return False, "", ""


# ----------------------------------------------------------------------------------------------------------------------
def ImgNothingCheck(FolURL2, FileName, conf, LoopVal):  # 画像がなければTrueを返す
    ImgURL = FolURL2 + "/" + FileName
    for x in range(LoopVal):
        try:
            p = pyautogui.locateOnScreen(ImgURL, confidence=conf)
            x, y = pyautogui.center(p)
            return False
        except:
            Flag = 0
    if Flag == 0:
        return True


# ----------------------------------------------------------------------------------------------------------------------
def ImgCheckForList(FolURL2, List, conf, LoopVal):  # リスト内の画像があればTrueと画像名を返す
    for x in range(LoopVal):
        for ListItem in List:
            ImgURL = FolURL2 + "/" + ListItem
            try:
                p = pyautogui.locateOnScreen(ImgURL, confidence=conf)
                x, y = pyautogui.center(p)
                return True, ListItem
                break
            except:
                Flag = 0
    if Flag == 0:
        return False, ""


# ----------------------------------------------------------------------------------------------------------------------
def ImgClick(FolURL2, FileName, conf, LoopVal):  # 画像があればクリックしてx,y軸を返す
    ImgURL = FolURL2 + "/" + FileName
    for x in range(10):
        if (
            ImgCheck(FolURL2, FileName, conf, LoopVal)[0] is True
        ):  # OMSメニューの年調起動ボタンを判定して初期処理分け
            # 正常待機後処理
            for y in range(10):
                try:
                    p = pyautogui.locateOnScreen(ImgURL, confidence=conf)
                    x, y = pyautogui.center(p)
                    pyautogui.click(x, y)
                    time.sleep(1)
                    return x, y
                except:
                    print("失敗")
        else:
            # 異常待機後処理
            print("要素取得に失敗しました。")


# ------------------------------------------------------------------------------------------------------------------
def NameSearch(NameDF, Rno):
    try:
        NameDFColumn = np.array(NameDF.columns)
        NameDF = np.array(NameDF)
        NC = np.where(NameDFColumn=="コード")
        KC = np.where(NameDFColumn=="顧問先名称")
        NameDFIndex = NameDF[:,NC]
        NR = np.where(NameDFIndex==Rno)
        N_L = NameDF[NR,KC]
        N_L = str(N_L[0]).replace("[","").replace("]","").replace("'","").replace('"',"")
        print(N_L)
        return N_L
    except:
        return "NameErr"


# ------------------------------------------------------------------------------------------------------------------
def KomonUpdate(TFolURL,ExRow):

    # 関与先コード入力ボックスをクリック------------------------------------
    ImgClick(TFolURL, r"\Komonsaki_Icon.png", 0.9, 10)
    while (
        pg.locateOnScreen(TFolURL + r"\Komonsaki_Open.png", confidence=0.9)
        is None
    ):
        time.sleep(1)

    p = pyautogui.locateOnScreen(TFolURL + r"\Komonsaki_CodeTxt.png", confidence=0.9)
    x, y = pyautogui.center(p)
    pyautogui.click(x + 100, y)
    pg.press("delete")
    pyperclip.copy(str(ExRow["関与先番号"]))
    pg.hotkey("ctrl", "v")
    pg.press(["return", "return"])

    time.sleep(1)

    p = pyautogui.locateOnScreen(TFolURL + r"\RensaouMeisyou.png", confidence=0.9)
    x, y = pyautogui.center(p)
    pyautogui.click(x + 100, y)
    pg.press("up")
    pg.press("down")
    pg.press("delete")
    pyperclip.copy(str(ExRow["関与先番号"]))
    pg.hotkey("ctrl", "v")
    pg.press(["return", "return"])

    pg.keyDown("alt")
    pg.press("u")
    pg.keyUp("alt")

    time.sleep(1)

    pg.keyDown("alt")
    pg.press("x")
    pg.keyUp("alt")

    while (
        pg.locateOnScreen(TFolURL + r"\SyonaiKanri.png", confidence=0.9)
        is None
    ):
        time.sleep(1)

    pg.keyDown("alt")
    pg.press("f4")
    pg.keyUp("alt")
    time.sleep(1)


# -----------------------------------------------------------------------


def ChildFlow(
    FolURL,
    TFolURL,
    ExRow,
    Ex,
    Eh,
    ExrcHeader,
    isnItem,
    Title,
    driver,
    Rno,
    Rn,
):
    if "会計大将" == Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_会計大将更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "会計大将更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = KaikeiUpDate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み--------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            print(WriteEx)
            WriteEx.save(XLSURL)
            WriteEx.close
            # ---------------------------------------------------------------
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_会計大将更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "会計大将更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        else:
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_会計大将更新処理エラー中断")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "会計大将更新エラー中断"], file=f)
            # ------------------------------------------------------------------------------------------
    elif "決算内訳書" == Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_決算内訳書更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "決算内訳書更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = KessanUpDate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み--------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_決算内訳書更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "決算内訳書更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "Noren":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "連動対象無エラー"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_決算内訳書更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "決算内訳書更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
    elif "減価償却" == Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_減価償却更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "減価償却更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = GenkasyoukyakuUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_減価償却更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "減価償却更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "Noren":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "決算未確定更新"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_決算未確定減価償却更新処理終了"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "決算未確定_減価償却更新処理終了"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
        # ---------------------------------------------------------------
    elif "法人税申告書" == Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法人税申告書更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "法人税申告書更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = HoujinzeiUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法人税申告書更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "法人税申告書更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "要データ再計算":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "要データ再計算"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法人税申告書更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "法人税申告書更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "要申告指定":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "要申告指定"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法人税申告書申告指定無しの為中断"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "法人税申告書申告指定無しの為中断"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
    elif "所得税確定申告" == Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_所得税更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "所得税更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = SyotokuzeiUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_所得税更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "所得税更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "Nocalc":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "計算未処理更新"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_所得税更新処理計算未処理で終了"
            )
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "所得税更新処理計算未処理で終了"], file=f)
        elif SystemUp[0] is False:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "関与先無"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_所得税更新関与先無しの為終了"
            )
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "所得税更新関与先無しの為終了"], file=f)
            # -----------------------------------------------
            # -------------------------------------------
    elif "財産評価明細書" == Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_財産評価明細書更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "財産評価明細書更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = ZaisanUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_財産評価明細書更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "財産評価明細書更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "更新対象年度無し":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "更新対象年度無し"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(
                dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_財産評価明細書更新処理終了_更新対象年度無し"
            )
            with open(LURL, "a") as f:
                print(
                    [dt_s, "関与先番号:" + str(Rno), str(Rn), "財産評価明細書更新処理終了_更新対象年度無し"],
                    file=f,
                )
            # ------------------------------------------------------------------------------------------
        # ---------------------------------------------------------------
    elif "年末調整" == Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_年末調整更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "年末調整更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = NencyouUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_年末調整更新処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "年末調整更新処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        elif SystemUp[1] == "次年度あり":
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
            # Log---------------------------------------------------------------------------------------
            dt_s = datetime.datetime.now()
            dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
            logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_年末調整更新次年度あり処理終了")
            with open(LURL, "a") as f:
                print([dt_s, "関与先番号:" + str(Rno), str(Rn), "年末調整更新次年度あり処理終了"], file=f)
            # ------------------------------------------------------------------------------------------
        # ---------------------------------------------------------------
    elif "法定調書" == Title:
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法定調書更新処理開始")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "法定調書更新処理開始"], file=f)
        # ------------------------------------------------------------------------------------------
        SystemUp = HouteiUpdate(FolURL, TFolURL, ExRow, driver)
        # Excel書き込み---------------------------------------------------
        if SystemUp[0] is True:
            dt_now = datetime.datetime.now()
            dt_now = dt_now.strftime("%Y/%m/%d %H:%M:%S")
            WriteEx = openpyxl.load_workbook(XLSURL, keep_vba=True)
            WriteExSheet = WriteEx[isnItem]
            WriteExSheet.cell(row=Ex + 5, column=Eh + 2).value = dt_now
            WriteExSheet.cell(row=Ex + 5, column=Eh + 1).value = "○"
            print("シート書き込み完了")
            WriteEx.save(XLSURL)
            WriteEx.close
        # ---------------------------------------------------------------
        # Log---------------------------------------------------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_関与先番号:" + str(Rno) + ":" + str(Rn) + "_法定調書更新処理終了")
        with open(LURL, "a") as f:
            print([dt_s, "関与先番号:" + str(Rno), str(Rn), "法定調書更新処理終了"], file=f)
        # ------------------------------------------------------------------------------------------
    else:
        print("NoSystem")


# ------------------------------------------------------------------------------------------------------------------
def HouteiUpdate(FolURL, TFolURL, ExRow, driver):
    """
    00:"関与先番号"
    01:"関与先名"
    02:"担当者_ｺｰﾄﾞ"
    03:"担当者_担当者名"
    04:"会計大将_繰越対象"
    07:"会計大将_繰越処理日"
    08:"決算内訳書_繰越対象"
    11:"決算内訳書_繰越処理日"
    12:"減価償却_繰越対象"
    15:"減価償却_繰越処理日"
    16:"法人税申告書_繰越対象"
    19:"法人税申告書_繰越処理日"
    20:"財産評価明細書_繰越対象"
    23:"財産評価明細書_繰越処理日"
    24:"財産評価明細書"
    25:"年末調整"
    26:"法定調書"
    27:"所得税確定申告"
    """
    """
    概要: 法定調書更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        ErrStr = ""  # Rpaエラー判別変数
        # 法定調書のアイコンを探す-------------------------------------------------
        ImgList = [r"\Houtei.png", r"\Houtei2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 法定調書のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 法定調書のアイコンをクリック
            # 法定調書フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\HouteiFlag.png", confidence=0.9) is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 他システムとメニューが違う-------------------------------------------------------
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return"])
            # 入力した関与先コードを取得------------
            pg.keyDown("shift")
            pg.press(["tab", "tab"])
            pg.keyUp("shift")
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisNo = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された年度を取得-----------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisYear = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            pg.press("return")
            time.sleep(1)
            # 他システムとメニューが違う-------------------------------------------------------
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 法定調書メニューが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\HouteiMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    # 顧問先情報変更ダイアログが表示されたら
                    CDQ = ImgCheck(
                        TFolURL,
                        r"ChangeDataQ.png",
                        0.9,
                        10,
                    )
                    if CDQ[0] is True:
                        pg.press("y")  # yで決定
                        # 顧問先情報取込メニューが表示されるまで待機--------------------------
                        while (
                            ImgCheckForList(
                                TFolURL,
                                [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        CDB = ImgCheckForList(
                            TFolURL,
                            [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                            0.9,
                            10,
                        )
                        ImgClick(TFolURL, CDB[1], 0.9, 10)  # 顧問先情報取込ボタンをクリック
                # --------------------------------------------------------------------
                ImgClick(TFolURL, r"\Siharaisya.png", 0.9, 10)  # 支払者基本情報をクリック
                time.sleep(1)
                SAN = ImgCheck(TFolURL, r"\SansyouOK.png", 0.9, 50) # 参照表示ダイアログを確認
                if SAN[0] is True:
                    time.sleep(1)
                    ImgClick(TFolURL, r"\DataInNo.png", 0.9, 50)
                    while (
                        pg.locateOnScreen(TFolURL + r"\DataInYes.png", confidence=0.9)
                        is not None
                    ):
                        ImgClick(TFolURL, r"\DataInYes.png", 0.9, 10)
                        time.sleep(1)
                # 顧問先情報取り込みアイコンが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\DataInIcon.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)         
                ImgClick(TFolURL, r"\DataInIcon.png", 0.9, 10)  # 顧問先情報取り込みアイコンをクリック
                # 取り込むボタンが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\DataInIcon2.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                ImgClick(TFolURL, r"\DataInIcon2.png", 0.9, 10)  #取り込むボタンをクリック
                time.sleep(1)
                pg.keyDown("alt")
                pg.press("x")
                pg.keyUp("alt")
                time.sleep(1)
                # 法定調書メニューが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\HouteiMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    SKE = ImgCheck(TFolURL, r"\SiharaiKihonEnd.png", 0.9, 10)  # 終了ダイアログ確認
                    if SKE[0] is True:
                        pg.press("y")
                ImgClick(TFolURL, r"\HouteiKousin.png", 0.9, 10)  # 法定調書更新をクリック
                # 法定調書メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\HouteiAll.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                time.sleep(1)
                # 更新区分フラグが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\HouteiZenken.png", confidence=0.9)
                    is None
                ):
                    pg.press("tab")

                FC = ImgCheckForList(
                    TFolURL,
                    [
                        r"IkkatuFind.png",
                        r"IkkatuFind2.png",
                    ],
                    0.9,
                    10,
                )
                if FC[0] is True:
                    ImgClick(TFolURL, FC[1], 0.9, 10)  # 一括検索メニューのアイコンをクリック
                    # 確認ウィンドウが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(TFolURL + r"\Rensou.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    p = pg.locateOnScreen(TFolURL + r"\Rensou.png", confidence=0.9)
                    x, y = pg.center(p)
                    pg.click(x + 100, y)
                    pyperclip.copy(str(ExRow["関与先番号"]))
                    pg.hotkey("ctrl", "v")
                    # 検索ボタンまでエンター-------------------------------------
                    while ImgCheck(TFolURL, r"ZFindFlag2.png", 0.9, 10)[0] is False:
                        time.sleep(1)
                        pg.press("return")
                pg.press("return")
                time.sleep(1)
                pg.press("space")

                # --------------------------------------------------------------------
                if ErrStr == "":
                    pg.press("return")
                    pg.press("space")
                    # チェックマークが表示されるまで待機-------------------------------------
                    while (
                        ImgCheckForList(
                            TFolURL,
                            [
                                r"IkkatuCheck.png",
                                r"ZaisanCheck.png",
                                r"NendCheck.png",
                                r"HouteiCheck.png",
                            ],
                            0.9,
                            10,
                        )[0]
                        is False
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    time.sleep(1)
                    ImgClick(TFolURL, r"\HouteiStart.png", 0.9, 10)  # 更新開始のアイコンをクリック
                    # 確認ウィンドウが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(TFolURL + r"\HouteiQ.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    pg.press("y")  # yで決定(nがキャンセル)
                    # 処理終了ウィンドウが表示されるまで待機----------------------------------
                    while (
                        pg.locateOnScreen(TFolURL + r"\HouteiEnd.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    pg.press("return")
                    # ロードを待機----------------------------------------------------------
                    while (
                        pg.locateOnScreen(
                            TFolURL + r"\HouteiNoCheck.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    # 法定調書終了確認が表示されるまで待機--------------------------------------
                    while (
                        pg.locateOnScreen(
                            TFolURL + r"\HouteiEndMsg.png", confidence=0.9
                        )
                        is None
                    ):
                        time.sleep(1)
                        ME = ImgCheckForList(
                            TFolURL, [r"\MenuEnd.png", r"\MenuEnd2.png"], 0.9, 10
                        )
                        if ME[0] is True:
                            ImgClick(TFolURL, ME[1], 0.9, 10)  # 終了アイコンをクリック
                    # --------------------------------------------------------------------
                    pg.press("y")
                    # 法定調書開始フラグが表示されるまで待機--------------------------------------
                    while (
                        pg.locateOnScreen(TFolURL + r"\HouteiMenu.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    # 閉じる処理--------------------------
                    pg.keyDown("alt")
                    pg.press("f4")
                    pg.keyUp("alt")
                    # -----------------------------------
                    # 法定調書フラグが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(TFolURL + r"\HouteiFlag.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    print("更新完了")
                    if ErrStr == "":
                        return True, ThisNo, ThisYear, "ThisMonth"
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "法定調書起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def NencyouUpdate(FolURL, TFolURL, ExRow, driver):
    """
    概要: 年調更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        ErrStr = ""  # Rpaエラー判別変数
        # 年調のアイコンを探す-------------------------------------------------
        ImgList = [r"\Nencyou.png", r"\Nencyou2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 年調のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 年調のアイコンをクリック
            # 年調フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\NencyouFlag.png", confidence=0.9) is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 他システムとメニューが違う-------------------------------------------------------
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return"])
            # 入力した関与先コードを取得------------
            pg.keyDown("shift")
            pg.press(["tab", "tab"])
            pg.keyUp("shift")
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisNo = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された年度を取得-----------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisYear = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            pg.press("return")
            time.sleep(1)
            # 他システムとメニューが違う-------------------------------------------------------
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 年調更新アイコンが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\NencyouOpenFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    # 顧問先情報変更ダイアログが表示されたら
                    CDQ = ImgCheck(
                        TFolURL,
                        r"ChangeDataQ.png",
                        0.9,
                        10,
                    )
                    if CDQ[0] is True:
                        pg.press("y")  # yで決定
                        # 顧問先情報取込メニューが表示されるまで待機--------------------------
                        while (
                            ImgCheckForList(
                                TFolURL,
                                [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        CDB = ImgCheckForList(
                            TFolURL,
                            [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                            0.9,
                            10,
                        )
                        ImgClick(TFolURL, CDB[1], 0.9, 10)  # 顧問先情報取込ボタンをクリック
                # --------------------------------------------------------------------





                ImgClick(TFolURL, r"\NencyouTopMenu.png", 0.9, 10)  # 印刷更新タブをクリック
                time.sleep(1)
                ImgClick(TFolURL, r"\NenjiKakutei.png", 0.9, 10)  # 年長確定処理ボタンをクリック
                # 年長確定確認メニューが表示されるまで待機--------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\NencyouKakuteiQ.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                time.sleep(1)
                NK = ImgCheck(TFolURL, r"\NencyouKubun.png", 0.99999, 10)  # 年調チェックボックスをクリック      
                if NK[0] is True:
                    ImgClick(TFolURL, r"\NencyouKubun.png", 0.99999, 10)  # ＯＫアイコンをクリック
                time.sleep(1)        
                ImgClick(TFolURL, r"\NencyouOK.png", 0.9, 10)  # ＯＫアイコンをクリック
                time.sleep(1)
                pg.press("y")
                # 年調更新アイコンが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\NencyouOpenFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                ImgClick(TFolURL, r"\DounyuuTAB.png", 0.9, 10)  # 導入タブをクリック
                time.sleep(1)
                ImgClick(TFolURL, r"\CamIcon.png", 0.9, 10)  # 会社基本情報アイコンをクリック
                # 顧問先情報取り込みアイコンが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\DataInIcon.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)         
                ImgClick(TFolURL, r"\DataInIcon.png", 0.9, 10)  # 顧問先情報取り込みアイコンをクリック
                # 更新確認ダイアログが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\KousinKakunin.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1) 
                pg.press("y")
                # 更新完了ダイアログが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\KousinKanryou.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                pg.press("return")
                # 取り込むボタンが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\DataInOK.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                ImgClick(TFolURL, r"\DataInOK.png", 0.9, 10)  #取り込むボタンをクリック
                time.sleep(1)
                pg.keyDown("alt")
                pg.press("x")
                pg.keyUp("alt")
                time.sleep(1)
                # 印刷更新タブが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\NencyouTopMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                ImgClick(TFolURL, r"\NencyouTopMenu.png", 0.9, 10)  # 印刷更新タブをクリック
                # 年調メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\NencyouKousin.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                time.sleep(1)




                # 年調更新をクリック---------------------------------------------------
                ImgClick(TFolURL, r"\NencyouKousin.png", 0.9, 10)
                # 更新区分フラグが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(
                        TFolURL + r"\NencyouKousinFlag.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)

                FC = ImgCheckForList(
                    TFolURL,
                    [
                        r"IkkatuFind.png",
                        r"IkkatuFind2.png",
                    ],
                    0.9,
                    10,
                )
                if FC[0] is True:
                    ImgClick(TFolURL, FC[1], 0.9, 10)  # 一括更新メニューのアイコンをクリック
                    pyperclip.copy(str(ExRow["関与先番号"]))
                    pg.hotkey("ctrl", "v")
                    # 検索ボタンまでエンター-------------------------------------
                    while ImgCheck(TFolURL, r"ZNBtn.png", 0.9, 10)[0] is False:
                        time.sleep(1)
                        pg.press("return")
                pg.press("return")
                time.sleep(1)
                pg.press("space")

                # --------------------------------------------------------------------
                if ErrStr == "":
                    # チェックマークが表示されるまで待機-------------------------------------
                    while (
                        ImgCheckForList(
                            TFolURL,
                            [
                                r"IkkatuCheck.png",
                                r"ZaisanCheck.png",
                                r"NendCheck.png",
                                r"HouteiCheck.png",
                            ],
                            0.9,
                            10,
                        )[0]
                        is False
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    time.sleep(1)
                    ImgClick(TFolURL, r"\NencyouStart.png", 0.9, 10)  # 更新開始のアイコンをクリック
                    # 確認ウィンドウが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(TFolURL + r"\NencyouQ.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    pg.press("y")  # yで決定(nがキャンセル)
                    # 確認ウィンドウが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(TFolURL + r"\NencyouQ2.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    pg.press("y")  # yで決定(nがキャンセル)
                    # 処理終了ウィンドウが表示されるまで待機----------------------------------
                    while (
                        pg.locateOnScreen(TFolURL + r"\NencyouFQ.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                        # 年調データが完了していない場合表示される----------------------------
                        ErrC = ImgCheck(TFolURL, r"NencyouQ3ErrMsg.png", 0.9, 10)
                        if ErrC[0] is True:
                            pg.press("y")  # yで決定(nがキャンセル)
                        # 自然データがすでにある場合表示される----------------------------
                        NNQ = ImgCheck(TFolURL, r"NextNoQ.png", 0.9, 10)
                        if NNQ[0] is True:
                            pg.press("n")  # yで決定(nがキャンセル)
                            ErrStr = "NoData"
                            break
                    if ErrStr != "NoData":
                       # --------------------------------------------------------------------
                        pg.press("return")
                        # チェックマークが表示されなくなるまで待機-------------------------------
                        while (
                            ImgCheckForList(
                                TFolURL,
                                [
                                    r"IkkatuCheck.png",
                                    r"ZaisanCheck.png",
                                    r"NendCheck.png",
                                    r"HouteiCheck.png",
                                ],
                                0.9,
                                10,
                            )[0]
                            is True
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        # 処理終了ウィンドウが表示されるまで待機----------------------------------
                        while (
                            pg.locateOnScreen(TFolURL + r"\NencyouEnd.png", confidence=0.9)
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        pg.press("return")

                        # --------------------------------------------------------------------
                        ME = ImgCheckForList(
                            TFolURL, [r"\MenuEnd.png", r"\MenuEnd2.png"], 0.9, 10
                        )
                        if ME[0] is True:
                            ImgClick(TFolURL, ME[1], 0.9, 10)  # 終了アイコンをクリック
                        # 年調開始フラグが表示されるまで待機--------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\NencyouOpenFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        # 閉じる処理--------------------------
                        pg.keyDown("alt")
                        pg.press("f4")
                        pg.keyUp("alt")
                        # -----------------------------------
                        # 年調フラグが表示されるまで待機-------------------------------------
                        while (
                            pg.locateOnScreen(TFolURL + r"\NencyouFlag.png", confidence=0.9)
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        print("更新完了")
                    else:
                        # --------------------------------------------------------------------
                        ME = ImgCheckForList(
                            TFolURL, [r"\MenuEnd.png", r"\MenuEnd2.png"], 0.9, 10
                        )
                        if ME[0] is True:
                            ImgClick(TFolURL, ME[1], 0.9, 10)  # 終了アイコンをクリック
                        # 年調開始フラグが表示されるまで待機--------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\NencyouOpenFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        # 閉じる処理--------------------------
                        pg.keyDown("alt")
                        pg.press("f4")
                        pg.keyUp("alt")
                        # -----------------------------------
                        # 年調フラグが表示されるまで待機-------------------------------------
                        while (
                            pg.locateOnScreen(TFolURL + r"\NencyouFlag.png", confidence=0.9)
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        print("更新完了")
                    if ErrStr == "":
                        return True, ThisNo, ThisYear, "ThisMonth"
                    else:
                        return False, "次年度あり", "", ""
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "年調起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def ZaisanUpdate(FolURL, TFolURL, ExRow, driver):
    """
    00:"関与先番号"
    01:"関与先名"
    02:"担当者_ｺｰﾄﾞ"
    03:"担当者_担当者名"
    04:"会計大将_繰越対象"
    07:"会計大将_繰越処理日"
    08:"決算内訳書_繰越対象"
    11:"決算内訳書_繰越処理日"
    12:"減価償却_繰越対象"
    15:"減価償却_繰越処理日"
    16:"法人税申告書_繰越対象"
    19:"法人税申告書_繰越処理日"
    20:"財産評価明細書_繰越対象"
    23:"財産評価明細書_繰越処理日"
    24:"財産評価明細書"
    25:"年末調整"
    26:"法定調書"
    27:"所得税確定申告"
    """
    """
    概要: 財産評価更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        ErrStr = ""  # Rpaエラー判別変数
        # 財産評価のアイコンを探す-------------------------------------------------
        ImgList = [r"\Zaisanhyouka.png", r"\Zaisanhyouka2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 財産評価のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 財産評価のアイコンをクリック
            # 財産評価フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\ZaisanhyoukaFlag.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 他システムとメニューが違う-------------------------------------------------------
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return"])
            # -----------------------------------
            if ImgCheck(TFolURL, r"\NotData.png", 0.9, 10)[0] is True:
                # 入力した関与先コードを取得------------
                pg.press(["return", "return"])
                pg.keyDown("shift")
                pg.press(["tab", "tab", "tab", "tab"])
                pg.keyUp("shift")
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisNo = pyperclip.paste()
                pg.press("return")
                # クリップボードをクリア----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                # ------------------------------------
                # 表示された年度を取得-----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisYear = pyperclip.paste()
                # -----------------------------------
                pg.press("return")
                # 表示された申告種類を取得---------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisMonth = pyperclip.paste()
                pg.press("return")
                # -----------------------------------
            else:
                # 入力した関与先コードを取得------------
                pg.press(["return", "return"])
                pg.keyDown("shift")
                pg.press(["tab", "tab", "tab", "tab"])
                pg.keyUp("shift")
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisNo = pyperclip.paste()
                pg.press("return")
                # 表示された年度を取得-----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisYear = pyperclip.paste()
                # -----------------------------------
                pg.press("return")
                # 表示された申告種類を取得---------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisMonth = pyperclip.paste()
                pg.press("return")
                # -----------------------------------
            # 他システムとメニューが違う-------------------------------------------------------
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 財産評価更新アイコンが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\ZaisanKousin.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    HQ = ImgCheck(TFolURL, r"\ZaisanOpenQ.png", 0.9, 10)
                    if HQ[0] is True:
                        ImgClick(TFolURL, r"\ZaisanOpenQCansel.png", 0.9, 10)
                    # 顧問先情報変更ダイアログが表示されたら
                    CDQ = ImgCheck(
                        TFolURL,
                        r"ZChangeDataQ.png",
                        0.9,
                        10,
                    )
                    if CDQ[0] is True:
                        pg.press("y")  # yで決定
                        # 顧問先情報取込メニューが表示されるまで待機--------------------------
                        while (
                            ImgCheckForList(
                                TFolURL,
                                [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        CDB = ImgCheckForList(
                            TFolURL,
                            [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                            0.9,
                            10,
                        )
                        ImgClick(TFolURL, CDB[1], 0.9, 10)  # 顧問先情報取込ボタンをクリック
                # --------------------------------------------------------------------
                ImgClick(TFolURL, r"\ZaisanKousin.png", 0.9, 10)  # 一括更新のアイコンをクリック
                # 財産評価メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\ZaisanOpenFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                time.sleep(3)
                # 更新区分を次年繰越に変更---------------------------------------------
                ImgClick(TFolURL, r"\Kousinkubun.png", 0.9, 10)
                pg.press("home")
                pg.press("return")
                # 更新区分フラグが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\KousinkubunFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                FC = ImgCheckForList(
                    TFolURL,
                    [
                        r"IkkatuFind.png",
                        r"IkkatuFind2.png",
                    ],
                    0.9,
                    10,
                )
                if FC[0] is True:
                    ImgClick(TFolURL, FC[1], 0.9, 10)  # 一括更新メニューのアイコンをクリック
                    pyperclip.copy(str(ExRow["関与先番号"]))
                    pg.hotkey("ctrl", "v")
                    # 検索ボタンまでエンター-------------------------------------
                    while ImgCheck(TFolURL, r"ZFindFlag.png", 0.9, 10)[0] is False:
                        time.sleep(1)
                        pg.press("return")
                pg.press("return")
                time.sleep(1)
                pg.press("space")

                if ErrStr == "":
                    # チェックマークが表示されるまで待機-------------------------------------
                    while (
                        ImgCheckForList(
                            TFolURL,
                            [
                                r"IkkatuCheck.png",
                                r"ZaisanCheck.png",
                                r"NendCheck.png",
                                r"HouteiCheck.png",
                            ],
                            0.9,
                            10,
                        )[0]
                        is False
                    ):
                        time.sleep(1)
                        ZND = ImgCheck(TFolURL, r"ZaisanNoData.png", 0.9, 10)
                        if ZND[0] is True:
                            ErrStr = "NoData"
                            break
                    # --------------------------------------------------------------------
                    time.sleep(1)
                    if not ErrStr == "NoData":
                        ImgClick(
                            TFolURL, r"\ZaisanStart.png", 0.9, 10
                        )  # 更新開始のアイコンをクリック
                        # 確認ウィンドウが表示されるまで待機-------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\ZaisanStartQ.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        pg.press("y")  # yで決定(nがキャンセル)
                        # 処理終了ウィンドウが表示されるまで待機----------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\ZaisanEnd.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        pg.press("return")
                        # チェックマークが表示されなくなるまで待機-------------------------------
                        while (
                            ImgCheckForList(
                                TFolURL,
                                [
                                    r"IkkatuCheck.png",
                                    r"ZaisanCheck.png",
                                    r"NendCheck.png",
                                    r"HouteiCheck.png",
                                ],
                                0.9,
                                10,
                            )[0]
                            is True
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        ME = ImgCheckForList(
                            TFolURL, [r"\MenuEnd.png", r"\MenuEnd2.png"], 0.9, 10
                        )
                        if ME[0] is True:
                            ImgClick(TFolURL, ME[1], 0.9, 10)  # 終了アイコンをクリック
                        # 一括更新のアイコンが表示されるまで待機----------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\ZaisanOpenFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        # 閉じる処理--------------------------
                        pg.keyDown("alt")
                        pg.press("f4")
                        pg.keyUp("alt")
                        # -----------------------------------
                        # 財産評価フラグが表示されるまで待機-------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\ZaisanhyoukaFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        # 初期画面で開封された財産評価項目を閉じる----------------------------------
                        HoujinList = [r"\Zaisanhyouka.png", r"\Zaisanhyouka2.png"]
                        HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
                        if HLI[0] is True:
                            ImgClick(TFolURL, HLI[1], 0.9, 10)
                        # --------------------------------------------------------------------
                        print("更新完了")
                    if ErrStr == "":
                        return True, ThisNo, ThisYear, ThisMonth
                    else:
                        # --------------------------------------------------------------------
                        DD = ImgCheck(
                            TFolURL, r"\DoubleDataQ.png", 0.9, 10
                        )
                        if DD[0] is True:
                            pg.press("return")
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        ME = ImgCheckForList(
                            TFolURL, [r"\MenuEnd.png", r"\MenuEnd2.png"], 0.9, 10
                        )
                        if ME[0] is True:
                            ImgClick(TFolURL, ME[1], 0.9, 10)  # 終了アイコンをクリック
                        # 一括更新のアイコンが表示されるまで待機----------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\ZaisanOpenFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        # 閉じる処理--------------------------
                        pg.keyDown("alt")
                        pg.press("f4")
                        pg.keyUp("alt")
                        # -----------------------------------
                        # 財産評価フラグが表示されるまで待機-------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\ZaisanhyoukaFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        # 初期画面で開封された財産評価項目を閉じる----------------------------------
                        HoujinList = [r"\Zaisanhyouka.png", r"\Zaisanhyouka2.png"]
                        HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
                        if HLI[0] is True:
                            ImgClick(TFolURL, HLI[1], 0.9, 10)
                        # --------------------------------------------------------------------
                        print("更新完了_更新対象年度無し")
                        return False, "更新対象年度無し", "", ""
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "財産評価起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def SyotokuzeiUpdate(FolURL, TFolURL, ExRow, driver):
    """
    概要: 所得税更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        ErrStr = ""  # Rpaエラー判別変数
        # 所得税のアイコンを探す-------------------------------------------------
        ImgList = [r"\Syotoku.png", r"\Syotoku2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 所得税のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 所得税のアイコンをクリック
            # 所得税フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\SyotokuFlag.png", confidence=0.9) is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 他システムとメニューが違う-------------------------------------------------------
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return"])
            # -----------------------------------
            if ImgCheck(TFolURL, r"\NotData.png", 0.9, 10)[0] is True:
                # 入力した関与先コードを取得------------
                pg.press("return")
                pg.keyDown("shift")
                pg.press(["tab", "tab", "tab"])
                pg.keyUp("shift")
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisNo = pyperclip.paste()
                pg.press("return")
                # クリップボードをクリア----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                # ------------------------------------
                # 表示された年度を取得-----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisYear = pyperclip.paste()
                # -----------------------------------
                pg.press("return")
                # 表示された申告種類を取得---------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisMonth = pyperclip.paste()
                pg.press("return")
                # -----------------------------------
            else:
                # 入力した関与先コードを取得------------
                pg.press("return")
                pg.keyDown("shift")
                pg.press(["tab", "tab", "tab"])
                pg.keyUp("shift")
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisNo = pyperclip.paste()
                pg.press("return")
                # 表示された年度を取得-----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisYear = pyperclip.paste()
                # -----------------------------------
                pg.press("return")
                # 表示された申告種類を取得---------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisMonth = pyperclip.paste()
                pg.press("return")
                # -----------------------------------
            # 他システムとメニューが違う-------------------------------------------------------
            if str(
                ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 所得税メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SyotokuMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    # 顧問先情報変更ダイアログが表示されたら
                    CDQ = ImgCheck(
                        TFolURL,
                        r"ChangeDataQ.png",
                        0.9,
                        10,
                    )
                    if CDQ[0] is True:
                        pg.press("y")  # yで決定
                        # 顧問先情報取込メニューが表示されるまで待機--------------------------
                        while (
                            ImgCheckForList(
                                TFolURL,
                                [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        CDB = ImgCheckForList(
                            TFolURL,
                            [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                            0.9,
                            10,
                        )
                        ImgClick(TFolURL, CDB[1], 0.9, 10)  # 顧問先情報取込ボタンをクリック
                # --------------------------------------------------------------------
                ImgClick(TFolURL, r"\SyotokuKousin.png", 0.9, 10)  # 一括更新のアイコンをクリック
                # 所得税メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SyotokuKMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                # チェックボックス直前までTAB------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SyotokuZenken.png", confidence=0.9)
                    is None
                ):
                    pg.press("tab")
                # --------------------------------------------------------------------
                pg.press("tab")
                pg.press("space")
                # チェックマークが表示されるまで待機-------------------------------------
                while (
                    ImgCheckForList(
                        TFolURL,
                        [
                            r"IkkatuCheck.png",
                            r"ZaisanCheck.png",
                            r"NendCheck.png",
                            r"HouteiCheck.png",
                        ],
                        0.9,
                        10,
                    )[0]
                    is False
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                SNC = ImgCheck(TFolURL, r"SyotokuNoCalc.png", 0.9, 10)
                if SNC[0] is True:
                    ErrStr = "Nocalc"
                    pg.press("y")  # yで決定(nがキャンセル)
                time.sleep(2)
                ImgClick(TFolURL, r"\SyotokuStart.png", 0.9, 10)  # 更新開始のアイコンをクリック
                # 確認ウィンドウが表示されるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SyotokuQ.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                pg.press("y")  # yで決定(nがキャンセル)
                # 処理終了ウィンドウが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SyotokuEnd.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                pg.press("return")  # 決定
                # 一括更新のアイコンが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SyotokuMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    ME = ImgCheckForList(
                        TFolURL, [r"\MenuEnd.png", r"\MenuEnd2.png"], 0.9, 10
                    )
                    if ME[0] is True:
                        ImgClick(TFolURL, ME[1], 0.9, 10)  # 終了アイコンをクリック
                # --------------------------------------------------------------------
                # 閉じる処理--------------------------
                pg.keyDown("alt")
                pg.press("f4")
                pg.keyUp("alt")
                # -----------------------------------
                # 所得税フラグが表示されるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SyotokuFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                # 初期画面で開封された所得税項目を閉じる----------------------------------
                HoujinList = [
                    r"\Syotoku.png",
                    r"\Syotoku2.png",
                ]
                HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
                if HLI[0] is True:
                    ImgClick(TFolURL, HLI[1], 0.9, 10)
                # --------------------------------------------------------------------
                print("更新完了")
                if ErrStr == "":
                    return True, ThisNo, ThisYear, ThisMonth
                elif ErrStr == "Nocalc":
                    return True, "Nocalc", ThisYear, ThisMonth
            else:
                print("関与先なし")
                # 初期画面で開封された所得税項目を閉じる----------------------------------
                HoujinList = [
                    r"\Syotoku.png",
                    r"\Syotoku2.png",
                ]
                HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
                if HLI[0] is True:
                    ImgClick(TFolURL, HLI[1], 0.9, 10)
                # --------------------------------------------------------------------
                return False, "関与先なし", "", ""
        else:
            return False, "所得税起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def HoujinzeiUpdate(FolURL, TFolURL, ExRow, driver):
    """
    概要: 法人税更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        ErrStr = ""  # Rpaエラー判別変数
        # 法人税のアイコンを探す-------------------------------------------------
        ImgList = [r"\Houjinzei.png", r"\Houjinzei2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 法人税のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 法人税のアイコンをクリック
            # 法人税フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\HoujinFlag.png", confidence=0.9) is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 他システムとメニューが違う-------------------------------------------------------
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return"])
            # -----------------------------------
            if ImgCheck(TFolURL, r"\NotData.png", 0.9, 10)[0] is True:
                # 入力した関与先コードを取得------------
                pg.press(["return", "return"])
                pg.keyDown("shift")
                pg.press(["tab", "tab", "tab", "tab"])
                pg.keyUp("shift")
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisNo = pyperclip.paste()
                pg.press("return")
                # クリップボードをクリア----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                # ------------------------------------
                # 表示された年度を取得-----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisYear = pyperclip.paste()
                # -----------------------------------
                pg.press("return")
                # 表示された申告種類を取得---------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisMonth = pyperclip.paste()
                pg.press("return")
                # -----------------------------------
            else:
                # 入力した関与先コードを取得------------
                pg.press(["return", "return"])
                pg.keyDown("shift")
                pg.press(["tab", "tab", "tab", "tab"])
                pg.keyUp("shift")
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisNo = pyperclip.paste()
                pg.press("return")
                # 表示された年度を取得-----------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisYear = pyperclip.paste()
                # -----------------------------------
                pg.press("return")
                # 表示された申告種類を取得---------------
                if windll.user32.OpenClipboard(None):
                    windll.user32.EmptyClipboard()
                    windll.user32.CloseClipboard()
                pg.hotkey("ctrl", "c")
                ThisMonth = pyperclip.paste()
                pg.press("return")
                # -----------------------------------
            # 他システムとメニューが違う-------------------------------------------------------
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 法人税メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\HoujinzeiMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    HQ = ImgCheck(TFolURL, r"\HoujinOpenQ.png", 0.9, 10)
                    if HQ[0] is True:
                        ImgClick(TFolURL, r"\HoujinOpenQCansel.png", 0.9, 10)
                    # 顧問先情報変更ダイアログが表示されたら
                    CDQ = ImgCheck(
                        TFolURL,
                        r"ChangeDataQ.png",
                        0.9,
                        10,
                    )
                    if CDQ[0] is True:
                        pg.press("y")  # yで決定
                        # 顧問先情報取込メニューが表示されるまで待機--------------------------
                        while (
                            ImgCheckForList(
                                TFolURL,
                                [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        CDB = ImgCheckForList(
                            TFolURL,
                            [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                            0.9,
                            10,
                        )
                        ImgClick(TFolURL, CDB[1], 0.9, 10)  # 顧問先情報取込ボタンをクリック
                # --------------------------------------------------------------------
                ImgClick(
                    TFolURL, r"\IkkatsuHoujinKousin.png", 0.9, 10
                )  # 一括更新のアイコンをクリック
                # 法人税メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\HoujinzeiMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                # 再計算ウィンドウが表示されていないかチェック-----------------------------
                DRC = ImgCheck(TFolURL, r"\DataReCalc.png", 0.9, 10)
                if DRC[0] is True:
                    pg.press("return")
                    ErrStr = "ReCalc"  # Rpaエラー判別変数
                # ------------------------------------------------------------------
                if ErrStr == "":
                    ListCheck = False  # 申告種類判定変数
                    # Excelの値から申告種類を判定------------------------------------
                    try:
                        ExPar = str(ExRow["法人税申告書_繰越対象"])
                        if ExPar == "1":
                            TaxPngName = r"\KakuteiNext.png"
                        elif ExPar == "2":
                            TaxPngName = r"\CyuukanNext.png"
                        elif ExPar == "3":
                            TaxPngName = r"\YoteiNext.png"
                        elif ExPar == "4":
                            TaxPngName = r"\SyuuseiThis.png"
                        else:
                            TaxPngName = "Err"
                    except:
                        TaxPngName = "Err"
                    time.sleep(1)
                    if TaxPngName == "Err":
                        # 確認ウィンドウが表示されるまで待機-------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\HoujinEndQ.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                            # --------------------------------------------------------------------
                            ME = ImgCheckForList(
                                TFolURL, [r"\MenuEnd.png", r"\MenuEnd2.png"], 0.9, 10
                            )
                            if ME[0] is True:
                                ImgClick(TFolURL, ME[1], 0.9, 10)  # 終了アイコンをクリック
                        # --------------------------------------------------------------------
                        pg.press("y")
                        # 法人税メニューが表示されるまで待機------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\HoujinzeiMenu.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        # 閉じる処理--------------------------
                        pg.keyDown("alt")
                        pg.press("f4")
                        pg.keyUp("alt")
                        # -----------------------------------
                        # 法人税フラグが表示されるまで待機-------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\HoujinFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        # 初期画面で開封された法人税項目を閉じる----------------------------------
                        HoujinList = [
                            r"\FastMenuHoujinzei.png",
                            r"\FastMenuHoujinzei2.png",
                        ]
                        HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
                        if HLI[0] is True:
                            ImgClick(TFolURL, HLI[1], 0.9, 10)
                        # --------------------------------------------------------------------
                        print("更新完了")

                        return False, "要申告指定", "", ""
                    else:
                        # 申告種類に応じたリスト画像までループ----------------------------
                        while ListCheck is False:
                            pg.press("down")
                            TPN = ImgCheck(TFolURL, TaxPngName, 0.9, 10)
                            if TPN[0] is True:
                                ListCheck = True
                                pg.press("return")
                        # ------------------------------------------------------------
                        time.sleep(1)
                        ImgClick(
                            TFolURL, r"\HoujinStart.png", 0.9, 10
                        )  # 更新開始のアイコンをクリック
                        # 確認ウィンドウが表示されるまで待機-------------------------------------
                        while (
                            pg.locateOnScreen(TFolURL + r"\HoujinQ.png", confidence=0.9)
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        pg.press("y")  # yで決定(nがキャンセル)
                        # 処理終了ウィンドウが表示されるまで待機----------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\HoujinEnd.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        pg.press("return")  # 決定
                        # 一括更新のアイコンが表示されるまで待機----------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\HoujinzeiMenu.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                        # --------------------------------------------------------------------
                        # 閉じる処理--------------------------
                        pg.keyDown("alt")
                        pg.press("f4")
                        pg.keyUp("alt")
                        # -----------------------------------
                        # 法人税フラグが表示されるまで待機-------------------------------------
                        while (
                            pg.locateOnScreen(
                                TFolURL + r"\HoujinFlag.png", confidence=0.9
                            )
                            is None
                        ):
                            time.sleep(1)
                            HEQ2 = ImgCheck(TFolURL, r"\HoujinEndQ2.png", 0.9, 10)
                            if HEQ2[0] is True:
                                pg.press("y")
                        # --------------------------------------------------------------------
                        # 初期画面で開封された法人税項目を閉じる----------------------------------
                        HoujinList = [
                            r"\FastMenuHoujinzei.png",
                            r"\FastMenuHoujinzei2.png",
                        ]
                        HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
                        if HLI[0] is True:
                            ImgClick(TFolURL, HLI[1], 0.9, 10)
                        # --------------------------------------------------------------------
                        print("更新完了")
                        if ErrStr == "":
                            return True, ThisNo, ThisYear, ThisMonth
                elif ErrStr == "ReCalc":
                    # 閉じる処理--------------------------
                    pg.keyDown("alt")
                    pg.press("f4")
                    pg.keyUp("alt")
                    # -----------------------------------
                    # 法人税フラグが表示されるまで待機-------------------------------------
                    while (
                        pg.locateOnScreen(TFolURL + r"\HoujinFlag.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    # --------------------------------------------------------------------
                    # 初期画面で開封された法人税項目を閉じる----------------------------------
                    HoujinList = [r"\FastMenuHoujinzei.png", r"\FastMenuHoujinzei2.png"]
                    HLI = ImgCheckForList(TFolURL, HoujinList, 0.9, 10)
                    if HLI[0] is True:
                        ImgClick(TFolURL, HLI[1], 0.9, 10)
                    # --------------------------------------------------------------------
                    return False, "要データ再計算", "", ""
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "法人税起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def GenkasyoukyakuUpdate(FolURL, TFolURL, ExRow, driver):
    """
    概要: 減価償却更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        ErrStr = ""  # Rpaエラー判別変数
        # 減価償却のアイコンを探す-------------------------------------------------
        ImgList = [r"\G_Syoukyaku.png", r"\G_Syoukyaku2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 減価償却のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 減価償却のアイコンをクリック
            # 減価償却フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\G_SyoukyakuFlag.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return", "return"])
            # 入力した関与先コードを取得------------
            pg.keyDown("shift")
            pg.press(["tab", "tab", "tab"])
            pg.keyUp("shift")
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisNo = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された年度を取得-----------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisYear = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された月を取得-------------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisMonth = pyperclip.paste()
            # -----------------------------------
            time.sleep(1)
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 減価償却メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\G_SyoukyakuMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    # アップデート情報画面が出たら閉じる-------------------------------
                    GSUM = ImgCheck(TFolURL, r"\G_SyoukyakuUpMsg.png", 0.9, 10)
                    if GSUM[0] is True:
                        ImgClick(TFolURL, r"\G_SyoukyakuUpMsgCansel.png", 0.9, 10)
                    # 顧問先情報変更ダイアログが表示されたら
                    CDQ = ImgCheck(
                        TFolURL,
                        r"ChangeDataQ.png",
                        0.9,
                        10,
                    )
                    if CDQ[0] is True:
                        pg.press("y")  # yで決定
                        # 顧問先情報取込メニューが表示されるまで待機--------------------------
                        while (
                            ImgCheckForList(
                                TFolURL,
                                [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        CDB = ImgCheckForList(
                            TFolURL,
                            [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                            0.9,
                            10,
                        )
                        ImgClick(TFolURL, CDB[1], 0.9, 10)  # 顧問先情報取込ボタンをクリック
                # --------------------------------------------------------------------
                ImgClick(TFolURL, r"\G_SyoukyakuMenu.png", 0.9, 10)  # 一括更新のアイコンをクリック
                ImgClick(TFolURL, r"\M_G_Sonota.png", 0.9, 10)  # 5.その他処理アイコンをクリック
                # 一括更新メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(
                        TFolURL + r"\IkkatsuGenkaKousin.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                ImgClick(
                    TFolURL, r"\IkkatsuGenkaKousin.png", 0.9, 10
                )  # 一括更新メニューのアイコンをクリック
                # 一括更新メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(
                        TFolURL + r"\IkkatuOpenGenkaFlag.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                ImgClick(
                    TFolURL, r"\IkkatuOpenGenkaFlag.png", 0.9, 10
                )  # 一括更新メニューのアイコンをクリック
                time.sleep(1)
                ImgClick(TFolURL, r"\IkkatuGenakStart.png", 0.9, 10)  # 一括更新開始のアイコンをクリック
                # 確認ウィンドウが表示されるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SakuseiGenkaQ.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                pg.press("y")  # yで決定(nがキャンセル)
                # 処理終了ウィンドウが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SakuseiGenkaEnd.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    Noren = ImgCheck(TFolURL, r"\No_Rendou.png", 0.9, 10)
                    if Noren[0] is True:
                        ErrStr = "Noren"
                        pg.press("y")  # yで決定(nがキャンセル)
                # --------------------------------------------------------------------
                pg.press("return")  # 決定
                # 一括更新のアイコンが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\G_SyoukyakuMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                # 閉じる処理--------------------------
                pg.keyDown("alt")
                pg.press("f4")
                pg.keyUp("alt")
                # -----------------------------------
                # 減価償却フラグが表示されるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\G_SyoukyakuFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    al4c = ImgCheck(TFolURL, r"\altf4Q.png", 0.9, 10)  # 終了確認が表示されたら
                    if al4c[0] is True:
                        pg.press("y")  # yで決定(nがキャンセル)
                # --------------------------------------------------------------------
                print("更新完了")
                if ErrStr == "":
                    return True, ThisNo, ThisYear, ThisMonth
                elif ErrStr == "Noren":
                    return False, ErrStr, "", ""
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "減価償却起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def KessanUpDate(FolURL, TFolURL, ExRow, driver):
    """
    概要: 決算内訳書更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        ErrStr = ""  # Rpaエラー判別変数
        # 決算内訳書のアイコンを探す-------------------------------------------------
        ImgList = [r"\K_Uchiwake.png", r"\K_Uchiwake2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 決算内訳書のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 決算内訳書のアイコンをクリック
            # 決算内訳書フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\Kessan_CFlag.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return", "return"])
            # 入力した関与先コードを取得------------
            pg.keyDown("shift")
            pg.press(["tab", "tab", "tab"])
            pg.keyUp("shift")
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisNo = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された年度を取得-----------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisYear = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された月を取得-------------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisMonth = pyperclip.paste()
            # -----------------------------------
            time.sleep(1)
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 決算内訳書メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\KessanMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    # 顧問先情報変更ダイアログが表示されたら
                    CDQ = ImgCheck(
                        TFolURL,
                        r"ChangeDataQ.png",
                        0.9,
                        10,
                    )
                    if CDQ[0] is True:
                        pg.press("y")  # yで決定
                        # 顧問先情報取込メニューが表示されるまで待機--------------------------
                        while (
                            ImgCheckForList(
                                TFolURL,
                                [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        CDB = ImgCheckForList(
                            TFolURL,
                            [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                            0.9,
                            10,
                        )
                        ImgClick(TFolURL, CDB[1], 0.9, 10)  # 顧問先情報取込ボタンをクリック
                # --------------------------------------------------------------------
                # 一括更新のアイコンが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(
                        TFolURL + r"\IkkatsuKessanKousin.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                ImgClick(
                    TFolURL, r"\IkkatsuKessanKousin.png", 0.9, 10
                )  # 一括更新のアイコンをクリック
                # 一括更新メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(
                        TFolURL + r"\IkkatuKessanOpenFlag.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                ImgClick(
                    TFolURL, r"\IkkatuKessanOpenFlag.png", 0.9, 10
                )  # 一括更新メニューのアイコンをクリック
                time.sleep(1)
                ImgClick(
                    TFolURL, r"\IkkatuKessanStart.png", 0.9, 10
                )  # 一括更新開始のアイコンをクリック
                # 確認ウィンドウが表示されるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SakuseiKessanQ.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                pg.press("y")  # yで決定(nがキャンセル)
                # 処理終了ウィンドウが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(
                        TFolURL + r"\IkkatuKessanEndFlag.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                    Noren = ImgCheckForList(
                        TFolURL, [r"No_Rendou.png", r"No_Rendou2.png"], 0.9, 10
                    )
                    if Noren[0] is True:
                        # ErrStr = "Noren"
                        break
                        # ImgClick(TFolURL, r"\No_Rendou_Cansel.png", 0.9, 10)
                # --------------------------------------------------------------------
                pg.press("return")  # 決定
                # 一括更新のアイコンが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(
                        TFolURL + r"\IkkatsuKessanKousin.png", confidence=0.9
                    )
                    is None
                ):
                    time.sleep(1)
                    ME = ImgCheckForList(
                        TFolURL, [r"\MenuEnd.png", r"\MenuEnd2.png"], 0.9, 10
                    )
                    if ME[0] is True:
                        pg.press("return")  # 決定
                        # ImgClick(TFolURL, ME[1], 0.9, 10)  # 終了アイコンをクリック
                # --------------------------------------------------------------------
                # 閉じる処理--------------------------
                pg.keyDown("alt")
                pg.press("f4")
                pg.keyUp("alt")
                # -----------------------------------
                # 決算内訳書フラグが表示されるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\Kessan_CFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    al4c = ImgCheck(TFolURL, r"\altf4Q.png", 0.9, 10)  # 終了確認が表示されたら
                    if al4c[0] is True:
                        pg.press("y")  # yで決定(nがキャンセル)
                # --------------------------------------------------------------------
                print("更新完了")
                if ErrStr == "":
                    return True, ThisNo, ThisYear, ThisMonth
                elif ErrStr == "Noren":
                    return False, ErrStr, "", ""
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "決算内訳書起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def KaikeiUpDate(FolURL, TFolURL, ExRow, driver):
    """
    概要: 会計大将更新処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param ExRow : Excel抽出行(obj)
    @param driver : 画面操作ドライバー(obj)
    @return : bool,ミロク入力関与先コード, ミロク入力処理年, ミロク入力処理月
    """
    try:
        ErrStr = ""  # Rpaエラー判別変数
        # 会計大将のアイコンを探す-------------------------------------------------
        ImgList = [r"\K_TaisyouIcon.png", r"\K_TaisyouIcon2.png"]
        ICFL = ImgCheckForList(TFolURL, ImgList, 0.9, 10)
        # -----------------------------------------------------------------------
        if ICFL[0] is True:  # 会計大将のアイコンがあれば
            ImgClick(TFolURL, ICFL[1], 0.9, 10)  # 会計大将のアイコンをクリック
            # 会計大将フラグが表示されるまで待機------------------------------------
            while (
                pg.locateOnScreen(TFolURL + r"\Kaikei_CFlag.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            # ------------------------------------------------------------------
            time.sleep(1)
            # 年度を最新に指定----------------------------------------------------
            IC = ImgCheck(TFolURL, r"\Nendo_Saisin.png", 0.9, 10)
            if IC[0] is False:
                IC2 = ImgCheck(TFolURL, r"\Nendo_All.png", 0.9, 10)
                if IC2[0] is False:
                    print("年度選択がない")
                else:
                    ImgClick(TFolURL, r"\Nendo_All.png", 0.9, 10)
                    pg.press("home")
                    time.sleep(1)
                    pg.press("down")
                    time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
            # ------------------------------------------------------------------
            # 関与先コード入力ボックスをクリック------------------------------------
            ImgClick(TFolURL, r"\K_NoBox.png", 0.9, 10)
            while (
                pg.locateOnScreen(TFolURL + r"\K_AfterNoBox.png", confidence=0.9)
                is None
            ):
                time.sleep(1)
            pg.write(str(ExRow["関与先番号"]))
            pg.press(["return", "return", "return"])
            # 入力した関与先コードを取得------------
            pg.keyDown("shift")
            pg.press(["tab", "tab", "tab"])
            pg.keyUp("shift")
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisNo = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された年度を取得-----------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisYear = pyperclip.paste()
            # -----------------------------------
            pg.press("return")
            # 表示された月を取得-------------------
            if windll.user32.OpenClipboard(None):
                windll.user32.EmptyClipboard()
                windll.user32.CloseClipboard()
            pg.hotkey("ctrl", "c")
            ThisMonth = pyperclip.paste()
            # -----------------------------------
            time.sleep(1)
            if str(ExRow["関与先番号"]) == ThisNo:
                print("関与先あり")
                pg.press(["return", "return", "return"])
                # 会計大将メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\K_TaisyouMenu.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    # 顧問先情報変更ダイアログが表示されたら
                    CDQ = ImgCheck(
                        TFolURL,
                        r"ChangeDataQ.png",
                        0.9,
                        10,
                    )
                    if CDQ[0] is True:
                        pg.press("y")  # yで決定
                        # 顧問先情報取込メニューが表示されるまで待機--------------------------
                        while (
                            ImgCheckForList(
                                TFolURL,
                                [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                                0.9,
                                10,
                            )[0]
                            is False
                        ):
                            time.sleep(1)
                        CDB = ImgCheckForList(
                            TFolURL,
                            [r"ChangeDataBtn.png", r"ChangeDataBtn2.png"],
                            0.9,
                            10,
                        )
                        ImgClick(TFolURL, CDB[1], 0.9, 10)  # 顧問先情報取込ボタンをクリック
                # --------------------------------------------------------------------
                ImgClick(TFolURL, r"\PrintOutTab.png", 0.9, 10)  # 2印刷タブクリック
                # 6月次締めが表示されるまで待機--------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SimeIcon.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                ImgClick(TFolURL, r"\SimeIcon.png", 0.9, 10)  # その他メニュ-のアイコンをクリック
                while (
                    pg.locateOnScreen(TFolURL + r"\GetujiIcon.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                ImgClick(TFolURL, r"\GetujiIcon.png", 0.9, 10)  # 月次処理アイコンをクリック
                # 月次確定済みか判定して処理分け-------------------------------------------------------
                # 月次処理解除-----------------------------------------------------------------------
                KUL = ImgCheck(TFolURL, r"\KakuteUnLock.png", 0.9, 10)  # 月次処理解除アイコンを検索
                if KUL[0] is True:
                    ImgClick(TFolURL, r"\KakuteUnLock.png", 0.9, 10)  # 月次処理アイコンをクリック
                    time.sleep(2)
                    pg.press("y")
                    while (
                        pg.locateOnScreen(TFolURL + r"\GetusjiKakutei.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    pg.press("return")
                    time.sleep(1)
                # ---------------------------------------------------------------------------------
                time.sleep(1)
                pg.keyDown("alt")
                pg.press("x")
                pg.keyUp("alt")
                while (
                    pg.locateOnScreen(TFolURL + r"\M_Sonota.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                ImgClick(TFolURL, r"\D_TourokuTAB.png", 0.9, 10)  # 6導入・登録タブクリック
                # 会社基本情報アイコンが表示されるまで待機--------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\CamIcon.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)      
                ImgClick(TFolURL, r"\CamIcon.png", 0.9, 10)  # 会社基本情報アイコンをクリック
                # 顧問先情報取り込みアイコンが表示されるまで待機--------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\DataInIcon.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)         
                ImgClick(TFolURL, r"\DataInIcon.png", 0.9, 10)  # 顧問先情報取り込みアイコンをクリック
                while (
                    pg.locateOnScreen(TFolURL + r"\DataInOK.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                ImgClick(TFolURL, r"\DataInOK.png", 0.9, 10)  #取り込むボタンをクリック
                time.sleep(1)
                pg.keyDown("alt")
                pg.press("x")
                pg.keyUp("alt")
                time.sleep(1)
                # --------------------------------------------------------------------------------
                ImgClick(TFolURL, r"\PrintOutTab.png", 0.9, 10)  # 2印刷タブクリック
                # マスター更新------------------------------------------------------------------------
                ImgClick(TFolURL, r"\MasterUp.png", 0.9, 10)  # マスター更新をクリック
                while (
                    pg.locateOnScreen(TFolURL + r"\MasterUpStart.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    ImgClick(TFolURL, r"\MasterUpStart.png", 0.9, 10)  # マスター更新開始をクリック
                time.sleep(1)
                pg.press("y")
                time.sleep(1)
                pg.press("return")
                time.sleep(1)
                pg.keyDown("alt")
                pg.press("x")
                pg.keyUp("alt")          
                # 月次確定--------------------------------------------------------------------------
                # 月次確定アイコンが表示されるまで待機--------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\AfterGetujiKakutei.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                ImgClick(TFolURL, r"\AfterGetujiKakutei.png", 0.9, 10)  # 月次確定アイコンクリック
                KL = ImgCheck(TFolURL, r"\KakuteiLock.png", 0.9, 10)  # 月次処理確定アイコンを検索
                if KL[0] is True:
                    while (
                        pg.locateOnScreen(TFolURL + r"\KakuteiKaijyo.png", confidence=0.99999)
                        is not None
                    ):
                        ImgClick(TFolURL, r"\KakuteiKaijyo.png", 0.99999, 10)  # 月次未確定チェックボックスをクリック
                        time.sleep(1)
                    ImgClick(TFolURL, r"\KakuteiLock.png", 0.9, 10)  # 月次処理確定アイコンをクリック
                    time.sleep(1)
                    pg.press("y")
                    while (
                        pg.locateOnScreen(TFolURL + r"\KessanKQ.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    pg.press("y")
                    while (
                        pg.locateOnScreen(TFolURL + r"\GetusjiKakutei.png", confidence=0.9)
                        is None
                    ):
                        time.sleep(1)
                    pg.press("return")
                time.sleep(1)
                pg.keyDown("alt")
                pg.press("x")
                pg.keyUp("alt")
                ImgClick(TFolURL, r"\M_Sonota.png", 0.9, 10)  # その他メニュ-のアイコンをクリック
                # 一括更新のアイコンが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\IkkatsuKousin.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                ImgClick(TFolURL, r"\IkkatsuKousin.png", 0.9, 10)  # その他メニュ-のアイコンをクリック
                # 一括更新メニューが表示されるまで待機------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\IkkatuOpenFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)                
                    # UW = ImgCheck(TFolURL, r"\Underwindow.png", 0.9, 10)
                    # if UW[0] is True:
                    #     ImgClick(TFolURL, r"\Underwindow.png", 0.9, 10)
                    #     UW2 = ImgCheck(TFolURL, r"\Underwindow2.png", 0.9, 10)
                    #     if UW2[0] is True:
                    #         ImgClick(TFolURL, r"\Underwindow2.png", 0.9, 10)
                # --------------------------------------------------------------------
                ImgClick(TFolURL, r"\IkkatuOpenFlag.png", 0.9, 10)  # 一括更新メニューのアイコンをクリック
                while ImgCheckForList(TFolURL,[r"IkkatuFind.png",r"IkkatuFind2.png"],0.9,10)[0] is False:
                    time.sleep(1)     
                FC = ImgCheckForList(
                    TFolURL,
                    [
                        r"IkkatuFind.png",
                        r"IkkatuFind2.png",
                    ],
                    0.9,
                    10,
                )
                if FC[0] is True:
                    ImgClick(TFolURL, FC[1], 0.9, 10)  # 一括更新メニューのアイコンをクリック
                    pyperclip.copy(str(ExRow["関与先番号"]))
                    pg.hotkey("ctrl", "v")
                    # 検索ボタンまでエンター-------------------------------------
                    while ImgCheck(TFolURL, r"FindFlag.png", 0.9, 10)[0] is False:
                        time.sleep(1)
                        pg.press("return")
                pg.press("return")
                time.sleep(1)
                pg.press("space")
                # チェックマークが表示されるまで待機-------------------------------------
                while (
                    ImgCheckForList(
                        TFolURL,
                        [
                            r"IkkatuCheck.png",
                            r"ZaisanCheck.png",
                            r"NendCheck.png",
                            r"HouteiCheck.png",
                        ],
                        0.9,
                        10,
                    )[0]
                    is False
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                time.sleep(1)
                ImgClick(TFolURL, r"\IkkatuStart.png", 0.9, 10)  # 一括更新開始のアイコンをクリック
                # 確認ウィンドウが表示されるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SakuseiQ.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                pg.press("y")  # yで決定(nがキャンセル)
                # 確認ウィンドウが表示されるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\SakuseiQ2.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                pg.press("return")  # 決定
                # 処理終了ウィンドウが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\IkkatuEndFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                # --------------------------------------------------------------------
                pg.press("return")  # 決定
                # チェックマークが表示されなくなるまで待機-------------------------------
                while (
                    ImgCheckForList(
                        TFolURL,
                        [
                            r"IkkatuCheck.png",
                            r"ZaisanCheck.png",
                            r"NendCheck.png",
                            r"HouteiCheck.png",
                        ],
                        0.9,
                        10,
                    )[0]
                    is True
                ):
                    time.sleep(1)
                # 一括更新のアイコンが表示されるまで待機----------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\IkkatsuKousin.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    # --------------------------------------------------------------------
                    ME = ImgCheckForList(
                        TFolURL, [r"\MenuEnd.png", r"\MenuEnd2.png"], 0.9, 10
                    )
                    if ME[0] is True:
                        ImgClick(TFolURL, ME[1], 0.9, 10)  # 終了アイコンをクリック
                # --------------------------------------------------------------------
                # 閉じる処理--------------------------
                pg.keyDown("alt")
                pg.press("f4")
                pg.keyUp("alt")
                # -----------------------------------
                # 会計大将フラグが表示されるまで待機-------------------------------------
                while (
                    pg.locateOnScreen(TFolURL + r"\Kaikei_CFlag.png", confidence=0.9)
                    is None
                ):
                    time.sleep(1)
                    al4c = ImgCheck(TFolURL, r"\altf4Q.png", 0.9, 10)  # 終了確認が表示されたら
                    if al4c[0] is True:
                        pg.press("y")  # yで決定(nがキャンセル)
                # --------------------------------------------------------------------
                print("更新完了")
                return True, ThisNo, ThisYear, ThisMonth
            else:
                print("関与先なし")
                return False, "関与先なし", "", ""
        else:
            return False, "会計大将起動失敗", "", ""
    except:
        return False, "exceptエラー", "", ""


# ------------------------------------------------------------------------------------------------------------------
def MainStarter(FolURL, TFolURL, NameDF, ExSheet, ExrcHeader, isnItem, driver):
    try:
        print(ExSheet[3 : len(ExSheet)])
        li = np.array(ExSheet[3 : len(ExSheet)])
        ExDf = pd.DataFrame(li, columns=ExrcHeader)
        print(ExDf)
        Exrc = np.array(ExDf).shape[0]  # 行数
        for Ex in range(1, Exrc):
            ExRow = ExDf.iloc[Ex]
            if ExRow["関与先番号"] == ExRow["関与先番号"]:  # nan判定
                # nanでない場合
                OpenSystem(
                    FolURL, TFolURL, NameDF, ExRow, Ex, ExrcHeader, isnItem, driver
                )
                print("")
            else:
                # nanの場合
                print("nan")
        return True, ""
    except:
        return False, ""


# ------------------------------------------------------------------------------------------------------------------
def OpenSystem(FolURL, TFolURL, NameDF, ExRow, Ex, ExrcHeader, isnItem, driver):
    try:
        Eh = 0
        for ExrcHeaderItem in ExrcHeader:
            if Eh < (len(ExrcHeader) - 1):
                if "_繰越対象" in ExrcHeaderItem:
                    SysN = ExrcHeaderItem.split("_")
                    Title = str(SysN[0])
                    if (
                        not ExRow[Title + "_繰越対象"] == "-"
                        and str(ExRow[Title + "_繰越対象"]) == "1"
                    ):
                        if "::" not in Title:
                            if ExRow[Title + "_繰越対象"] == ExRow[Title + "_繰越対象"]:
                                # nanでない場合
                                if ExRow[Title + "_繰越処理日"] == ExRow[Title + "_繰越処理日"]:
                                    # nanでない場合
                                    print(ExRow[Title + "_繰越処理日"])
                                else:
                                    # nanの場合
                                    # Log--------------------------------------------
                                    dt_s = datetime.datetime.now()
                                    dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                                    Rno = ExRow["関与先番号"]
                                    Rn = NameSearch(NameDF, Rno)
                                    Rn = Rn.replace("\u3000", "")
                                    KomonUpdate(TFolURL,ExRow) # 顧問先情報更新
                                    logger.debug(
                                        dt_s
                                        + "_関与先番号:"
                                        + str(Rno)
                                        + ":"
                                        + str(Rn)
                                        + "_メイン処理開始"
                                    )
                                    with open(LURL, "a") as f:
                                        print(
                                            [
                                                dt_s,
                                                "関与先番号:" + str(Rno),
                                                str(Rn),
                                                "メイン処理開始",
                                            ],
                                            file=f,
                                        )
                                    # -----------------------------------------------
                                    ChildFlow(
                                        FolURL,
                                        TFolURL,
                                        ExRow,
                                        Ex,
                                        Eh,
                                        ExrcHeader,
                                        isnItem,
                                        Title,
                                        driver,
                                        Rno,
                                        Rn,
                                    )
                            else:
                                # nanでない場合
                                if ExRow[Title + "_繰越処理日"] == ExRow[Title + "_繰越処理日"]:
                                    # nanでない場合
                                    print("スタート")
                                else:
                                    # nanの場合
                                    # Log--------------------------------------------
                                    dt_s = datetime.datetime.now()
                                    dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                                    Rno = ExRow["関与先番号"]
                                    Rn = NameSearch(NameDF, Rno)
                                    Rn = Rn.replace("\u3000", "")
                                    KomonUpdate(TFolURL,ExRow) # 顧問先情報更新
                                    logger.debug(
                                        dt_s
                                        + "_関与先番号:"
                                        + str(Rno)
                                        + ":"
                                        + str(Rn)
                                        + "_メイン処理開始"
                                    )
                                    with open(LURL, "a") as f:
                                        print(
                                            [
                                                dt_s,
                                                "関与先番号:" + str(Rno),
                                                str(Rn),
                                                "メイン処理開始",
                                            ],
                                            file=f,
                                        )
                                    # -----------------------------------------------
                                    ChildFlow(
                                        FolURL,
                                        TFolURL,
                                        ExRow,
                                        Ex,
                                        Eh,
                                        ExrcHeader,
                                        isnItem,
                                        Title,
                                        driver,
                                        Rno,
                                        Rn,
                                    )
                Eh += 1
        return True
    except:
        print("TEST")
        return False


# ------------------------------------------------------------------------------------------------------------------
def MainFlow(FolURL, TFolURL, Exlsx, driver, XLSURL):
    """
    概要: メイン処理
    @param FolURL : ミロク起動関数のフォルダ(str)
    @param TFolURL : このpyファイルのフォルダ(str)
    @param Exlsx : Excel指示シート(obj)
    @return : bool
    """
    try:
        # ----------------------------------------------------------------------------------------------------------------------
        logger.debug("xlsxをDataFrameに")
        # sheet_namesメソッドでExcelブック内の各シートの名前をリストで取得できる
        input_sheet_name = Exlsx.sheet_names
        # lenでシートの総数を確認
        num_sheet = len(input_sheet_name)
        # シートの数とシートの名前のリストの表示
        print("Sheet の数:", num_sheet)
        print(input_sheet_name)
        for isnItem in input_sheet_name:
            Exlsx = EFA.XlsmRead(XLSURL)[1]
            # DataFrameとしてsheetのデータ読込み
            if "更新申請" in isnItem:
                ExSheet = ""
                NameSheet = ""
                ExSheet = Exlsx.parse(isnItem, skiprows=0)
                NameSheet = Exlsx.parse("関与先一覧")
                print(ExSheet)
                # 初回読込時の保存--------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H-%M-%S")
                DF = pd.DataFrame(ExSheet)
                NameDF = pd.DataFrame(NameSheet)
                DF.to_csv(
                    r"\\NAS-SV\B_監査etc\B2_電子ﾌｧｲﾙ\RPA_ミロクシステム次年更新\一括更新申請\MJSLog\初回起動_"
                    + dt_s
                    + ".csv",
                    encoding="cp932",
                    index=False,
                )
                # ----------------------------------------
                Exrc = np.array(ExSheet).shape[1]  # 列数
                ExrcHeader = []
                for Ex in range(Exrc):
                    ExRow = ExSheet.iloc[0]
                    ExSecondRow = ExSheet.iloc[1]
                    if ExRow[Ex] == ExRow[Ex]:  # nan判定
                        # nanでない場合
                        Txt = ExRow[Ex]
                        if ExSecondRow[Ex] == ExSecondRow[Ex]:  # nan判定
                            # nanでない場合
                            ExrcHeader.append(ExRow[Ex] + "_" + ExSecondRow[Ex])
                        else:
                            # nanの場合
                            ExrcHeader.append(ExRow[Ex])
                    else:
                        # nanの場合
                        if ExSecondRow[Ex] == ExSecondRow[Ex]:  # nan判定
                            # nanでない場合
                            ExrcHeader.append(Txt + "_" + ExSecondRow[Ex])
                        else:
                            # nanの場合
                            ExrcHeader.append(Txt)
                # Log--------------------------------------------
                dt_s = datetime.datetime.now()
                dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
                logger.debug(dt_s + "_Excelシート読込完了")
                with open(LURL, "a") as f:
                    print([dt_s, "Excelシート読込完了"], file=f)
                # -----------------------------------------------
                MainStarter(
                    FolURL,
                    TFolURL,
                    NameDF,
                    ExSheet,
                    ExrcHeader,
                    isnItem,
                    driver,
                )  # データ送信画面までの関数
                print("")
    except Exception as e:
        logger.debug(e)


def Main(FolURL, TFolURL, XLSDir, LURL):
    try:
        # Log--------------------------------------------
        dt_s = datetime.datetime.now()
        dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
        logger.debug(dt_s + "_MJSシステム更新開始")
        # -----------------------------------------------
        BatUrl = FolURL + "/bat/AWADriverOpen.bat"  # 4724ポート指定でappiumサーバー起動バッチを開く
        driver = MJSOpen.MainFlow(
            BatUrl, FolURL, "RPAPhoto/MJS_DensiSinkoku"
        )  # MJSを起動しログイン後インスタンス化
        # driver = []
        FolURL = FolURL + "/RPAPhoto/MJS_DensiSinkoku"
        for fd_path, sb_folder, sb_file in os.walk(XLSDir):
            FDP = fd_path
            if not len(sb_folder) == 0:
                for sb_fileItem in sb_file:
                    print(sb_fileItem)
                    if (
                        "一括更新申請ミロク" in sb_fileItem
                        and not "一括更新申請ミロク(原本).xlsm" == sb_fileItem
                    ):
                        XLSURL = (
                            FDP + r"\\" + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        MoveXLSURL = (
                            FDP
                            + r"\\MJSLog\\"
                            + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        os.rename(XLSURL, MoveXLSURL)
                        MoveXLSURL = (
                            FDP + r"\\" + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        XLSURL = (
                            FDP
                            + r"\\MJSLog\\"
                            + sb_fileItem.replace("~", "").replace("$", "")
                        )
                        open(LURL, "w").close()
                        Exlsx = EFA.XlsmRead(XLSURL)
                        if Exlsx[0] is True:
                            try:
                                MainFlow(FolURL, TFolURL, Exlsx[1], driver, XLSURL)
                            except:
                                traceback.print_exc()
                            Exlsx = "閉じろや"
                            os.rename(XLSURL, MoveXLSURL)
                        else:
                            print("Excel読み込みエラー")
                            logger.debug("Excel読み込みエラー")
        return True
    except:
        return False


if __name__ == "__main__":
    # RPA用画像フォルダの作成---------------------------------------------------------
    FolURL = os.getcwd().replace("\\", "/")  # 先
    TFolURL = FolURL + r"\RPAPhoto\MJS_SystemNextCreate"  # 先
    XLSDir = r"\\NAS-SV\B_監査etc\B2_電子ﾌｧｲﾙ\RPA_ミロクシステム次年更新\一括更新申請"
    LURL = r"\\NAS-SV\B_監査etc\B2_電子ﾌｧｲﾙ\RPA_ミロクシステム次年更新\一括更新申請\MJSLog\MJSSysUpLog.txt"  # 処理状況CSVのURL
    # --------------------------------------------------------------------------------
    # Log--------------------------------------------
    dt_s = datetime.datetime.now()
    dt_s = dt_s.strftime("%Y-%m-%d %H:%M:%S")
    logger.debug(dt_s + "_MJSシステム更新開始")
    # -----------------------------------------------
    BatUrl = FolURL + "/bat/AWADriverOpen.bat"  # 4724ポート指定でappiumサーバー起動バッチを開く
    driver = MJSOpen.MainFlow(
        BatUrl, FolURL, "RPAPhoto/MJS_DensiSinkoku"
    )  # MJSを起動しログイン後インスタンス化
    # driver = []
    FolURL = FolURL + "/RPAPhoto/MJS_DensiSinkoku"
    for fd_path, sb_folder, sb_file in os.walk(XLSDir):
        FDP = fd_path
        if not len(sb_folder) == 0:
            for sb_fileItem in sb_file:
                print(sb_fileItem)
                if (
                    "一括更新申請ミロク" in sb_fileItem
                    and not "一括更新申請ミロク(原本).xlsm" == sb_fileItem
                ):
                    XLSURL = FDP + r"\\" + sb_fileItem.replace("~", "").replace("$", "")
                    MoveXLSURL = (
                        FDP
                        + r"\\MJSLog\\"
                        + sb_fileItem.replace("~", "").replace("$", "")
                    )
                    os.rename(XLSURL, MoveXLSURL)
                    MoveXLSURL = (
                        FDP + r"\\" + sb_fileItem.replace("~", "").replace("$", "")
                    )
                    XLSURL = (
                        FDP
                        + r"\\MJSLog\\"
                        + sb_fileItem.replace("~", "").replace("$", "")
                    )
                    open(LURL, "w").close()
                    Exlsx = EFA.XlsmRead(XLSURL)
                    if Exlsx[0] is True:
                        try:
                            MainFlow(FolURL, TFolURL, Exlsx[1], driver, XLSURL)
                        except:
                            traceback.print_exc()
                        Exlsx = "閉じろや"
                        os.rename(XLSURL, MoveXLSURL)
                    else:
                        print("Excel読み込みエラー")
                        logger.debug("Excel読み込みエラー")
